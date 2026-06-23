"""文件解析工具模块

目前包括的功能有：
- xlsx文件解析和转换为jsonl
"""
import io
from typing import List

from fastapi import HTTPException

"""
## 转换示例

### 示例1：只有prompt和response列

**XLSX文件内容：**
| prompt | response |
|--------|----------|
| 你好   | 你好！    |
| 今天天气怎么样？ | 今天天气很好！ |

**转换后的JSONL：**
```jsonl
{"prompt": "你好", "response": "你好！"}
{"prompt": "今天天气怎么样？", "response": "今天天气很好！"}
```

### 示例2：包含system列

**XLSX文件内容：**
| prompt | response | system |
|--------|----------|--------|
| 你好   | 你好！    | 你是一个专业的助手 |
| 今天天气怎么样？ | 今天天气很好！ | 你是一个专业的助手 |

**转换后的JSONL：**
```jsonl
{"system": "你是一个专业的助手", "prompt": "你好", "response": "你好！"}
{"system": "你是一个专业的助手", "prompt": "今天天气怎么样？", "response": "今天天气很好！"}
```

### 示例3：JSON数组转换为JSONL

**JSON文件内容：**
```json
[
  {
    "prompt": "你好",
    "response": "你好！",
    "system": "你是一个专业的助手"
  },
  {
    "prompt": "今天天气怎么样？",
    "response": "今天天气很好！",
    "system": "你是一个专业的助手"
  }
]
```

**转换后的JSONL：**
```jsonl
[{"system": "你是一个专业的助手", "prompt": "你好", "response": "你好！"}]
[{"system": "你是一个专业的助手", "prompt": "今天天气怎么样？", "response": "今天天气很好！"}]
```

### 示例4：单个JSON对象转换为JSONL

**JSON文件内容：**
```json
{
  "prompt": "你好",
  "response": "你好！",
  "system": "你是一个专业的助手"
}
```

**转换后的JSONL：**
```jsonl
{"system": "你是一个专业的助手", "prompt": "你好", "response": "你好！"}
```
"""
import json
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

async def convert_to_jsonl(file_content: bytes, file_type: str) -> bytes:
    """
    jsonl转化器
    将不同类型的数据集文件转化为标准的jsonl文件

    Args:
        file_content: 原始文件字节内容
        file_type: 原始文件文件格式

    Returns:
        转换后的jsonl格式字节内容
    """
    match file_type:
        case 'xlsx':
            return await convert_xlsx_to_jsonl(file_content)
        case 'json':
            return await convert_json_to_jsonl(file_content)
        case _:
            raise HTTPException(status_code=400, detail=f"当前数据集格式暂不支持：{file_type}")

async def convert_xlsx_to_jsonl(file_content: bytes) -> bytes:
    """
    将xlsx文件转换为符合规范的jsonl格式

    Args:
        file_content: xlsx文件的字节内容

    Returns:
        转换后的jsonl格式字节内容
    """
    try:
        # 1. 使用openpyxl读取xlsx文件
        # 将bytes转化为io
        file_stream = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(filename=file_stream, data_only=True)
        worksheet = workbook.active

        if worksheet is None:
            raise ValueError("工作簿中没有找到活动工作表")

        if worksheet.max_row < 2:
            raise ValueError("Excel文件至少需要包含标题行和一行数据")

        # 2. 验证列结构 - 支持多种列名
        ## 提取当前sheet页所有的列名
        headers = []
        for col_num in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col_num)
            header_value = str(header_cell.value) if header_cell.value else f"列{col_num}"
            headers.append(header_value)

        # 定义可能的列名映射（不区分大小写），目前列名必须固定
        prompt_aliases = ['prompt']
        response_aliases = ['response']
        system_aliases = ['system']

        # 查找各列的位置，定位需要的列的index
        prompt_col = None
        response_col = None
        system_col = None

        for i, header in enumerate(headers):
            header_lower = header.lower()
            # 定位prompt列
            if any(alias in header_lower for alias in prompt_aliases) and prompt_col is None:
                prompt_col = i + 1

            # 定位response列
            elif any(alias in header_lower for alias in response_aliases) and response_col is None:
                response_col = i + 1

            # 定位system列
            elif any(alias in header_lower for alias in system_aliases) and system_col is None:
                system_col = i + 1

        # 验证必需列（prompt与response）
        if prompt_col is None or response_col is None:
            raise ValueError("Excel文件必须包含prompt和response列")

        # 3. 转换数据
        jsonl_lines = []

        # 从第二行开始读取数据
        for row_num in range(2, worksheet.max_row + 1):
            prompt_cell = worksheet.cell(row=row_num, column=prompt_col)
            response_cell = worksheet.cell(row=row_num, column=response_col)

            # 跳过空行（prompt和response不能同时为空）
            if not prompt_cell.value or not response_cell.value:
                continue

            # 构造符合规范的json对象，封装prompt与reponse数据
            data = {
                "prompt": str(prompt_cell.value),
                "response": str(response_cell.value)
            }

            # 如果存在system列且有值，则添加system字段
            if system_col:
                system_cell = worksheet.cell(row=row_num, column=system_col)
                if system_cell.value:  # 只有当system列有值时才添加
                    data["system"] = str(system_cell.value)

            # 转换为jsonl格式（每行一个JSON对象）
            jsonl_lines.append('[' + json.dumps(data, ensure_ascii=False) + ']')

        # 4. 返回jsonl字节内容
        return '\n'.join(jsonl_lines).encode('utf-8')

    except InvalidFileException:
        raise ValueError("无效的Excel文件格式")
    except Exception as e:
        raise ValueError(f"转换Excel文件失败: {str(e)}")

async def convert_json_to_jsonl(file_content: bytes) -> bytes:
    """
    将json文件转换为符合规范的jsonl格式

    Args:
        file_content: json文件的字节内容

    Returns:
        转换后的jsonl格式字节内容
    """
    try:
        # 1. 解析JSON内容
        json_str = file_content.decode('utf-8')
        data = json.loads(json_str)

        # 2. 处理不同的JSON结构
        jsonl_lines = []

        if isinstance(data, list):
            # 如果是数组格式，遍历每个对象
            for item in data:
                if isinstance(item, dict):
                    # 验证必需字段
                    if "prompt" not in item or "response" not in item:
                        continue  # 跳过缺少必需字段的条目

                    # 确保prompt和response不为空
                    if not item["prompt"] or not item["response"]:
                        continue  # 跳过空值

                    # 转换为jsonl格式（每行一个JSON对象）
                    jsonl_lines.append('[' + json.dumps(item, ensure_ascii=False) + ']')
        elif isinstance(data, dict):
            # 如果是单个对象格式
            if "prompt" in data and "response" in data:
                # 确保prompt和response不为空
                if data["prompt"] and data["response"]:
                    # 转换为jsonl格式
                    jsonl_lines.append('[' + json.dumps(data, ensure_ascii=False) + ']')

        # 3. 返回jsonl字节内容
        return '\n'.join(jsonl_lines).encode('utf-8')

    except json.JSONDecodeError as e:
        raise ValueError(f"JSON文件格式无效: {str(e)}")
    except Exception as e:
        raise ValueError(f"转换JSON文件失败: {str(e)}")

# 上传数据集文件类型配置字典
# 这里应该与对应的 TrainingDatasetUploadTypeCategory （上传数据集文件类型枚举）同步
# 添加配置字典，为的是在文件下载时，针对不同文件，设置不同的media_type与read_mode
FILE_TYPE_CONFIG = {
    'jsonl': {
        'media_type': 'application/jsonl',
        'read_mode': 'r',
        'write_mode': 'w',
        'encoding': 'utf-8',
        'file_suffix': '.jsonl'
    },
    'xlsx': {
        'media_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'read_mode': 'rb',
        'write_mode': 'wb',
        'encoding': None,
        'file_suffix': '.xlsx'
    },
    'json': {
        'media_type': 'application/json',
        'read_mode': 'r',
        'write_mode': 'w',
        'encoding': 'utf-8',
        'file_suffix': '.json'
    }
}

def generate_filenames(base_filename: str) -> List[str]:
    """生成所有可能的文件名
    
    Args:
        base_filename: 基础文件名
        
    Returns:
        所有可能的文件名列表
    """
    # 若文件名称带有后缀，去除后缀
    if '.' in base_filename:
        base_filename = base_filename.split('.')[0]
    else:
        base_filename = base_filename

    # 根据配置生成所有支持的文件类型
    filenames = []
    for config in FILE_TYPE_CONFIG.values():
        filename = f"{base_filename}{config['file_suffix']}"
        filenames.append(filename)

    # 添加索引文件类型
    index_filename = f"{base_filename}_index.cache"
    filenames.append(index_filename)
    
    return filenames
