"""
文件解析工具
"""

import asyncio
import csv
import io
import json
import os
import pickle
import tempfile
import zipfile
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from io import BytesIO
from typing import Iterable, List, Dict, Any, Optional, Tuple

from fastapi import HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.logging import logger
from app.models import TrainingDataset
from app.schemas.training_dataset import TrainingDatasetUploadTypeCategory, DatasetUsage, DatasetFormat, \
    TrainingDatasetExportTypeCategory
from app.schemas.training_task import TrainingTypeCategory, TrainingMethodType
from app.services.storage.interface import StorageService
from app.utils.storage_enum import StoragePath
from app.utils.validators import validate_dataset_upload_file_type

# 上传数据集文件类型配置字典
# 这里应该与对应的 TrainingDatasetUploadTypeCategory （上传数据集文件类型枚举）同步
# 添加配置字典，为的是在文件下载时，针对不同文件，设置不同的media_type与read_mode
FILE_TYPE_CONFIG = {
    'jsonl': {
        'media_type': 'application/jsonl',
        'read_mode': 'r',
        'write_mode': 'w',
        'encoding': 'utf-8',
        'file_suffix': '.jsonl'
    },
    'xlsx': {
        'media_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'read_mode': 'rb',
        'write_mode': 'wb',
        'encoding': None,
        'file_suffix': '.xlsx'
    },
    'json': {
        'media_type': 'application/json',
        'read_mode': 'r',
        'write_mode': 'w',
        'encoding': 'utf-8',
        'file_suffix': '.json'
    },
    'zip': {
        'media_type': 'application/zip',
        'read_mode': 'rb',
        'write_mode': 'wb',
        'encoding': None,
        'file_suffix': '.zip'
    },
    'csv': {
        'media_type': 'text/csv',
        'read_mode': 'r',
        'write_mode': 'w',
        'encoding': 'utf-8',
        'file_suffix': '.csv'
    }
}


def write_jsonl_lines_in_batches(file_obj, jsonl_lines: List[str], batch_size: int = 500) -> None:
    """按固定行数批量写入 JSONL，避免 JuiceFS 单次写入超大 buffer。"""
    total_lines = len(jsonl_lines)
    for start in range(0, total_lines, batch_size):
        batch_lines = jsonl_lines[start:start + batch_size]
        batch_content = "\n".join(batch_lines)
        if start + batch_size < total_lines:
            batch_content += "\n"
        file_obj.write(batch_content.encode('utf-8'))


@dataclass
class ColumnMapping:
    """列映射信息"""
    prompt_col: int
    response_col: int
    system_col: Optional[int] = None


@dataclass
class LineIndex:
    """单行数据的索引信息 - 用于优化大数据集的随机访问"""
    line_number: int      # 行号（从0开始）
    file_offset: int      # 该行在文件中的字节偏移量
    line_length: int      # 该行的字节长度


@dataclass
class TextGenerationParseResult:
    """文本生成数据集文件解析结果"""
    total_rows: int # 总样本数
    processed_rows: int # 总处理行数
    skip_rows: int # 跳过行数
    total_characters: int # 总有效字符数
    convert_file_content: bytes # 转化后的文件字节


@dataclass
class ImageUnderstandingParseResult:
    """图像理解数据集文件处理结果"""
    jsonl_content: bytes  # data.jsonl 文件内容
    images: Dict[str, bytes]  # 图片字典 {图片文件名: 图片内容}
    total_samples: int  # 总样本数
    total_characters: int  # 总字符数

class TextGenerationDatasetFileParser:
    """文本生成数据集文件解析器"""
    
    def __init__(self, max_workers: int = 2, chunk_size: int = 1000):
        """
        初始化解析器

        Args:
            max_workers: 线程池最大工作线程数，默认2线程
            chunk_size: 批处理块大小，默认1000
        """
        self.executor = ThreadPoolExecutor(max_workers=max_workers)
        self.chunk_size = chunk_size

    @staticmethod
    def _map_columns(worksheet) -> ColumnMapping:
        """
        映射列信息（同步方法）
        """
        # 读取第一行（表头）
        headers = []
        for cell in next(worksheet.rows):
            header_value = str(cell.value) if cell.value else ""
            headers.append(header_value.lower())
        
        # 定义列名别名（小写）
        prompt_aliases = ['prompt']
        response_aliases = ['response']
        system_aliases = ['system']
        
        # 使用字典加速查找
        header_to_index = {header: i for i, header in enumerate(headers)}
        
        # 查找各列位置
        prompt_col = None
        response_col = None
        system_col = None
        
        for alias in prompt_aliases:
            if alias in header_to_index:
                prompt_col = header_to_index[alias] + 1  # +1因为openpyxl索引从1开始
                break
        
        for alias in response_aliases:
            if alias in header_to_index:
                response_col = header_to_index[alias] + 1
                break
        
        for alias in system_aliases:
            if alias in header_to_index:
                system_col = header_to_index[alias] + 1
                break
        
        # 验证必需列
        if prompt_col is None or response_col is None:
            logger.error("Excel文件必须包含prompt和response列")
            raise ValueError("文件格式错误")
        
        return ColumnMapping(
            prompt_col=prompt_col,
            response_col=response_col,
            system_col=system_col
        )

    @staticmethod
    def _process_xlsx_batch(batch: List, column_mapping: ColumnMapping) -> Tuple[List[str], int, int, int]:
        """
        批量处理数据行

        Tips:
        注意：根据业务需求，JSONL行格式为 [{"prompt": "...", "response": "..."}]
        这是非标准格式，但与下游系统兼容
        """
        prompt_col = column_mapping.prompt_col - 1  # 转换为0基索引
        response_col = column_mapping.response_col - 1
        system_col = column_mapping.system_col - 1 if column_mapping.system_col else None # system列为可选

        processed_lines = [] # 处理完成的样本字符列表
        processed_count = 0 # 处理行数
        skip_count = 0 # 跳过行数
        char_count = 0 # 总字符数

        for row in batch:
            # 检查行是否有效
            if len(row) <= max(col for col in [prompt_col, response_col] if col is not None):
                # 记录跳过处理的行数
                skip_count += 1
                continue
            
            prompt_cell = row[prompt_col]
            response_cell = row[response_col]
            
            # 跳过空行
            if not prompt_cell.value or not response_cell.value:
                # 记录跳过处理的行数
                skip_count += 1
                continue
            
            # 构造数据对象
            data = {
                "prompt": str(prompt_cell.value),
                "response": str(response_cell.value)
            }
            
            # 添加system字段（如果有且有值）
            if system_col is not None and system_col < len(row):
                system_cell = row[system_col]
                if system_cell.value:
                    data["system"] = str(system_cell.value)
            
            # 序列化为JSONL格式
            jsonl_line = '[' + json.dumps(data, ensure_ascii=False) + ']'
            processed_lines.append(jsonl_line)

            # 记录处理样本数
            processed_count += 1

            # 记录字符数
            char_count += len(jsonl_line.strip())

        return processed_lines, processed_count, skip_count, char_count

    @staticmethod
    def _process_json_batch(batch: List[Dict[str, Any]]) -> Tuple[List[str], int, int, int]:
        """
        批量处理JSON数据

        Args:
            batch: JSON对象批次

        Returns:
            处理后的JSONL行列表

        Tips:
        注意：根据业务需求，JSONL行格式为 [{"prompt": "...", "response": "..."}]
        这是非标准格式，但与下游系统兼容
        """
        processed_lines = []  # 处理完成的样本字符列表
        processed_count = 0  # 处理行数
        skip_count = 0  # 跳过行数
        char_count = 0  # 总字符数

        for item in batch:
            if not isinstance(item, dict):
                raise ValueError("字段/格式错误")

            # 验证必需字段
            if "prompt" not in item or "response" not in item:
                skip_count += 1
                continue

            # 验证字段不为空
            if not item["prompt"] or not item["response"]:
                skip_count += 1
                continue

            # 序列化
            jsonl_line = '[' + json.dumps(item, ensure_ascii=False) + ']'
            processed_lines.append(jsonl_line)
            processed_count += 1

            # 记录字符数
            char_count += len(jsonl_line.strip())

        return processed_lines, processed_count, skip_count, char_count

    @staticmethod
    def _process_csv_batch(batch: List[List[str]], prompt_col: int, response_col: int, system_col: Optional[int]) -> Tuple[List[str], int, int, int]:
        """
        批量处理CSV数据行

        Args:
            batch: CSV行数据批次
            prompt_col: prompt列索引
            response_col: response列索引
            system_col: system列索引（可选）

        Returns:
            处理后的JSONL行列表、处理行数、跳过行数、字符数
        """
        processed_lines = []
        processed_count = 0
        skip_count = 0
        char_count = 0

        for row in batch:
            # 检查行是否有效
            if len(row) <= max(prompt_col, response_col):
                skip_count += 1
                continue

            prompt_value = row[prompt_col].strip() if prompt_col < len(row) else ""
            response_value = row[response_col].strip() if response_col < len(row) else ""

            # 跳过空行
            if not prompt_value or not response_value:
                skip_count += 1
                continue

            # 构造数据对象
            data = {
                "prompt": prompt_value,
                "response": response_value
            }

            # 添加system字段（如果有且有值）
            if system_col is not None and system_col < len(row):
                system_value = row[system_col].strip() if row[system_col] else ""
                if system_value:
                    data["system"] = system_value

            # 序列化为JSONL格式
            jsonl_line = '[' + json.dumps(data, ensure_ascii=False) + ']'
            processed_lines.append(jsonl_line)
            processed_count += 1
            char_count += len(jsonl_line.strip())

        return processed_lines, processed_count, skip_count, char_count

    async def process_text_generation_file(
        self,
        file_content: bytes,
        file_type: str,
        dataset_format: str,
        training_method_type: str = TrainingMethodType.SFT.value,
    ) -> TextGenerationParseResult | None:
        """
        数据集文件解析

        Args:
            file_content: 原始文件字节内容
            file_type: 原始文件文件格式
            dataset_format: 数据集格式

        Returns:
            解析结果
        tips:
            在转化为jsonl串时，默认在前后加上[和]
        """
        try:
            if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
                # prompt-response
                # 若文件类型不为标准的jsonl，先转化为jsonl，并且同时保存转化后的jsonl文件与源文件（目前支持xlsx、csv、json和jsonl类型）
                if file_type == 'xlsx':
                    return await self.analyze_prompt_response_xlsx_content(file_content)
                elif file_type == 'csv':
                    return await self.analyze_prompt_response_csv_content(file_content)
                elif file_type == 'json':
                    return await self.analyze_prompt_response_json_content(file_content)
                elif file_type == 'jsonl':
                    return await self.analyze_prompt_response_jsonl_content(file_content)
                else:
                    logger.error(f"当前数据集格式暂不支持：{file_type}")
                    raise ValueError("文件格式错误")

            elif dataset_format == DatasetFormat.ALPACA.value:
                if file_type == 'json':
                    return await self.analyze_dpo_prompt_response_json_content(file_content)
                elif file_type == 'jsonl':
                    return await self.analyze_dpo_prompt_response_jsonl_content(file_content)
                elif file_type == 'xlsx':
                    return await self.analyze_dpo_prompt_response_xlsx_content(file_content)
                elif file_type == 'csv':
                    return await self.analyze_dpo_prompt_response_csv_content(file_content)
                else:
                    logger.error(f"alpaca 当前数据集格式暂不支持：{file_type}")
                    raise ValueError("文件格式错误")

            elif dataset_format == DatasetFormat.ROLE_BASED.value:
                # role-based
                # 目前支持json、jsonl、xlsx格式
                # 文件解析逻辑参考 图像理解数据集 只是不需要判断图片标签
                if training_method_type == TrainingMethodType.DPO.value:
                    if file_type == 'jsonl':
                        return await self.analyze_dpo_role_based_jsonl_content(file_content)
                    elif file_type == 'json':
                        return await self.analyze_dpo_role_based_json_content(file_content)
                    elif file_type == 'xlsx':
                        return await self.analyze_dpo_role_based_xlsx_content(file_content)
                    elif file_type == 'csv':
                        return await self.analyze_dpo_role_based_csv_content(file_content)
                    else:
                        logger.error(f"DPO role-based 当前数据集格式暂不支持：{file_type}")
                        raise ValueError("文件格式错误")
                if file_type == 'jsonl':
                    return await self.analyze_role_based_jsonl_content(file_content)
                elif file_type == 'xlsx':
                    return await self.analyze_role_based_xlsx_content(file_content)
                elif file_type == 'json':
                    return await self.analyze_role_based_json_content(file_content)
                else:
                    logger.error(f"当前数据集格式暂不支持：{file_type}")
                    raise ValueError("文件格式错误")

            elif dataset_format == DatasetFormat.GRPO.value:
                if training_method_type != TrainingMethodType.GRPO.value:
                    logger.error(f"grpo 数据集格式仅支持 grpo 训练方法，当前为: {training_method_type}")
                    raise ValueError("文件格式错误")
                if file_type == 'json':
                    return await self.analyze_grpo_json_content(file_content)
                elif file_type == 'jsonl':
                    return await self.analyze_grpo_jsonl_content(file_content)
                elif file_type == 'xlsx':
                    return await self.analyze_grpo_xlsx_content(file_content)
                logger.error(f"GRPO 数据集仅支持 json/jsonl/xlsx 格式，当前为: {file_type}")
                raise ValueError("文件格式错误")

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"文件解析服务异常: {str(e)}")
            raise ValueError("文件解析服务异常")

    @staticmethod
    def _validate_dpo_prompt_response_item(item: Dict[str, Any], line_num: int, obj_index: Optional[int] = None) -> None:
        prefix = f"第{line_num}行"
        if obj_index is not None:
            prefix += f"第{obj_index + 1}个对象"

        required_fields = ("input", "chosen", "rejected")
        for field_name in required_fields:
            if field_name not in item:
                logger.error(f"{prefix}：缺少必需字段 {field_name}")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if not isinstance(item[field_name], str):
                logger.error(f"{prefix}：{field_name} 必须是字符串")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if not item[field_name].strip():
                logger.error(f"{prefix}：{field_name} 不能为空")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        if "input" in item and item["input"] is not None and not isinstance(item["input"], str):
            logger.error(f"{prefix}：input 必须是字符串")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

    @staticmethod
    def _validate_grpo_item(item: Dict[str, Any], line_num: int) -> None:
        prefix = f"第{line_num}行"
        for field_name in ("data_source", "prompt", "reward_model"):
            if field_name not in item:
                logger.error(f"{prefix}：缺少必需字段 {field_name}")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        if not isinstance(item["data_source"], str) or not item["data_source"].strip():
            logger.error(f"{prefix}：data_source 必须是非空字符串")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        prompt = item["prompt"]
        if not isinstance(prompt, list) or not prompt:
            logger.error(f"{prefix}：prompt 必须是非空数组")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        for msg_idx, message in enumerate(prompt):
            if not isinstance(message, dict):
                logger.error(f"{prefix} prompt[{msg_idx}]：必须是对象")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            role = message.get("role")
            if role not in ("system", "user", "assistant"):
                logger.error(f"{prefix} prompt[{msg_idx}]：role 必须是 system/user/assistant")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            content = message.get("content")
            if not isinstance(content, str) or not content.strip():
                logger.error(f"{prefix} prompt[{msg_idx}]：content 必须是非空字符串")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        reward_model = item["reward_model"]
        if not isinstance(reward_model, dict):
            logger.error(f"{prefix}：reward_model 必须是对象")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        for field_name in ("style", "ground_truth"):
            if field_name not in reward_model:
                logger.error(f"{prefix}：reward_model 缺少必需字段 {field_name}")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if not isinstance(reward_model["style"], str) or not reward_model["style"].strip():
            logger.error(f"{prefix}：reward_model.style 必须是非空字符串")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if reward_model["ground_truth"] is None:
            logger.error(f"{prefix}：reward_model.ground_truth 不能为空")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if "extra_info" in item and item["extra_info"] is not None and not isinstance(item["extra_info"], dict):
            logger.error(f"{prefix}：extra_info 必须是对象")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

    @staticmethod
    def _parse_grpo_json_cell(value: Any, field_name: str, line_num: int, expected_type: type | tuple[type, ...]) -> Any:
        if isinstance(value, expected_type):
            return value
        if value is None:
            logger.error(f"第{line_num}行：{field_name} 不能为空")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if not isinstance(value, str):
            logger.error(f"第{line_num}行：{field_name} 必须是JSON字符串")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        value = value.strip()
        if not value:
            logger.error(f"第{line_num}行：{field_name} 不能为空")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        try:
            parsed_value = json.loads(value)
        except json.JSONDecodeError as e:
            logger.error(f"第{line_num}行：{field_name} JSON格式错误: {str(e)}")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if not isinstance(parsed_value, expected_type):
            logger.error(f"第{line_num}行：{field_name} JSON类型错误")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        return parsed_value

    @staticmethod
    def _set_nested_value(target: Dict[str, Any], dotted_key: str, value: Any) -> None:
        current = target
        parts = dotted_key.split(".")
        for part in parts[:-1]:
            if part not in current or not isinstance(current[part], dict):
                current[part] = {}
            current = current[part]
        current[parts[-1]] = value

    def _normalize_grpo_xlsx_row(self, row_data: Dict[str, Any], line_num: int) -> Dict[str, Any]:
        item: Dict[str, Any] = {}

        if row_data.get("data_source") is not None:
            item["data_source"] = str(row_data["data_source"]).strip()
        if row_data.get("prompt") is not None:
            item["prompt"] = self._parse_grpo_json_cell(row_data["prompt"], "prompt", line_num, list)
        if row_data.get("ability") is not None:
            item["ability"] = str(row_data["ability"]).strip()

        if row_data.get("reward_model") is not None:
            item["reward_model"] = self._parse_grpo_json_cell(row_data["reward_model"], "reward_model", line_num, dict)
        else:
            reward_model: Dict[str, Any] = {}
            if row_data.get("reward_model.style") is not None:
                reward_model["style"] = str(row_data["reward_model.style"]).strip()
            elif row_data.get("reward_style") is not None:
                reward_model["style"] = str(row_data["reward_style"]).strip()
            if row_data.get("reward_model.ground_truth") is not None:
                reward_model["ground_truth"] = row_data["reward_model.ground_truth"]
            elif row_data.get("ground_truth") is not None:
                reward_model["ground_truth"] = row_data["ground_truth"]
            if reward_model:
                item["reward_model"] = reward_model

        if row_data.get("extra_info") is not None:
            item["extra_info"] = self._parse_grpo_json_cell(row_data["extra_info"], "extra_info", line_num, dict)

        for key, value in row_data.items():
            if value is None or value == "":
                continue
            if key.startswith("extra_info."):
                if "extra_info" not in item:
                    item["extra_info"] = {}
                self._set_nested_value(item["extra_info"], key.removeprefix("extra_info."), value)
            elif key.startswith("reward_model.") and key not in ("reward_model.style", "reward_model.ground_truth"):
                if "reward_model" not in item:
                    item["reward_model"] = {}
                self._set_nested_value(item["reward_model"], key.removeprefix("reward_model."), value)
            elif key == "images":
                item["images"] = self._parse_grpo_json_cell(value, "images", line_num, list)

        self._validate_grpo_item(item, line_num)
        return item

    @staticmethod
    def _validate_role_based_messages(
        messages: Any,
        line_num: int,
        *,
        allow_images: bool = False,
        images_list: Optional[Any] = None,
    ) -> None:
        if not isinstance(messages, list) or len(messages) == 0:
            logger.error(f"第{line_num}行：messages必须是非空数组")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        expected_role = None
        image_tag_count = 0
        for msg_idx, message in enumerate(messages):
            if not isinstance(message, dict):
                logger.error(f"第{line_num}行第{msg_idx + 1}个message：必须是对象格式")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if "role" not in message:
                logger.error(f"第{line_num}行第{msg_idx + 1}个message：缺少role字段")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if "content" not in message:
                logger.error(f"第{line_num}行第{msg_idx + 1}个message：缺少content字段")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            role = message["role"]
            content = message["content"]
            if role not in ["user", "assistant", "system"]:
                logger.error(f"第{line_num}行第{msg_idx + 1}个message：role必须是user、assistant或system")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if not isinstance(content, str):
                logger.error(f"第{line_num}行第{msg_idx + 1}个message：content必须是字符串")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if role != "system" and not content.strip():
                logger.error(f"第{line_num}行第{msg_idx + 1}个message：content不能为空")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if allow_images:
                image_tag_count += content.count("<image>")

            if role != "system":
                if expected_role is None:
                    if role != "user":
                        logger.error(f"第{line_num}行：第一个非system的message必须是user")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                    expected_role = "assistant"
                else:
                    if role != expected_role:
                        logger.error(f"第{line_num}行：messages中role必须交替出现")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                    expected_role = "assistant" if role == "user" else "user"

        if allow_images:
            if images_list is None:
                images_list = []
            if not isinstance(images_list, list):
                logger.error(f"第{line_num}行：images字段必须是数组")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if image_tag_count != len(images_list):
                logger.error(f"第{line_num}行：<image>标签数量必须等于images数组长度")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            for image_name in images_list:
                if not isinstance(image_name, str) or not image_name.strip():
                    logger.error(f"第{line_num}行：images数组中的元素必须是非空字符串")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

    @staticmethod
    def _validate_dpo_role_based_response(item: Any, field_name: str, line_num: int) -> None:
        if not isinstance(item, dict):
            logger.error(f"第{line_num}行：{field_name} 必须是对象格式")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if item.get("role") != "assistant":
            logger.error(f"第{line_num}行：{field_name}.role 必须是 assistant")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        content = item.get("content")
        if not isinstance(content, str) or not content.strip():
            logger.error(f"第{line_num}行：{field_name}.content 不能为空")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

    async def analyze_prompt_response_json_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        JSON转JSONL并解析，同时对文件内容进行初步的清洗与处理，无需再进行二次处理，节省时间

        Args:
            file_content: json文件的字节内容

        Returns:
            解析结果
        """
        try:
            # 1. 解析JSON内容
            json_str = file_content.decode('utf-8')
            data = json.loads(json_str)

            # 2. 流式处理数据
            jsonl_lines = []
            total_rows = 0
            processed_rows = 0
            skip_rows = 0
            total_characters = 0

            if isinstance(data, list):
                # 统计总行数（数组长度）
                total_rows = len(data)
                # 批量处理数组
                for i in range(0, len(data), self.chunk_size):
                    batch = data[i:i + self.chunk_size]
                    processed_batch, processed, skipped, chars = self._process_json_batch(batch)
                    jsonl_lines.extend(processed_batch)

                    # 更新处理结果数据
                    processed_rows += processed
                    skip_rows += skipped
                    total_characters += chars

            elif isinstance(data, dict):
                # 处理单个对象
                if "prompt" in data and "response" in data:
                    if data["prompt"] and data["response"]:
                        jsonl_line = '[' + json.dumps(data, ensure_ascii=False) + ']'
                        jsonl_lines.append(jsonl_line)
                        total_rows += 1
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())
                    else:
                        skip_rows += 1

            # 3. 返回结果
            content = '\n'.join(jsonl_lines).encode('utf-8')

            logger.info(
                f"文件总共处理：{total_rows}行\n其中：{skip_rows}行有误被跳过\n剩余{processed_rows}行转换成功"
            )

            # 返回数据总行数等相关信息
            return TextGenerationParseResult(
                total_rows=total_rows,
                processed_rows=processed_rows,
                skip_rows=skip_rows,
                total_characters=total_characters,
                convert_file_content=content
            )

        except json.JSONDecodeError as e:
            logger.error(f"文件JSON格式异常: {str(e)}")
            raise ValueError("文件格式错误")

        except Exception as e:
            logger.error(f"文件解析服务异常: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_prompt_response_xlsx_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        xlsx文件转化为jsonl并解析，同时对文件内容进行初步的清洗与处理，无需再进行二次处理，节省时间

        Args:
            file_content: xlsx文件的字节内容

        Returns:
            转换后的jsonl格式字节内容
        """

        def _parse_xlsx_sync() -> TextGenerationParseResult:
            """同步解析XLSX文件（在线程池中执行）"""
            try:
                # 使用只读模式加载工作簿，大幅减少内存占用
                file_stream = io.BytesIO(file_content)

                # 使用read_only=True和data_only=True
                # read_only=True: 只读模式，不加载样式信息，内存占用减少70%
                # data_only=True: 只读取计算后的值，不保留公式
                workbook = load_workbook(
                    filename=file_stream,
                    read_only=True,  # 只读模式，关键优化
                    data_only=True  # 只读取值，不保留公式
                )
                worksheet = workbook.active

                if worksheet is None:
                    raise ValueError("无有效数据")

                if worksheet.max_row < 2:
                    raise ValueError("无有效数据")

                # 1. 读取表头并映射列
                column_mapping = self._map_columns(worksheet)

                # 2. 流式处理数据行
                jsonl_lines = []
                total_rows = worksheet.max_row - 1  # 总样本数（减1因为第一行是标题）
                processed_rows = 0 # 已转化样本数
                skip_rows = 0 # 跳过样本数
                total_characters = 0 # 总字符数

                # 使用worksheet.rows迭代器，避免逐个访问单元格
                # 跳过第一行（标题行）
                row_iter = iter(worksheet.rows)
                next(row_iter)  # 跳过标题行

                # 批量处理行数据
                batch = []
                for row in row_iter:
                    batch.append(row)

                    # 凑够一批再统一处理
                    if len(batch) >= self.chunk_size:
                        # 处理一批数据
                        processed_batch, processed, skipped, chars = self._process_xlsx_batch(batch, column_mapping)
                        jsonl_lines.extend(processed_batch)

                        # 累加，修改处理结果
                        processed_rows += processed
                        skip_rows += skipped
                        total_characters += chars

                        batch = []  # 清空批次

                # 处理最后凑不够一批的数据
                if batch:
                    processed_batch, processed, skipped, chars = self._process_xlsx_batch(batch, column_mapping)
                    jsonl_lines.extend(processed_batch)

                    # 累加，修改处理结果
                    processed_rows += processed
                    skip_rows += skipped
                    total_characters += chars

                # 返回结果（json串使用urf-8编码）
                content = '\n'.join(jsonl_lines).encode('utf-8')

                # 返回数据总行数等相关信息
                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters = total_characters,
                    convert_file_content = content
                )

            except InvalidFileException as e:
                logger.error(f"转换Excel文件失败: {str(e)}")
                raise ValueError("文件格式错误")

            except ValueError:
                raise
            except Exception as e:
                logger.error(f"转换Excel文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        # 在线程池中执行同步解析
        current_loop = asyncio.get_running_loop()
        result = await current_loop.run_in_executor(self.executor, _parse_xlsx_sync)

        logger.info(
            f"文件总共处理：{result.total_rows}行\n其中：{result.skip_rows}行有误被跳过\n剩余{result.processed_rows}行转换成功"
        )
        return result

    async def analyze_prompt_response_jsonl_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        分析上传文件内容，返回样本数和字符数、仅支持jsonl文件，最基础的解析逻辑
        """
        total_samples = 0
        total_characters = 0

        try:
            # 将字节内容转换为字符串
            content_str = file_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue

                # 跳过注释行（支持 # 前面有空格的情况）
                if line.lstrip().startswith('#'):
                    continue

                try:
                    # 解析JSON
                    parsed_data = json.loads(line)

                    # 验证数据格式：应该是一个数组
                    if not isinstance(parsed_data, list):
                        logger.error(f"第{line_num}行：应为数组格式")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    if len(parsed_data) == 0:
                        logger.error(f"第{line_num}行：数组为空")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 验证数组中的每个对象
                    for obj_index, item in enumerate(parsed_data):
                        if not isinstance(item, dict):
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：应是json对象格式")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证必需字段
                        if 'prompt' not in item:
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：缺少必需字段 prompt")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        if 'response' not in item:
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：缺少必需字段 response")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证字段类型
                        if not isinstance(item['prompt'], str):
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：prompt 必须是字符串")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        if not isinstance(item['response'], str):
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：response 必须是字符串")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证可选字段（如果存在）
                        if 'system' in item and not isinstance(item['system'], str):
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：system 必须是字符串")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 检查字段内容是否为空
                        if not item['prompt'].strip():
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：prompt 不能为空")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        if not item['response'].strip():
                            logger.error(f"第{line_num}行第{obj_index + 1}个对象：response 不能为空")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 计算样本数：每行算一个样本（与预览逻辑保持一致）
                    total_samples += 1
                    total_characters += len(line)

                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if total_samples == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_samples,
                skip_rows=0,
                processed_rows=total_samples,
                total_characters=total_characters,
                convert_file_content=file_content
            )

        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_grpo_jsonl_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 GRPO JSONL 文件，校验 verl 所需字段并原样保存。"""
        total_samples = 0
        total_characters = 0
        normalized_lines: List[str] = []

        try:
            content_str = file_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line or line.lstrip().startswith('#'):
                    continue

                try:
                    parsed_data = json.loads(line)
                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                if not isinstance(parsed_data, dict):
                    logger.error(f"第{line_num}行：GRPO样本必须是JSON对象")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                self._validate_grpo_item(parsed_data, line_num)
                jsonl_line = json.dumps(parsed_data, ensure_ascii=False)
                normalized_lines.append(jsonl_line)
                total_samples += 1
                total_characters += len(jsonl_line)

            if total_samples == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_samples,
                skip_rows=0,
                processed_rows=total_samples,
                total_characters=total_characters,
                convert_file_content="\n".join(normalized_lines).encode("utf-8"),
            )
        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析GRPO JSONL文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_grpo_json_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 GRPO JSON 文件，支持单对象或对象数组，内部转换为 JSONL。"""
        try:
            json_str = file_content.decode("utf-8")
            data = json.loads(json_str)

            if isinstance(data, dict):
                data = [data]
            if not isinstance(data, list):
                raise ValueError("字段/格式错误")

            normalized_lines = []
            total_characters = 0
            for index, item in enumerate(data, 1):
                if not isinstance(item, dict):
                    raise ValueError(f"第 {index} 个样本：字段/格式错误")
                self._validate_grpo_item(item, index)
                jsonl_line = json.dumps(item, ensure_ascii=False)
                normalized_lines.append(jsonl_line)
                total_characters += len(jsonl_line)

            if not normalized_lines:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=len(data),
                processed_rows=len(normalized_lines),
                skip_rows=0,
                total_characters=total_characters,
                convert_file_content="\n".join(normalized_lines).encode("utf-8"),
            )
        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except json.JSONDecodeError as e:
            logger.error(f"文件JSON格式异常: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析GRPO JSON文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_grpo_xlsx_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 GRPO XLSX 文件，复杂字段使用 JSON 字符串，内部转换为 JSONL。"""

        def _parse_xlsx_sync() -> TextGenerationParseResult:
            try:
                file_stream = io.BytesIO(file_content)
                workbook = load_workbook(
                    filename=file_stream,
                    read_only=True,
                    data_only=True,
                )
                worksheet = workbook.active
                if worksheet is None or worksheet.max_row < 2:
                    raise ValueError("无有效数据")

                row_iter = worksheet.iter_rows(values_only=True)
                raw_headers = next(row_iter)
                headers = [str(cell).strip() if cell is not None else "" for cell in raw_headers]
                if not any(headers):
                    raise ValueError("无有效数据")

                normalized_lines: List[str] = []
                total_rows = worksheet.max_row - 1
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                for row_index, row in enumerate(row_iter, 2):
                    row_data = {
                        headers[col_index]: value
                        for col_index, value in enumerate(row)
                        if col_index < len(headers)
                        and headers[col_index]
                        and value is not None
                    }
                    if not row_data:
                        skip_rows += 1
                        continue

                    item = self._normalize_grpo_xlsx_row(row_data, row_index)
                    jsonl_line = json.dumps(item, ensure_ascii=False)
                    normalized_lines.append(jsonl_line)
                    processed_rows += 1
                    total_characters += len(jsonl_line)

                if processed_rows == 0:
                    raise ValueError("无有效数据")

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content="\n".join(normalized_lines).encode("utf-8"),
                )
            except InvalidFileException as e:
                logger.error(f"转换GRPO Excel文件失败: {str(e)}")
                raise ValueError("文件格式错误")
            except ValueError:
                raise
            except Exception as e:
                logger.error(f"转换GRPO Excel文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        current_loop = asyncio.get_running_loop()
        result = await current_loop.run_in_executor(self.executor, _parse_xlsx_sync)
        logger.info(
            f"文件总共处理：{result.total_rows}行\n其中：{result.skip_rows}行有误被跳过\n剩余{result.processed_rows}行转换成功"
        )
        return result

    async def analyze_dpo_prompt_response_json_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO prompt-response JSON 文件。"""
        try:
            json_str = file_content.decode('utf-8')
            data = json.loads(json_str)

            if isinstance(data, dict):
                data = [data]
            if not isinstance(data, list):
                raise ValueError("字段/格式错误")

            jsonl_lines = []
            total_rows = len(data)
            processed_rows = 0
            total_characters = 0

            for index, item in enumerate(data, 1):
                if not isinstance(item, dict):
                    raise ValueError(f"第 {index} 个样本：字段/格式错误")
                self._validate_dpo_prompt_response_item(item, index)
                normalized_item = {
                    "instruction": item["instruction"],
                    "input": item.get("input", "") if item.get("input") is not None else "",
                    "chosen": item["chosen"],
                    "rejected": item["rejected"],
                }
                jsonl_line = json.dumps(normalized_item, ensure_ascii=False)
                jsonl_lines.append(jsonl_line)
                processed_rows += 1
                total_characters += len(jsonl_line.strip())

            if processed_rows == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_rows,
                processed_rows=processed_rows,
                skip_rows=0,
                total_characters=total_characters,
                convert_file_content="\n".join(jsonl_lines).encode("utf-8"),
            )
        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析DPO prompt-response JSON文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_dpo_prompt_response_jsonl_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO prompt-response JSONL 文件。"""
        total_samples = 0
        total_characters = 0

        try:
            content_str = file_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line or line.lstrip().startswith('#'):
                    continue

                try:
                    parsed_data = json.loads(line)
                    if not isinstance(parsed_data, dict):
                        logger.error(f"第{line_num}行：应为对象格式")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                    self._validate_dpo_prompt_response_item(parsed_data, line_num)

                    total_samples += 1
                    total_characters += len(line)
                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if total_samples == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_samples,
                skip_rows=0,
                processed_rows=total_samples,
                total_characters=total_characters,
                convert_file_content=file_content,
            )
        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析DPO prompt-response JSONL文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_dpo_prompt_response_xlsx_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO prompt-response XLSX 文件。"""

        def _parse_xlsx_sync() -> TextGenerationParseResult:
            try:
                file_stream = io.BytesIO(file_content)
                workbook = load_workbook(filename=file_stream, read_only=True, data_only=True)
                worksheet = workbook.active

                if worksheet is None or worksheet.max_row < 2:
                    raise ValueError("无有效数据")

                headers = []
                for cell in next(worksheet.rows):
                    headers.append(str(cell.value).strip().lower() if cell.value else "")

                header_to_index = {header: i for i, header in enumerate(headers)}
                required_headers = ("instruction", "input", "chosen", "rejected")
                missing_headers = [header for header in required_headers if header not in header_to_index]
                if missing_headers:
                    logger.error(f"Excel文件缺少必需列: {missing_headers}")
                    raise ValueError("文件格式错误")

                jsonl_lines = []
                total_rows = worksheet.max_row - 1
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                row_iter = iter(worksheet.rows)
                next(row_iter)

                for row_num, row in enumerate(row_iter, start=2):
                    try:
                        item = {
                            "instruction": str(row[header_to_index["instruction"]].value).strip() if row[header_to_index["instruction"]].value is not None else "",
                            "input": str(row[header_to_index["input"]].value).strip() if row[header_to_index["input"]].value is not None else "",
                            "chosen": str(row[header_to_index["chosen"]].value).strip() if row[header_to_index["chosen"]].value is not None else "",
                            "rejected": str(row[header_to_index["rejected"]].value).strip() if row[header_to_index["rejected"]].value is not None else "",
                        }
                        self._validate_dpo_prompt_response_item(item, row_num - 1)
                        jsonl_line = json.dumps(item, ensure_ascii=False)
                        jsonl_lines.append(jsonl_line)
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())
                    except ValueError:
                        skip_rows += 1
                        continue

                if processed_rows == 0:
                    raise ValueError("无有效数据")

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content="\n".join(jsonl_lines).encode("utf-8"),
                )
            except InvalidFileException as e:
                logger.error(f"无效的文件格式: {str(e)}")
                raise ValueError("文件格式错误")
            except ValueError:
                raise
            except Exception as e:
                logger.error(f"解析DPO prompt-response XLSX文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        current_loop = asyncio.get_running_loop()
        return await current_loop.run_in_executor(self.executor, _parse_xlsx_sync)

    async def analyze_dpo_prompt_response_csv_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO prompt-response CSV 文件。"""

        def _parse_csv_sync() -> TextGenerationParseResult:
            try:
                encodings_to_try = ['utf-8', 'gbk', 'gb18030', 'latin-1', 'cp1252']
                csv_content = None
                for encoding in encodings_to_try:
                    try:
                        csv_content = file_content.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue

                if csv_content is None:
                    logger.error("CSV文件编码不支持，请使用UTF-8编码")
                    raise ValueError("文件格式错误")

                file_stream = io.StringIO(csv_content)
                csv_reader = csv.reader(file_stream)
                try:
                    headers = next(csv_reader)
                    headers = [h.strip().lower() if h else "" for h in headers]
                except StopIteration:
                    raise ValueError("无有效数据")

                header_to_index = {header: i for i, header in enumerate(headers)}
                required_headers = ("instruction", "input", "chosen", "rejected")
                missing_headers = [header for header in required_headers if header not in header_to_index]
                if missing_headers:
                    logger.error(f"CSV文件缺少必需列: {missing_headers}")
                    raise ValueError("文件格式错误")

                jsonl_lines = []
                total_rows = 0
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                for row_num, row in enumerate(csv_reader, start=2):
                    total_rows += 1
                    try:
                        item = {
                            "instruction": row[header_to_index["instruction"]].strip() if header_to_index["instruction"] < len(row) and row[header_to_index["instruction"]] else "",
                            "input": row[header_to_index["input"]].strip() if header_to_index["input"] < len(row) and row[header_to_index["input"]] else "",
                            "chosen": row[header_to_index["chosen"]].strip() if header_to_index["chosen"] < len(row) and row[header_to_index["chosen"]] else "",
                            "rejected": row[header_to_index["rejected"]].strip() if header_to_index["rejected"] < len(row) and row[header_to_index["rejected"]] else "",
                        }
                        self._validate_dpo_prompt_response_item(item, row_num - 1)
                        jsonl_line = json.dumps(item, ensure_ascii=False)
                        jsonl_lines.append(jsonl_line)
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())
                    except ValueError:
                        skip_rows += 1
                        continue

                if processed_rows == 0:
                    raise ValueError("无有效数据")

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content="\n".join(jsonl_lines).encode("utf-8"),
                )
            except ValueError:
                raise
            except Exception as e:
                logger.error(f"解析DPO prompt-response CSV文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        current_loop = asyncio.get_running_loop()
        return await current_loop.run_in_executor(self.executor, _parse_csv_sync)

    async def analyze_prompt_response_csv_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        CSV文件转化为jsonl并解析，同时对文件内容进行初步的清洗与处理，无需再进行二次处理，节省时间

        Args:
            file_content: CSV文件的字节内容

        Returns:
            转换后的jsonl格式字节内容
        """

        def _parse_csv_sync() -> TextGenerationParseResult:
            """同步解析CSV文件（在线程池中执行）"""
            try:
                # 尝试多种编码
                encodings_to_try = ['utf-8', 'gbk', 'gb18030', 'latin-1', 'cp1252']
                csv_content = None
                encoding_used = None

                for encoding in encodings_to_try:
                    try:
                        csv_content = file_content.decode(encoding)
                        encoding_used = encoding
                        break
                    except UnicodeDecodeError:
                        continue

                if csv_content is None:
                    logger.error("CSV文件编码不支持，请使用UTF-8编码")
                    raise ValueError("文件格式错误")

                # 使用csv模块解析CSV内容
                file_stream = io.StringIO(csv_content)
                csv_reader = csv.reader(file_stream)

                # 读取表头
                try:
                    headers = next(csv_reader)
                    headers = [h.strip().lower() if h else "" for h in headers]
                except StopIteration:
                    raise ValueError("无有效数据")

                # 映射列信息
                header_to_index = {header: i for i, header in enumerate(headers)}
                
                # 定义列名别名（小写）
                prompt_aliases = ['prompt']
                response_aliases = ['response']
                system_aliases = ['system']
                
                # 查找各列位置
                prompt_col = None
                response_col = None
                system_col = None
                
                for alias in prompt_aliases:
                    if alias in header_to_index:
                        prompt_col = header_to_index[alias]
                        break
                
                for alias in response_aliases:
                    if alias in header_to_index:
                        response_col = header_to_index[alias]
                        break
                
                for alias in system_aliases:
                    if alias in header_to_index:
                        system_col = header_to_index[alias]
                        break
                
                # 验证必需列
                if prompt_col is None or response_col is None:
                    logger.error("CSV文件必须包含prompt和response列")
                    raise ValueError("文件格式错误")

                # 流式处理数据行
                jsonl_lines = []
                total_rows = 0
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                # 批量处理行数据
                batch = []
                for row in csv_reader:
                    total_rows += 1
                    batch.append(row)

                    # 凑够一批再统一处理
                    if len(batch) >= self.chunk_size:
                        processed_batch, processed, skipped, chars = self._process_csv_batch(batch, prompt_col, response_col, system_col)
                        jsonl_lines.extend(processed_batch)
                        processed_rows += processed
                        skip_rows += skipped
                        total_characters += chars
                        batch = []

                # 处理最后凑不够一批的数据
                if batch:
                    processed_batch, processed, skipped, chars = self._process_csv_batch(batch, prompt_col, response_col, system_col)
                    jsonl_lines.extend(processed_batch)
                    processed_rows += processed
                    skip_rows += skipped
                    total_characters += chars

                if total_rows == 0:
                    raise ValueError("无有效数据")

                # 返回结果（json串使用utf-8编码）
                content = '\n'.join(jsonl_lines).encode('utf-8')

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content=content
                )
            except ValueError:
                raise
            except Exception as e:
                logger.error(f"转换CSV文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        # 在线程池中执行同步解析
        current_loop = asyncio.get_running_loop()
        result = await current_loop.run_in_executor(self.executor, _parse_csv_sync)

        logger.info(
            f"CSV文件总共处理：{result.total_rows}行\n其中：{result.skip_rows}行有误被跳过\n剩余{result.processed_rows}行转换成功"
        )
        return result

    async def analyze_role_based_json_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        分析role-based格式的JSON文件内容
        参考图像理解数据集的解析规则，但无需验证图片标签

        Args:
            file_content: json文件的字节内容

        Returns:
            解析结果
        """
        try:
            # 1. 解析JSON内容
            json_str = file_content.decode('utf-8')
            data = json.loads(json_str)

            # 2. 流式处理数据
            jsonl_lines = []
            total_rows = 0
            processed_rows = 0
            skip_rows = 0
            total_characters = 0

            # 将数据转换为列表格式统一处理
            if isinstance(data, list):
                # 数组格式：每个元素是一个对话样本
                data_list = data
            elif isinstance(data, dict):
                # 单个对象格式：转换为列表
                data_list = [data]
            else:
                logger.error("根节点既非数组也非对象")
                raise ValueError("文件格式错误")

            # 统计总行数
            total_rows = len(data_list)

            # 批量处理数据
            for i in range(0, len(data_list), self.chunk_size):
                batch = data_list[i:i + self.chunk_size]
                
                for item_idx, item in enumerate(batch):
                    sample_num = i + item_idx + 1
                    
                    try:
                        if not isinstance(item, dict):
                            logger.warning(f"第{sample_num}个样本：不是标准的json对象格式")
                            raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                        # 验证 messages 字段
                        if 'messages' not in item:
                            logger.warning(f"第{sample_num}个样本：缺少messages字段")
                            raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                        messages = item['messages']
                        if not isinstance(messages, list) or len(messages) == 0:
                            logger.warning(f"第{sample_num}个样本：messages必须是非空数组")
                            raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                        # 验证每个 message
                        expected_role = None
                        for msg_idx, message in enumerate(messages):
                            if not isinstance(message, dict):
                                logger.warning(f"第{sample_num}个样本第{msg_idx + 1}个message：必须是对象格式")
                                raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                            # 验证 role 字段
                            if 'role' not in message:
                                logger.warning(f"第{sample_num}个样本第{msg_idx + 1}个message：缺少role字段")
                                raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                            role = message['role']
                            if role not in ['user', 'assistant', 'system']:
                                logger.warning(f"第{sample_num}个样本第{msg_idx + 1}个message：role必须是user、assistant或system")
                                raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                            # 验证 content 字段
                            if 'content' not in message:
                                logger.warning(f"第{sample_num}个样本第{msg_idx + 1}个message：缺少content字段")
                                raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                            content = message['content']
                            if not isinstance(content, str) or not content.strip():
                                logger.warning(f"第{sample_num}个样本第{msg_idx + 1}个message：content不能为空")
                                raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")

                            # 验证 role 交替出现（system 不影响交替规则）
                            if role != 'system':
                                # 如果是第一个非 system 的消息，必须是 user
                                if expected_role is None:
                                    if role != 'user':
                                        logger.warning(f"第{sample_num}个样本：第一个非system的message必须是user")
                                        raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")
                                    expected_role = 'assistant'
                                else:
                                    # 后续消息必须与期望的角色匹配
                                    if role != expected_role:
                                        logger.warning(f"第{sample_num}个样本：messages中role必须交替出现")
                                        raise ValueError(f"第 {sample_num} 个样本：字段/格式错误")
                                    # 切换期望的角色
                                    expected_role = 'assistant' if role == 'user' else 'user'

                        # 验证通过，序列化为JSONL格式
                        jsonl_line = json.dumps(item, ensure_ascii=False)
                        jsonl_lines.append(jsonl_line)
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())

                    except ValueError:
                        # 验证失败，跳过该样本
                        skip_rows += 1
                        continue

            if processed_rows == 0:
                raise ValueError("无有效数据")

            # 3. 返回结果
            content = '\n'.join(jsonl_lines).encode('utf-8')

            logger.info(
                f"文件总共处理：{total_rows}行\n其中：{skip_rows}行有误被跳过\n剩余{processed_rows}行转换成功"
            )

            return TextGenerationParseResult(
                total_rows=total_rows,
                processed_rows=processed_rows,
                skip_rows=skip_rows,
                total_characters=total_characters,
                convert_file_content=content
            )

        except json.JSONDecodeError as e:
            logger.error(f"文件格式异常: {str(e)}")
            raise ValueError("文件格式错误")

        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"文件解析服务异常: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_role_based_xlsx_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        分析role-based格式的xlsx文件内容
        XLSX格式说明：
        - 单轮数据集：表头为 system、user、assistant，每一行对应一个对话样本
        - 多轮数据集：表头为 system、user1、assistant1、user2、assistant2...，每一行对应一个多轮对话样本

        Args:
            file_content: xlsx文件的字节内容

        Returns:
            解析结果
        """

        def _parse_xlsx_sync() -> TextGenerationParseResult:
            """同步解析XLSX文件（在线程池中执行）"""
            try:
                # 使用只读模式加载工作簿
                file_stream = io.BytesIO(file_content)
                workbook = load_workbook(
                    filename=file_stream,
                    read_only=True,
                    data_only=True
                )
                worksheet = workbook.active

                if worksheet is None:
                    raise ValueError("无有效数据")

                if worksheet.max_row < 2:
                    raise ValueError("无有效数据")

                # 1. 读取表头
                headers = []
                header_row = next(worksheet.rows)
                for cell in header_row:
                    header_value = str(cell.value).strip() if cell.value else ""
                    headers.append(header_value)

                # 2. 判断是单轮还是多轮格式，并找到对应的列索引
                system_col = None
                user_cols = []  # 存储所有user列的索引
                assistant_cols = []  # 存储所有assistant列的索引
                
                # 查找system列
                for i, header in enumerate(headers):
                    header_lower = header.lower()
                    if header_lower == 'system':
                        system_col = i
                        break

                # 判断格式：检查是否有user1或user（单轮）
                is_multi_turn = False
                for i, header in enumerate(headers):
                    header_lower = header.lower()
                    if header_lower.startswith('user') and header_lower != 'user':
                        # 发现user1, user2等，说明是多轮格式
                        is_multi_turn = True
                        break

                # 根据格式提取列索引
                if is_multi_turn:
                    # 多轮格式：system、user1、assistant1、user2、assistant2...
                    for i, header in enumerate(headers):
                        header_lower = header.lower()
                        if header_lower.startswith('user') and header_lower != 'user':
                            # 提取数字，如user1 -> 1, user2 -> 2
                            try:
                                num = int(header_lower.replace('user', ''))
                                user_cols.append((num, i))
                            except ValueError:
                                continue
                        elif header_lower.startswith('assistant') and header_lower != 'assistant':
                            try:
                                num = int(header_lower.replace('assistant', ''))
                                assistant_cols.append((num, i))
                            except ValueError:
                                continue
                    
                    # 按数字排序
                    user_cols.sort(key=lambda x: x[0])
                    assistant_cols.sort(key=lambda x: x[0])
                else:
                    # 单轮格式：system、user、assistant
                    for i, header in enumerate(headers):
                        header_lower = header.lower()
                        if header_lower == 'user':
                            user_cols.append((1, i))
                        elif header_lower == 'assistant':
                            assistant_cols.append((1, i))

                # 验证必需的列
                if system_col is None:
                    logger.error("Excel文件必须包含system列")
                    raise ValueError("文件格式错误")
                if len(user_cols) == 0:
                    logger.error("Excel文件必须包含user列（单轮）或user1、user2...列（多轮）")
                    raise ValueError("文件格式错误")
                if len(assistant_cols) == 0:
                    logger.error("Excel文件必须包含assistant列（单轮）或assistant1、assistant2...列（多轮）")
                    raise ValueError("文件格式错误")
                if len(user_cols) != len(assistant_cols):
                    logger.error("Excel文件中user列和assistant列的数量必须相等")
                    raise ValueError("文件格式错误")

                # 3. 流式处理数据行
                jsonl_lines = []
                total_rows = worksheet.max_row - 1  # 总样本数（减1因为第一行是标题）
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                # 跳过第一行（标题行）
                row_iter = iter(worksheet.rows)
                next(row_iter)

                # 处理每一行数据
                for row_num, row in enumerate(row_iter, start=2):
                    try:
                        # 构建messages数组
                        messages = []

                        # 添加system消息（如果存在）
                        if system_col < len(row) and row[system_col].value:
                            system_content = str(row[system_col].value).strip()
                            if system_content:
                                messages.append({
                                    "role": "system",
                                    "content": system_content
                                })

                        # 添加user和assistant消息对
                        for turn_num, (user_num, user_col_idx) in enumerate(user_cols, start=1):
                            # 查找对应的assistant列
                            assistant_col_idx = None
                            for ast_num, ast_idx in assistant_cols:
                                if ast_num == user_num:
                                    assistant_col_idx = ast_idx
                                    break

                            if assistant_col_idx is None:
                                logger.warning(f"第{row_num}个样本：找不到与user{user_num}对应的assistant列")
                                raise ValueError(f"第 {row_num} 个样本：字段/格式错误")

                            # 获取user内容
                            if user_col_idx >= len(row):
                                continue
                            user_cell = row[user_col_idx]
                            user_content = str(user_cell.value).strip() if user_cell.value else ""

                            # 获取assistant内容
                            if assistant_col_idx >= len(row):
                                continue
                            assistant_cell = row[assistant_col_idx]
                            assistant_content = str(assistant_cell.value).strip() if assistant_cell.value else ""

                            # 如果user或assistant为空，跳过这一轮
                            if not user_content or not assistant_content:
                                if turn_num == 1:
                                    # 第一轮必须有内容
                                    logger.warning(f"第{row_num}个样本：第一轮对话的user和assistant不能为空")
                                    raise ValueError(f"第 {row_num} 个样本：字段/格式错误")
                                else:
                                    # 后续轮次可以为空，表示对话结束
                                    break

                            # 添加user消息
                            messages.append({
                                "role": "user",
                                "content": user_content
                            })

                            # 添加assistant消息
                            messages.append({
                                "role": "assistant",
                                "content": assistant_content
                            })

                        # 验证messages数组
                        if len(messages) == 0:
                            logger.warning(f"第{row_num}个样本：没有找到有效的对话内容")
                            raise ValueError(f"第 {row_num} 个样本：字段/格式错误")

                        # 验证role交替规则
                        expected_role = None
                        for msg_idx, message in enumerate(messages):
                            role = message['role']
                            if role != 'system':
                                if expected_role is None:
                                    if role != 'user':
                                        logger.warning(f"第{row_num}个样本：第一个非system的message必须是user")
                                        raise ValueError(f"第 {row_num} 个样本：字段/格式错误")
                                    expected_role = 'assistant'
                                else:
                                    if role != expected_role:
                                        logger.warning(f"第{row_num}个样本：messages中role必须交替出现")
                                        raise ValueError(f"第 {row_num} 个样本：字段/格式错误")
                                    expected_role = 'assistant' if role == 'user' else 'user'

                        # 构建数据对象
                        data = {
                            "messages": messages
                        }

                        # 序列化为JSONL格式
                        jsonl_line = json.dumps(data, ensure_ascii=False)
                        jsonl_lines.append(jsonl_line)
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())

                    except ValueError:
                        # 验证失败，跳过该行
                        skip_rows += 1
                        continue
                    except Exception as e:
                        skip_rows += 1
                        logger.warning(f"处理第{row_num}个样本时发生错误: {str(e)}")
                        continue

                if processed_rows == 0:
                    raise ValueError("无有效数据")

                # 返回结果
                content = '\n'.join(jsonl_lines).encode('utf-8')

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content=content
                )

            except InvalidFileException as e:
                logger.error(f"无效的文件格式: {str(e)}")
                raise ValueError("文件格式错误")

            except ValueError:
                raise
            except Exception as e:
                logger.error(f"文件解析服务异常: {str(e)}")
                raise ValueError("文件解析服务异常")

        # 在线程池中执行同步解析
        current_loop = asyncio.get_running_loop()
        result = await current_loop.run_in_executor(self.executor, _parse_xlsx_sync)

        logger.info(
            f"文件总共处理：{result.total_rows}行\n其中：{result.skip_rows}行有误被跳过\n剩余{result.processed_rows}行转换成功"
        )
        return result

    async def analyze_role_based_jsonl_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        分析role-based格式的JSONL文件内容
        参考图像理解数据集的解析规则，但无需验证图片标签

        Args:
            file_content: jsonl文件的字节内容

        Returns:
            解析结果
        """
        total_samples = 0
        total_characters = 0

        try:
            # 将字节内容转换为字符串
            content_str = file_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue

                # 跳过注释行（支持 # 前面有空格的情况）
                if line.lstrip().startswith('#'):
                    continue

                try:
                    # 解析JSON
                    parsed_data = json.loads(line)

                    if not isinstance(parsed_data, dict):
                        logger.error(f"第{line_num}行：数据应是JSON对象格式")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 验证 messages 字段
                    if 'messages' not in parsed_data:
                        logger.error(f"第{line_num}行：缺少messages字段")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    messages = parsed_data['messages']
                    if not isinstance(messages, list) or len(messages) == 0:
                        logger.error(f"第{line_num}行：messages必须是非空数组")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 验证每个 message
                    # 找到第一个非 system 的消息，它必须是 user
                    expected_role = None
                    for msg_idx, message in enumerate(messages):
                        if not isinstance(message, dict):
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：必须是对象格式")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证 role 字段
                        if 'role' not in message:
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：缺少role字段")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        role = message['role']
                        if role not in ['user', 'assistant', 'system']:
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：role必须是user、assistant或system")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证 content 字段
                        if 'content' not in message:
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：缺少content字段")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        content = message['content']
                        if not isinstance(content, str) or not content.strip():
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：content不能为空")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证 role 交替出现（system 不影响交替规则）
                        if role != 'system':
                            # 如果是第一个非 system 的消息，必须是 user
                            if expected_role is None:
                                if role != 'user':
                                    logger.error(f"第{line_num}行：第一个非system的message必须是user")
                                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                                expected_role = 'assistant'
                            else:
                                # 后续消息必须与期望的角色匹配
                                if role != expected_role:
                                    logger.error(f"第{line_num}行：messages中role必须交替出现")
                                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                                # 切换期望的角色
                                expected_role = 'assistant' if role == 'user' else 'user'

                    # 计算样本数和字符数
                    total_samples += 1
                    total_characters += len(line)

                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if total_samples == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_samples,
                skip_rows=0,
                processed_rows=total_samples,
                total_characters=total_characters,
                convert_file_content=file_content
            )

        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_dpo_role_based_json_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO role-based JSON 文件。"""
        try:
            json_str = file_content.decode('utf-8')
            data = json.loads(json_str)

            if isinstance(data, dict):
                data = [data]
            if not isinstance(data, list):
                raise ValueError("字段/格式错误")

            jsonl_lines = []
            total_rows = len(data)
            processed_rows = 0
            total_characters = 0

            for index, item in enumerate(data, 1):
                if not isinstance(item, dict):
                    raise ValueError(f"第 {index} 个样本：字段/格式错误")
                if 'messages' not in item or 'chosen' not in item or 'rejected' not in item:
                    raise ValueError(f"第 {index} 个样本：字段/格式错误")
                self._validate_role_based_messages(
                    item['messages'],
                    index,
                    allow_images='images' in item,
                    images_list=item.get('images'),
                )
                self._validate_dpo_role_based_response(item['chosen'], 'chosen', index)
                self._validate_dpo_role_based_response(item['rejected'], 'rejected', index)
                jsonl_line = json.dumps(item, ensure_ascii=False)
                jsonl_lines.append(jsonl_line)
                processed_rows += 1
                total_characters += len(jsonl_line.strip())

            if processed_rows == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_rows,
                processed_rows=processed_rows,
                skip_rows=0,
                total_characters=total_characters,
                convert_file_content="\n".join(jsonl_lines).encode("utf-8"),
            )
        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析DPO role-based JSON文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_dpo_role_based_jsonl_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO role-based JSONL 文件。"""
        total_samples = 0
        total_characters = 0

        try:
            content_str = file_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line or line.lstrip().startswith('#'):
                    continue

                try:
                    parsed_data = json.loads(line)
                    if not isinstance(parsed_data, dict):
                        logger.error(f"第{line_num}行：数据应是JSON对象格式")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                    if 'messages' not in parsed_data or 'chosen' not in parsed_data or 'rejected' not in parsed_data:
                        logger.error(f"第{line_num}行：缺少messages/chosen/rejected字段")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    self._validate_role_based_messages(
                        parsed_data['messages'],
                        line_num,
                        allow_images='images' in parsed_data,
                        images_list=parsed_data.get('images'),
                    )
                    self._validate_dpo_role_based_response(parsed_data['chosen'], 'chosen', line_num)
                    self._validate_dpo_role_based_response(parsed_data['rejected'], 'rejected', line_num)

                    total_samples += 1
                    total_characters += len(line)
                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if total_samples == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_samples,
                skip_rows=0,
                processed_rows=total_samples,
                total_characters=total_characters,
                convert_file_content=file_content,
            )
        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析DPO role-based JSONL文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_dpo_role_based_xlsx_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO role-based XLSX 文件。"""

        def _parse_xlsx_sync() -> TextGenerationParseResult:
            try:
                file_stream = io.BytesIO(file_content)
                workbook = load_workbook(filename=file_stream, read_only=True, data_only=True)
                worksheet = workbook.active

                if worksheet is None or worksheet.max_row < 2:
                    raise ValueError("无有效数据")

                headers = []
                for cell in next(worksheet.rows):
                    headers.append(str(cell.value).strip().lower() if cell.value else "")

                header_to_index = {header: i for i, header in enumerate(headers)}
                required_headers = ("messages", "chosen", "rejected")
                missing_headers = [header for header in required_headers if header not in header_to_index]
                if missing_headers:
                    logger.error(f"Excel文件缺少必需列: {missing_headers}")
                    raise ValueError("文件格式错误")

                jsonl_lines = []
                total_rows = worksheet.max_row - 1
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                row_iter = iter(worksheet.rows)
                next(row_iter)

                for row_num, row in enumerate(row_iter, start=2):
                    try:
                        item = self._build_dpo_role_based_row_item_from_mapping(
                            lambda key: str(row[header_to_index[key]].value).strip() if header_to_index[key] < len(row) and row[header_to_index[key]].value is not None else "",
                            header_to_index,
                            row_num - 1,
                        )
                        jsonl_line = json.dumps(item, ensure_ascii=False)
                        jsonl_lines.append(jsonl_line)
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())
                    except ValueError:
                        skip_rows += 1
                        continue

                if processed_rows == 0:
                    raise ValueError("无有效数据")

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content="\n".join(jsonl_lines).encode("utf-8"),
                )
            except InvalidFileException as e:
                logger.error(f"无效的文件格式: {str(e)}")
                raise ValueError("文件格式错误")
            except ValueError:
                raise
            except Exception as e:
                logger.error(f"解析DPO role-based XLSX文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        current_loop = asyncio.get_running_loop()
        return await current_loop.run_in_executor(self.executor, _parse_xlsx_sync)

    async def analyze_dpo_role_based_csv_content(self, file_content: bytes) -> TextGenerationParseResult:
        """解析 DPO role-based CSV 文件。"""

        def _parse_csv_sync() -> TextGenerationParseResult:
            try:
                encodings_to_try = ['utf-8', 'gbk', 'gb18030', 'latin-1', 'cp1252']
                csv_content = None
                for encoding in encodings_to_try:
                    try:
                        csv_content = file_content.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue

                if csv_content is None:
                    logger.error("CSV文件编码不支持，请使用UTF-8编码")
                    raise ValueError("文件格式错误")

                file_stream = io.StringIO(csv_content)
                csv_reader = csv.reader(file_stream)
                try:
                    headers = next(csv_reader)
                    headers = [h.strip().lower() if h else "" for h in headers]
                except StopIteration:
                    raise ValueError("无有效数据")

                header_to_index = {header: i for i, header in enumerate(headers)}
                required_headers = ("messages", "chosen", "rejected")
                missing_headers = [header for header in required_headers if header not in header_to_index]
                if missing_headers:
                    logger.error(f"CSV文件缺少必需列: {missing_headers}")
                    raise ValueError("文件格式错误")

                jsonl_lines = []
                total_rows = 0
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                for row_num, row in enumerate(csv_reader, start=2):
                    total_rows += 1
                    try:
                        item = self._build_dpo_role_based_row_item_from_mapping(
                            lambda key: row[header_to_index[key]].strip() if header_to_index[key] < len(row) and row[header_to_index[key]] else "",
                            header_to_index,
                            row_num - 1,
                        )
                        jsonl_line = json.dumps(item, ensure_ascii=False)
                        jsonl_lines.append(jsonl_line)
                        processed_rows += 1
                        total_characters += len(jsonl_line.strip())
                    except ValueError:
                        skip_rows += 1
                        continue

                if processed_rows == 0:
                    raise ValueError("无有效数据")

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content="\n".join(jsonl_lines).encode("utf-8"),
                )
            except ValueError:
                raise
            except Exception as e:
                logger.error(f"解析DPO role-based CSV文件失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        current_loop = asyncio.get_running_loop()
        return await current_loop.run_in_executor(self.executor, _parse_csv_sync)

    def _build_dpo_role_based_row_item_from_mapping(self, getter, header_to_index: Dict[str, int], line_num: int) -> Dict[str, Any]:
        item = {
            "messages": json.loads(getter("messages")),
            "chosen": json.loads(getter("chosen")),
            "rejected": json.loads(getter("rejected")),
        }
        if "images" in header_to_index:
            images_raw = getter("images")
            if images_raw:
                item["images"] = json.loads(images_raw)

        self._validate_role_based_messages(
            item["messages"],
            line_num,
            allow_images='images' in item,
            images_list=item.get('images'),
        )
        self._validate_dpo_role_based_response(item["chosen"], "chosen", line_num)
        self._validate_dpo_role_based_response(item["rejected"], "rejected", line_num)
        return item

class BusinessDatasetFileParser:
    """业务数据集文件解析器（格式要求宽松，不验证对象内容）"""

    def __init__(self, max_workers: int = 2, chunk_size: int = 1000):
        """
        初始化解析器

        Args:
            max_workers: 线程池最大工作线程数，默认2线程
            chunk_size: 批处理块大小，默认1000
        """
        self.executor = ThreadPoolExecutor(max_workers=max_workers)
        self.chunk_size = chunk_size

    @staticmethod
    def _process_json_batch(batch: List[Dict[str, Any]]) -> Tuple[List[str], int, int, int]:
        """
        批量处理JSON数据（业务数据集版本）
        不验证对象内容，只保证是有效的dict对象

        Args:
            batch: JSON对象批次

        Returns:
            处理后的JSONL行列表
        """
        processed_lines = []  # 处理完成的样本字符列表
        processed_count = 0  # 处理行数
        skip_count = 0  # 跳过行数
        char_count = 0  # 总字符数

        for item in batch:
            if not isinstance(item, dict):
                skip_count += 1
                continue

            # 序列化（不验证对象内容）
            jsonl_line = json.dumps(item, ensure_ascii=False)
            processed_lines.append(jsonl_line)
            processed_count += 1

            # 记录字符数
            char_count += len(jsonl_line.strip())

        return processed_lines, processed_count, skip_count, char_count

    @staticmethod
    def _process_xlsx_batch(batch: List, headers: List[str]) -> Tuple[List[str], int, int, int]:
        """
        批量处理XLSX数据行（业务数据集版本）
        将每一行转换为一个JSON对象，列名作为键

        Args:
            batch: 数据行批次
            headers: 表头列表

        Returns:
            处理后的JSONL行列表
        """
        processed_lines = []
        processed_count = 0
        skip_count = 0
        char_count = 0

        for row in batch:
            # 构造数据对象，使用表头作为键
            data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    # 将单元格值转换为字符串，None转为空字符串
                    cell_value = str(cell.value) if cell.value is not None else ""
                    data[headers[i]] = cell_value

            # 跳过完全空的行
            if not any(data.values()):
                skip_count += 1
                continue

            # 序列化为JSONL格式
            jsonl_line = json.dumps(data, ensure_ascii=False)
            processed_lines.append(jsonl_line)
            processed_count += 1
            char_count += len(jsonl_line.strip())

        return processed_lines, processed_count, skip_count, char_count

    @staticmethod
    def _process_csv_batch(batch: List[List[str]], headers: List[str]) -> Tuple[List[str], int, int, int]:
        """
        批量处理业务数据集CSV数据行
        将每一行转换为一个JSON对象，列名作为键，不验证对象内容

        Args:
            batch: CSV行数据批次
            headers: CSV表头列表

        Returns:
            处理后的JSONL行列表、处理行数、跳过行数、字符数
        """
        processed_lines = []
        processed_count = 0
        skip_count = 0
        char_count = 0

        for row in batch:
            # 跳过空行
            if not row or all(not cell.strip() for cell in row):
                skip_count += 1
                continue

            # 将行数据转换为字典，列名作为键
            data = {}
            for i, header in enumerate(headers):
                value = row[i].strip() if i < len(row) and row[i] else ""
                # 处理 NaN 值，转换为 None
                if value.lower() in ['nan', 'none', 'null', '']:
                    data[header] = None
                else:
                    data[header] = value

            # 序列化为JSONL格式（业务数据集：直接序列化为对象，不需要数组格式）
            jsonl_line = json.dumps(data, ensure_ascii=False)
            processed_lines.append(jsonl_line)
            processed_count += 1
            char_count += len(jsonl_line.strip())

        return processed_lines, processed_count, skip_count, char_count

    async def process_business_file(self, file_content: bytes, file_type: str) -> TextGenerationParseResult:
        """
        业务数据集文件解析

        Args:
            file_content: 原始文件字节内容
            file_type: 原始文件文件格式

        Returns:
            解析结果
        """
        try:
            if file_type == 'xlsx':
                return await self.analyze_xlsx_content(file_content)
            elif file_type == 'csv':
                return await self.analyze_csv_content(file_content)
            elif file_type == 'json':
                return await self.analyze_json_content(file_content)
            elif file_type == 'jsonl':
                return await self.analyze_jsonl_content(file_content)
            else:
                logger.error(f"当前数据集格式暂不支持：{file_type}")
                raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"文件解析服务异常: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_jsonl_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        分析业务数据集JSONL文件内容
        支持两种格式：
        1. 对象格式：每行一个JSON对象 {"key": "value"}
        2. 数组格式：每行一个JSON数组 [{"key": "value"}]（兼容文本生成数据集格式）
        不验证对象内容，只验证每行都是有效的JSON

        Args:
            file_content: jsonl文件的字节内容

        Returns:
            解析结果
        """
        total_samples = 0
        total_characters = 0

        try:
            # 将字节内容转换为字符串
            content_str = file_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue

                # 跳过注释行（支持 # 前面有空格的情况）
                if line.lstrip().startswith('#'):
                    continue

                try:
                    # 解析JSON
                    parsed_data = json.loads(line)

                    # 支持两种格式：
                    # 1. 对象格式：{"key": "value"}
                    # 2. 数组格式：[{"key": "value"}]（兼容文本生成数据集格式）
                    if isinstance(parsed_data, dict):
                        # 对象格式：直接使用
                        total_samples += 1
                        total_characters += len(line)
                    elif isinstance(parsed_data, list):
                        # 数组格式：提取数组中的第一个对象（兼容文本生成数据集格式）
                        if len(parsed_data) > 0 and isinstance(parsed_data[0], dict):
                            total_samples += 1
                            total_characters += len(line)
                        else:
                            logger.error(f"第{line_num}行：数组格式中应包含至少一个JSON对象")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                    else:
                        logger.error(f"第{line_num}行：数据应是JSON对象或包含对象的数组格式")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if total_samples == 0:
                raise ValueError("无有效数据")

            return TextGenerationParseResult(
                total_rows=total_samples,
                skip_rows=0,
                processed_rows=total_samples,
                total_characters=total_characters,
                convert_file_content=file_content
            )

        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"文件解析服务异常: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_json_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        JSON转JSONL（业务数据集版本）
        不验证对象内容，只保证每行都是有效的JSON对象

        Args:
            file_content: json文件的字节内容

        Returns:
            解析结果
        """
        try:
            # 1. 解析JSON内容
            json_str = file_content.decode('utf-8')
            data = json.loads(json_str)

            # 2. 流式处理数据
            jsonl_lines = []
            total_rows = 0
            processed_rows = 0
            skip_rows = 0
            total_characters = 0

            if isinstance(data, list):
                # 统计总行数（数组长度）
                total_rows = len(data)
                # 批量处理数组
                for i in range(0, len(data), self.chunk_size):
                    batch = data[i:i + self.chunk_size]
                    processed_batch, processed, skipped, chars = self._process_json_batch(batch)
                    jsonl_lines.extend(processed_batch)

                    # 更新处理结果数据
                    processed_rows += processed
                    skip_rows += skipped
                    total_characters += chars

            elif isinstance(data, dict):
                # 处理单个对象
                jsonl_line = json.dumps(data, ensure_ascii=False)
                jsonl_lines.append(jsonl_line)
                total_rows += 1
                processed_rows += 1
                total_characters += len(jsonl_line.strip())

            # 3. 返回结果
            content = '\n'.join(jsonl_lines).encode('utf-8')

            logger.info(
                f"文件总共处理：{total_rows}行\n其中：{skip_rows}行有误被跳过\n剩余{processed_rows}行转换成功"
            )

            # 返回数据总行数等相关信息
            return TextGenerationParseResult(
                total_rows=total_rows,
                processed_rows=processed_rows,
                skip_rows=skip_rows,
                total_characters=total_characters,
                convert_file_content=content
            )

        except json.JSONDecodeError as e:
            logger.error(f"文件格式异常: {str(e)}")
            raise ValueError("文件格式错误")

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"文件解析服务异常: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_xlsx_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        xlsx文件转化为jsonl（业务数据集版本）
        将每一行转换为一个JSON对象，列名作为键，不验证对象内容

        Args:
            file_content: xlsx文件的字节内容

        Returns:
            转换后的jsonl格式字节内容
        """

        def _parse_xlsx_sync() -> TextGenerationParseResult:
            """同步解析XLSX文件（在线程池中执行）"""
            try:
                # 使用只读模式加载工作簿，大幅减少内存占用
                file_stream = io.BytesIO(file_content)

                # 使用read_only=True和data_only=True
                workbook = load_workbook(
                    filename=file_stream,
                    read_only=True,
                    data_only=True
                )
                worksheet = workbook.active

                if worksheet is None:
                    raise ValueError("无有效数据")

                if worksheet.max_row < 2:
                    raise ValueError("无有效数据")

                # 1. 读取表头
                headers = []
                for cell in next(worksheet.rows):
                    header_value = str(cell.value) if cell.value else f"column_{len(headers)}"
                    headers.append(header_value)

                # 2. 流式处理数据行
                jsonl_lines = []
                total_rows = worksheet.max_row - 1  # 总样本数（减1因为第一行是标题）
                processed_rows = 0  # 已转化样本数
                skip_rows = 0  # 跳过样本数
                total_characters = 0  # 总字符数

                # 使用worksheet.rows迭代器，跳过第一行（标题行）
                row_iter = iter(worksheet.rows)
                next(row_iter)  # 跳过标题行

                # 批量处理行数据
                batch = []
                for row in row_iter:
                    batch.append(row)

                    # 凑够一批再统一处理
                    if len(batch) >= self.chunk_size:
                        processed_batch, processed, skipped, chars = self._process_xlsx_batch(batch, headers)
                        jsonl_lines.extend(processed_batch)

                        processed_rows += processed
                        skip_rows += skipped
                        total_characters += chars

                        batch = []  # 清空批次

                # 处理最后凑不够一批的数据
                if batch:
                    processed_batch, processed, skipped, chars = self._process_xlsx_batch(batch, headers)
                    jsonl_lines.extend(processed_batch)

                    processed_rows += processed
                    skip_rows += skipped
                    total_characters += chars

                # 返回结果（json串使用utf-8编码）
                content = '\n'.join(jsonl_lines).encode('utf-8')

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content=content
                )

            except InvalidFileException as e:
                logger.error(f"文件格式错误: {str(e)}")
                raise ValueError("文件格式错误")

            except ValueError:
                raise
            except Exception as e:
                logger.error(f"文件转换服务异常: {str(e)}")
                raise ValueError("文件解析服务异常")

        # 在线程池中执行同步解析
        current_loop = asyncio.get_running_loop()
        result = await current_loop.run_in_executor(self.executor, _parse_xlsx_sync)

        logger.info(
            f"文件总共处理：{result.total_rows}行\n其中：{result.skip_rows}行有误被跳过\n剩余{result.processed_rows}行转换成功"
        )
        return result

    async def analyze_csv_content(self, file_content: bytes) -> TextGenerationParseResult:
        """
        CSV文件转化为jsonl（业务数据集版本）
        将每一行转换为一个JSON对象，列名作为键，不验证对象内容

        Args:
            file_content: CSV文件的字节内容

        Returns:
            转换后的jsonl格式字节内容
        """

        def _parse_csv_sync() -> TextGenerationParseResult:
            """同步解析CSV文件（在线程池中执行）"""
            try:
                # 尝试多种编码
                encodings_to_try = ['utf-8', 'gbk', 'gb18030', 'latin-1', 'cp1252']
                csv_content = None
                encoding_used = None

                for encoding in encodings_to_try:
                    try:
                        csv_content = file_content.decode(encoding)
                        encoding_used = encoding
                        break
                    except UnicodeDecodeError:
                        continue

                if csv_content is None:
                    logger.error("文件编码不支持，请使用UTF-8编码")
                    raise ValueError("文件格式错误")

                # 使用csv模块解析CSV内容
                file_stream = io.StringIO(csv_content)
                csv_reader = csv.reader(file_stream)

                # 读取表头
                try:
                    headers = next(csv_reader)
                    headers = [h.strip() if h else f"column_{i}" for i, h in enumerate(headers)]
                except StopIteration:
                    raise ValueError("无有效数据")

                # 流式处理数据行
                jsonl_lines = []
                total_rows = 0
                processed_rows = 0
                skip_rows = 0
                total_characters = 0

                # 批量处理行数据
                batch = []
                for row in csv_reader:
                    total_rows += 1
                    batch.append(row)

                    # 凑够一批再统一处理
                    if len(batch) >= self.chunk_size:
                        processed_batch, processed, skipped, chars = self._process_csv_batch(batch, headers)
                        jsonl_lines.extend(processed_batch)
                        processed_rows += processed
                        skip_rows += skipped
                        total_characters += chars
                        batch = []

                # 处理最后凑不够一批的数据
                if batch:
                    processed_batch, processed, skipped, chars = self._process_csv_batch(batch, headers)
                    jsonl_lines.extend(processed_batch)
                    processed_rows += processed
                    skip_rows += skipped
                    total_characters += chars

                if total_rows == 0:
                    raise ValueError("无有效数据")

                # 返回结果（json串使用utf-8编码）
                content = '\n'.join(jsonl_lines).encode('utf-8')

                return TextGenerationParseResult(
                    total_rows=total_rows,
                    processed_rows=processed_rows,
                    skip_rows=skip_rows,
                    total_characters=total_characters,
                    convert_file_content=content
                )

            except ValueError:
                raise
            except Exception as e:
                logger.error(f"CSV文件解析失败: {str(e)}")
                raise ValueError("文件解析服务异常")

        # 在线程池中执行同步解析
        current_loop = asyncio.get_running_loop()
        result = await current_loop.run_in_executor(self.executor, _parse_csv_sync)

        logger.info(
            f"业务数据集CSV文件总共处理：{result.total_rows}行\n其中：{result.skip_rows}行有误被跳过\n剩余{result.processed_rows}行转换成功"
        )
        return result

class ImageUnderstandingDatasetFileParser:
    """图像理解数据集文件解析器"""
    def __init__(self):
        pass

    @staticmethod
    def _normalize_zip_member_path(file_name: str) -> str:
        normalized = (file_name or "").replace('\\', '/').strip('/')
        while normalized.startswith('./'):
            normalized = normalized[2:]
        return normalized

    @classmethod
    def _image_relative_path_from_zip_member(cls, file_name: str) -> Optional[str]:
        normalized = cls._normalize_zip_member_path(file_name)
        parts = [part for part in normalized.split('/') if part]
        for index, part in enumerate(parts[:-1]):
            if part.lower() == 'images':
                relative_parts = parts[index + 1:]
                if relative_parts:
                    return '/'.join(relative_parts)
        return None

    @classmethod
    def _image_relative_path_from_sibling_images_dir(cls, file_name: str, dataset_dir: str) -> Optional[str]:
        normalized = cls._normalize_zip_member_path(file_name)
        dataset_dir = cls._normalize_zip_member_path(dataset_dir)
        expected_prefix = f"{dataset_dir}/images/" if dataset_dir else "images/"
        if normalized.lower().startswith(expected_prefix.lower()):
            return normalized[len(expected_prefix):]
        return None

    @classmethod
    def _image_reference_candidates(cls, image_name: str) -> set[str]:
        normalized = cls._normalize_zip_member_path(image_name)
        candidates = {normalized, os.path.basename(normalized)}
        relative_path = cls._image_relative_path_from_zip_member(normalized)
        if relative_path:
            candidates.add(relative_path)
            candidates.add(os.path.basename(relative_path))
        return {candidate for candidate in candidates if candidate}

    async def process_image_understanding_file(
            self,
            file_content: bytes,
            dataset_format: Optional[str] = None,
            training_method_type: Optional[str] = None,
    ) -> ImageUnderstandingParseResult:
        """
        处理图像理解数据集的zip文件

        Args:
            file_content: zip文件的字节内容

        Returns:
            ImageUnderstandingParseResult: 处理结果，包含jsonl内容和图片字典
        """
        images = {}
        jsonl_content = None
        data_jsonl_files = []

        try:
            # 创建临时文件来解压zip
            with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
                tmp_file.write(file_content)
                tmp_file_path = tmp_file.name

            try:
                # 解压zip文件
                with zipfile.ZipFile(tmp_file_path, 'r') as zip_ref:
                    # 获取所有文件列表
                    file_list = zip_ref.namelist()

                    # 查找 .jsonl数据集 文件
                    for file_name in file_list:
                        normalized_file_name = self._normalize_zip_member_path(file_name)
                        # 忽略目录
                        if (file_name or '').replace('\\', '/').endswith('/'):
                            continue

                        # 查找 xxxx.jsonl 文件（不区分大小写）
                        basename = os.path.basename(normalized_file_name).lower()
                        if basename.endswith('.jsonl'):
                            data_jsonl_files.append(file_name)

                    is_image_prompt = dataset_format == DatasetFormat.IMAGE_PROMPT.value

                    # 验证 .jsonl数据集 文件。图像生成 image-prompt 允许未标注素材包仅包含 images/。
                    if len(data_jsonl_files) == 0 and not is_image_prompt:
                        logger.error("zip文件中未找到.jsonl数据集文件")
                        raise ValueError("文件格式错误")
                    if len(data_jsonl_files) > 1:
                        logger.error(f"zip文件中包含多个.jsonl数据集文件: {data_jsonl_files}")
                        raise ValueError("文件格式错误")

                    dataset_dir = ""
                    if data_jsonl_files:
                        # 读取 .jsonl数据集 文件内容
                        data_jsonl_file = data_jsonl_files[0]
                        jsonl_content = zip_ref.read(data_jsonl_file)
                        dataset_dir = os.path.dirname(self._normalize_zip_member_path(data_jsonl_file)).replace('\\', '/')

                    # 查找 images 文件夹。有 data.jsonl 时要求与其同级；无 data.jsonl 时使用 zip 中的 images/。
                    images_folder_found = False
                    for file_name in file_list:
                        image_relative_path = (
                            self._image_relative_path_from_sibling_images_dir(file_name, dataset_dir)
                            if data_jsonl_files
                            else self._image_relative_path_from_zip_member(file_name)
                        )
                        # 检查是否是 data.jsonl 同级 images 文件夹下的文件，兼容 xxx/images/1.jpg 等多层目录
                        if image_relative_path and not (file_name or '').replace('\\', '/').endswith('/'):
                            images_folder_found = True
                            image_name = os.path.basename(image_relative_path)
                            # 只处理图片文件
                            if image_name.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
                                images[image_name] = zip_ref.read(file_name)

                    if not images_folder_found:
                        logger.error("zip文件中未找到images文件夹")
                        raise ValueError("文件格式错误")

                    if len(images) == 0:
                        logger.error("images文件夹中未找到任何图片文件")
                        raise ValueError("文件格式错误")

                    if is_image_prompt and not jsonl_content:
                        lines = [
                            json.dumps(
                                {"prompt": "", "images": [f"images/{image_name}"]},
                                ensure_ascii=False,
                            )
                            for image_name in sorted(images.keys())
                        ]
                        jsonl_content = ("\n".join(lines) + "\n").encode("utf-8")

            finally:
                # 清理临时文件
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)

            # 验证 jsonl 内容格式
            if dataset_format == DatasetFormat.IMAGE_PROMPT.value and data_jsonl_files:
                total_samples, total_characters = await validate_image_generation_jsonl_content(jsonl_content, images)
            elif dataset_format == DatasetFormat.IMAGE_PROMPT.value:
                total_samples = len(images)
                total_characters = 0
            else:
                total_samples, total_characters = await self.analyze_jsonl_content(
                    jsonl_content,
                    images,
                    dataset_format=dataset_format,
                    training_method_type=training_method_type,
                )

            return ImageUnderstandingParseResult(
                jsonl_content=jsonl_content,
                images=images,
                total_samples=total_samples,
                total_characters=total_characters
            )

        except zipfile.BadZipFile as e:
            logger.error(f"无效的zip文件格式: {str(e)}")
            raise ValueError("文件格式错误")
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"处理zip文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    async def analyze_jsonl_content(
            self,
            jsonl_content: bytes,
            images: Dict[str, bytes],
            dataset_format: Optional[str] = None,
            training_method_type: Optional[str] = None,
    ) -> Tuple[int, int]:
        """
        验证图像理解数据集的jsonl格式

        Args:
            jsonl_content: jsonl文件内容
            images: 图片字典

        Returns:
            Tuple[int, int]: (总样本数, 总字符数)
        """
        total_samples = 0
        total_characters = 0

        try:
            content_str = jsonl_content.decode('utf-8')
            lines = content_str.splitlines()

            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue

                # 跳过注释行（支持 # 前面有空格的情况）
                if line.lstrip().startswith('#'):
                    continue

                try:
                    # 解析JSON
                    parsed_data = json.loads(line)

                    if not isinstance(parsed_data, dict):
                        logger.error(f"第{line_num}行：数据应是JSON对象格式")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    is_grpo_format = (
                        dataset_format == DatasetFormat.GRPO.value
                        or training_method_type == TrainingMethodType.GRPO.value
                        or (
                            dataset_format is None
                            and training_method_type is None
                            and "prompt" in parsed_data
                            and "reward_model" in parsed_data
                        )
                    )
                    if is_grpo_format:
                        self._validate_grpo_image_item(parsed_data, line_num, images)
                        total_samples += 1
                        total_characters += len(line)
                        continue

                    # 验证 messages 字段
                    if 'messages' not in parsed_data:
                        logger.error(f"第{line_num}行：缺少messages字段")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    messages = parsed_data['messages']
                    if not isinstance(messages, list) or len(messages) == 0:
                        logger.error(f"第{line_num}行：messages必须是非空数组")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 验证 images 字段（如果存在）
                    images_list = parsed_data.get('images', [])
                    if not isinstance(images_list, list):
                        logger.error(f"第{line_num}行：images字段必须是数组")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 验证每个 message
                    expected_role = 'user'  # 第一个消息应该是 user
                    for msg_idx, message in enumerate(messages):
                        if not isinstance(message, dict):
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：必须是对象格式")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证 role 字段
                        if 'role' not in message:
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：缺少role字段")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        role = message['role']
                        if role not in ['user', 'assistant', 'system']:
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：role必须是user、assistant或system")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证 content 字段
                        if 'content' not in message:
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：缺少content字段")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 除开system之外，其他message中，content内容不允许为空
                        content = message['content']
                        if role != 'system' and (not isinstance(content, str) or not content.strip()):
                            logger.error(f"第{line_num}行第{msg_idx + 1}个message：content不能为空")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                        # 验证 role 交替出现（system 不影响交替规则）
                        if role != 'system':
                            if role != expected_role:
                                logger.error(f"第{line_num}行：messages中role必须交替出现")
                                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                            # 切换期望的角色
                            expected_role = 'assistant' if role == 'user' else 'user'

                    # 验证 <image> 标签数量与 images 数组长度匹配
                    image_tag_count = 0
                    for message in messages:
                        content = message.get('content', '')
                        if isinstance(content, str):
                            image_tag_count += content.count('<image>')

                    if image_tag_count != len(images_list):
                        logger.error(f"第{line_num}行：<image>标签数量必须等于images数组长度")
                        raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 验证 images 数组中的图片文件名是否存在于 zip 文件中
                    image_names = set(images.keys())
                    for img_name in images_list:
                        if not isinstance(img_name, str):
                            logger.error(f"第{line_num}行：images数组中的元素必须是字符串")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                        if not (self._image_reference_candidates(img_name) & image_names):
                            logger.error(f"第{line_num}行：images中引用的图片在zip中不存在: {img_name}")
                            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

                    # 计算样本数和字符数
                    total_samples += 1
                    total_characters += len(line)

                except json.JSONDecodeError as e:
                    logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if total_samples == 0:
                raise ValueError("无有效数据")

            return total_samples, total_characters

        except UnicodeDecodeError as e:
            logger.error(f"文件编码错误: {str(e)}")
            raise ValueError("文件格式错误")

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"解析文件失败: {str(e)}")
            raise ValueError("文件解析服务异常")

    def _validate_grpo_image_item(self, item: Dict[str, Any], line_num: int, images: Dict[str, bytes]) -> None:
        TextGenerationDatasetFileParser._validate_grpo_item(item, line_num)

        images_list = item.get("images", [])
        if not isinstance(images_list, list):
            logger.error(f"第{line_num}行：images字段必须是数组")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        image_tag_count = 0
        for message in item.get("prompt", []):
            content = message.get("content", "")
            if isinstance(content, str):
                image_tag_count += content.count("<image>")

        if image_tag_count != len(images_list):
            logger.error(f"第{line_num}行：prompt中的<image>标签数量必须等于images数组长度")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
        if image_tag_count == 0:
            logger.error(f"第{line_num}行：图像理解GRPO样本必须包含至少一个<image>标签")
            raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

        for img_name in images_list:
            if not isinstance(img_name, str) or not img_name.strip():
                logger.error(f"第{line_num}行：images数组中的元素必须是非空字符串")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
            if img_name.replace("images/", "") not in images:
                logger.error(f"第{line_num}行：images中引用的图片在zip中不存在: {img_name}")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

async def analyze_text_generation_dataset_file_content(
    file_content: bytes,
    file_type: str,
    dataset_format: str,
    training_method_type: str = TrainingMethodType.SFT.value,
) -> TextGenerationParseResult:
    """
    文本生成文件解析入口函数

    Args:
        file_content: 原始文件字节内容
        file_type: 原始文件文件格式
        dataset_format: 数据集/结果集格式

    Returns:
        解析结果
    """
    # 每次创建新实例，完全隔离
    parser = TextGenerationDatasetFileParser(max_workers=2, chunk_size=1000)
    try:
        return await parser.process_text_generation_file(
            file_content,
            file_type,
            dataset_format,
            training_method_type,
        )
    finally:
        parser.executor.shutdown(wait=False)  # 清理资源

async def analyze_business_dataset_file_content(file_content: bytes, file_type: str) -> TextGenerationParseResult:
    """
    业务数据集文件解析入口函数

    Args:
        file_content: 原始文件字节内容
        file_type: 原始文件文件格式

    Returns:
        解析结果
    """
    # 每次创建新实例，完全隔离
    parser = BusinessDatasetFileParser(max_workers=2, chunk_size=1000)
    try:
        return await parser.process_business_file(file_content, file_type)
    finally:
        parser.executor.shutdown(wait=False)  # 清理资源

async def analyze_image_understanding_dataset_file_content(
        file_content: bytes,
        file_type: str,
        dataset_format: Optional[str] = None,
        training_method_type: Optional[str] = None,
) -> ImageUnderstandingParseResult:
    """
    图像理解文件解析入口函数

    Args:
        file_content: 原始文件字节内容
        file_type: 原始文件文件格式

    Returns:
        解析结果
        jsonl_content: data.jsonl 文件内容
        images: 图片字典 {图片文件名: 图片内容}
        total_samples: 总样本数
        total_characters: 总字符数
    """
    if file_type != TrainingDatasetUploadTypeCategory.ZIP_TYPE:
        logger.error(f"图像理解数据集仅支持zip格式，当前为: {file_type}")
        raise ValueError("文件格式错误")
    parser = ImageUnderstandingDatasetFileParser()
    return await parser.process_image_understanding_file(
        file_content,
        dataset_format=dataset_format,
        training_method_type=training_method_type,
    )

async def analyze_image_generation_dataset_file_content(
        file_content: bytes,
        file_type: str,
) -> ImageUnderstandingParseResult:
    """
    图像生成 image-prompt zip 解析入口。

    有标注 zip 结构：data.jsonl + images/。
    未标注 zip 结构：仅 images/，系统会按图片生成空 prompt 样本。
    data.jsonl 每行包含 prompt、images[]，可选 negative_prompt、metadata。
    """
    if file_type != TrainingDatasetUploadTypeCategory.ZIP_TYPE:
        logger.error(f"图像生成数据集仅支持zip格式，当前为: {file_type}")
        raise ValueError("文件格式错误")

    parser = ImageUnderstandingDatasetFileParser()
    return await parser.process_image_understanding_file(
        file_content,
        dataset_format=DatasetFormat.IMAGE_PROMPT.value,
        training_method_type=TrainingMethodType.SFT.value,
    )


async def validate_image_generation_jsonl_content(
        jsonl_content: bytes,
        images: Dict[str, bytes],
) -> Tuple[int, int]:
    """校验 image-prompt 格式：prompt、images[] 必填，negative_prompt/metadata 可选。"""
    total_samples = 0
    total_characters = 0
    try:
        content_str = jsonl_content.decode('utf-8')
        image_names = set(images.keys())
        for line_num, line in enumerate(content_str.splitlines(), 1):
            line = line.strip()
            if not line or line.lstrip().startswith('#'):
                continue

            try:
                parsed_data = json.loads(line)
            except json.JSONDecodeError as e:
                logger.error(f"第{line_num}行JSON格式错误: {str(e)}")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            if not isinstance(parsed_data, dict):
                logger.error(f"第{line_num}行：数据应是JSON对象格式")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            prompt = parsed_data.get('prompt')
            if not isinstance(prompt, str) or not prompt.strip():
                logger.error(f"第{line_num}行：prompt不能为空")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            images_list = parsed_data.get('images')
            if not isinstance(images_list, list) or len(images_list) == 0:
                logger.error(f"第{line_num}行：images必须是非空数组")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            negative_prompt = parsed_data.get('negative_prompt')
            if negative_prompt is not None and not isinstance(negative_prompt, str):
                logger.error(f"第{line_num}行：negative_prompt必须是字符串")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            metadata = parsed_data.get('metadata')
            if metadata is not None and not isinstance(metadata, dict):
                logger.error(f"第{line_num}行：metadata必须是对象")
                raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            for img_name in images_list:
                if not isinstance(img_name, str) or not img_name.strip():
                    logger.error(f"第{line_num}行：images数组中的元素必须是非空字符串")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")
                if not (ImageUnderstandingDatasetFileParser._image_reference_candidates(img_name) & image_names):
                    logger.error(f"第{line_num}行：images中引用的图片在zip中不存在: {img_name}")
                    raise ValueError(f"第 {line_num} 个样本：字段/格式错误")

            total_samples += 1
            total_characters += len(line)

        if total_samples == 0:
            raise ValueError("无有效数据")

        return total_samples, total_characters
    except UnicodeDecodeError as e:
        logger.error(f"文件编码错误: {str(e)}")
        raise ValueError("文件格式错误")

# ========= 其他基础公共方法 ==========

def generate_filenames(base_filename: str) -> List[str]:
    """
    生成所有可能的文件名

    Args:
        base_filename: 基础文件名

    Returns:
        所有可能的文件名列表
    """
    # 若文件名称带有后缀，去除后缀
    if '.' in base_filename:
        base_filename = base_filename.split('.')[0]
    else:
        base_filename = base_filename

    # 根据配置生成所有支持的文件类型
    filenames = []
    for config in FILE_TYPE_CONFIG.values():
        filename = f"{base_filename}{config['file_suffix']}"
        filenames.append(filename)

    # 添加索引文件类型（该文件类型并不参与文件下载，所以无需添加相对于的下载配置）
    index_filename = f"{base_filename}_index.cache"
    filenames.append(index_filename)

    return filenames

def get_content_type_by_extension(file_extension: str) -> str:
    """
    根据文件扩展名获取对应的 Content-Type

    Args:
        file_extension: 文件扩展名（如 '.jsonl', '.xlsx'）

    Returns:
        Content-Type 字符串，如果未找到则返回 'application/octet-stream'
    """
    content_type_map = {
        '.jsonl': 'application/x-ndjson',
        '.json': 'application/json',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.zip': 'application/zip',
        '.csv': 'text/csv'
    }
    return content_type_map.get(file_extension.lower(), 'application/octet-stream')

# ========= 数据集文件处理工具函数 =========

def format_file_size(size_bytes: int) -> str:
    """格式化文件大小（B/KB/MB/GB/TB）"""
    if size_bytes == 0:
        return "0 B"

    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0
    size = float(size_bytes)

    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1

    if size >= 100:
        return f"{size:.0f} {units[unit_index]}"
    elif size >= 10:
        return f"{size:.1f} {units[unit_index]}"
    else:
        return f"{size:.2f} {units[unit_index]}"
# ------------- 数据集路径相关函数 -----------------------

def generate_base_path(namespace: str, usage: str) -> str:
    """
    生成数据集保存基础路径

    图像理解数据集基础路径格式：/{namespace}/{usage}/datasets/imageUnderstanding/
    文本生成数据集路径格式：/{namespace}/{usage}/datasets/
    业务数据集路径格式：/{namespace}/{usage}/datasets/
    """
    if usage == DatasetUsage.TRAINING.value:
        base_path = StoragePath.REGISTERED_TRAINED_DATASETS.format_storage_path(namespace=namespace)
    elif usage == DatasetUsage.VALIDATION.value:
        base_path = StoragePath.REGISTERED_VALIDATION_DATASETS.format_storage_path(namespace=namespace)
    elif usage == DatasetUsage.TEST.value:
        base_path = StoragePath.REGISTERED_TEST_DATASETS.format_storage_path(namespace=namespace)
    elif usage == DatasetUsage.BUSINESS_TRAINING.value:
        base_path = StoragePath.REGISTERED_BUSINESS_TRAINING_DATASETS.format_storage_path(namespace=namespace)
    elif usage == DatasetUsage.BUSINESS_TEST.value:
        base_path = StoragePath.REGISTERED_BUSINESS_TEST_DATASETS.format_storage_path(namespace=namespace)
    else:
        raise ValueError(f"不支持当前所选的数据集类型：{usage}")

    return base_path


def generate_dataset_path(
    namespace: str,
    dataset_name: str,
    version: str,
    file_extension: str,
    usage: str,
    dataset_type: Optional[str] = None
) -> str:
    """
    生成数据集文件保存路径

    图像理解数据集路径格式：/{namespace}/{usage}/datasets/imageUnderstanding/{dataset_name}_{version}/data.jsonl
    文本生成数据集路径格式：/{namespace}/{usage}/datasets/{dataset_name}_{version}.jsonl
    业务数据集逻辑格式：/{namespace}/{usage}/datasets/{dataset_name}_{version}/data.jsonl
    """

    # 先生成基础目录
    base_path = generate_base_path(namespace, usage)

    # 图像类数据集需要额外添加类型子目录和版本目录
    if dataset_type in (TrainingTypeCategory.IMAGE_UNDERSTANDING.value, TrainingTypeCategory.IMAGE_GENERATION.value):
        image_dir = 'imageGeneration' if dataset_type == TrainingTypeCategory.IMAGE_GENERATION.value else 'imageUnderstanding'
        base_path = os.path.join(base_path, image_dir)
        dataset_dir = os.path.join(base_path, f"{dataset_name}_{version}")
        filename = f"data.{file_extension}"
        return os.path.join(dataset_dir, filename).replace('\\', '/')

    elif dataset_type == TrainingTypeCategory.TEXT_GENERATION.value:
        # 文本生成数据集
        filename = f"{dataset_name}_{version}.{file_extension}"
        return os.path.join(base_path, filename).replace('\\', '/')  # 统一使用 / 分隔符

    elif dataset_type == TrainingTypeCategory.BUSINESS.value:
        # 业务数据集
        filename = f"{dataset_name}_{version}.{file_extension}"
        return os.path.join(base_path, filename).replace('\\', '/')  # 统一使用 / 分隔符

    else:
        raise ValueError(f"不支持当前所选的数据集格式：{dataset_type}")


# ----------------- 图像理解数据集相关函数 ----------------

def generate_image_folder_path(
    namespace: str,
    dataset_name: str,
    version: str,
    usage: str,
    dataset_type: str = TrainingTypeCategory.IMAGE_UNDERSTANDING.value,
) -> str:
    """
    生成图像理解数据集的图片文件夹路径

    路径格式：/{namespace}/{usage}/datasets/imageUnderstanding/{dataset_name}_{version}/images

    Args:
        namespace: 命名空间
        dataset_name: 数据集名称
        version: 版本号
        usage: 数据集用途

    Returns:
        str: 图片文件夹路径
    """
    # 先生成基础目录
    base_path = generate_base_path(namespace, usage)

    image_dir = 'imageGeneration' if dataset_type == TrainingTypeCategory.IMAGE_GENERATION.value else 'imageUnderstanding'
    base_path = os.path.join(base_path, image_dir)
    dataset_dir = os.path.join(base_path, f"{dataset_name}_{version}")
    return os.path.join(dataset_dir, 'images').replace('\\', '/')

def generate_image_dataset_directory_path(
        namespace: str,
        dataset_name: str,
        version: str,
        usage: str,
        dataset_type: str = TrainingTypeCategory.IMAGE_UNDERSTANDING.value,
) -> str:
    """
    生成图像理解数据集的目录路径（用于删除整个数据集目录）

    路径格式：/{namespace}/{usage}/datasets/imageUnderstanding/{dataset_name}_{version}/

    Args:
        namespace: 命名空间
        dataset_name: 数据集名称
        version: 版本号
        usage: 数据集用途

    Returns:
        str: 数据集目录路径
    """
    # 先生成基础目录
    # 这里调用file_parser内的公共方法
    base_path = generate_base_path(namespace, usage)

    image_dir = 'imageGeneration' if dataset_type == TrainingTypeCategory.IMAGE_GENERATION.value else 'imageUnderstanding'
    base_path = os.path.join(base_path, image_dir)
    return os.path.join(base_path, f"{dataset_name}_{version}").replace('\\', '/')

# ------------------ JFS操作相关函数 --------------------

async def get_juicefs_client(storage_service: StorageService) -> Any:
    """获取JuiceFS客户端（通过注入的StorageService）"""
    return await storage_service.JUICEFS_CLIENT()


async def save_file_content_to_jfs(storage_service: StorageService, file_content: bytes, dataset_path: str) -> None:
    """保存文件内容到JuiceFS"""
    try:
        jfs = await get_juicefs_client(storage_service)
        # 确保目录存在
        remote_dir = os.path.dirname(dataset_path)
        if remote_dir and not jfs.exists(remote_dir):
            jfs.makedirs(remote_dir, exist_ok=True)
            logger.info(f"创建JuiceFS目录: {remote_dir}")
        # 写入文件
        with jfs.open(dataset_path, 'wb') as f:
            f.write(file_content)
        logger.info(f"文件保存成功: {dataset_path}")
    except Exception as e:
        logger.error(f"JuiceFS保存失败: {str(e)}")
        raise


# ----------------- 文件索引相关函数 ---------------------
def get_index_cache_path(dataset_path: str) -> str:
    """
    获取索引缓存文件路径 - 用于优化大数据集的随机访问

    优化原理：
    1. 为每个数据集文件生成对应的索引缓存文件
    2. 索引文件存储每行的偏移量和长度信息
    3. 后续访问可以直接跳转到指定位置，无需顺序扫描

    Args:
        dataset_path: 数据集文件路径

    Returns:
        str: 索引缓存文件路径
    """
    # 移除文件扩展名，添加索引缓存后缀
    base_path = dataset_path.rsplit('.', 1)[0]
    return f"{base_path}_index.cache"


async def build_line_index(executor: ThreadPoolExecutor, jfs, dataset_path: str) -> List[LineIndex]:
    """
    构建文件行索引

    优化原理：
    1. 一次性扫描文件，记录每行的偏移量和长度
    2. 后续访问可以直接跳转到任意行，时间复杂度O(1)
    3. 原方法需要顺序扫描到目标行，时间复杂度O(n)

    Args:
        executor: 线程池执行器
        jfs: JuiceFS客户端
        dataset_path: 数据集文件路径

    Returns:
        List[LineIndex]: 行索引列表
    """
    logger.info(f"开始构建索引: {dataset_path}")

    def _build_index_sync():
        """同步构建索引的函数，在线程池中执行"""
        line_indices = []
        current_offset = 0
        line_number = 0

        """
        使用二进制的方式读取文件
        原因：
        - 1. 不同相同中，换行符所占字节可能不相同，直接在偏移量后面加1，可能会导致偏移量错误
        - 2. 在读取索引时，seek方法读取的是直接偏移量，而不是字符偏移量
        - 3. 二进制模式读取文本比文本模式更高效，省去了编码和解码的时间
        """
        with jfs.open(dataset_path, 'rb') as f:

            while True:
                line_start = current_offset
                line_bytes = f.readline()

                if not line_bytes:
                    break

                # 对当前读取的行进行解码，主要是为了判断当前行是否为注释或空行
                line_str = line_bytes.decode('utf-8', errors='ignore')

                # 跳过空行和注释行
                if not line_str.strip() or line_str.strip().startswith('#'):
                    current_offset += len(line_bytes)
                    continue

                # 记录每个json对象有效的索引行信息
                line_indices.append(LineIndex(
                    line_number=line_number, # 行信息
                    file_offset=line_start, # 偏移量
                    line_length=len(line_bytes) # 每行读取字符数
                ))

                # 更新偏移量（行内容 + 换行符）
                current_offset += len(line_bytes)
                line_number += 1

        logger.info(f"索引构建完成，总有效行数: {len(line_indices)}")
        return line_indices

    # 在线程池中执行索引构建，避免阻塞
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(executor, _build_index_sync)


async def load_or_build_index(executor: ThreadPoolExecutor, jfs, dataset_path: str) -> List[LineIndex]:
    """
    加载或构建行索引

    优化原理：
    1. 检查是否存在索引缓存文件
    2. 如果存在，验证文件哈希是否匹配（判断文件是否被修改）
    3. 如果匹配，直接加载缓存的索引
    4. 如果不匹配或不存在，重建索引并缓存

    Args:
        executor: 现场池执行器
        jfs: JuiceFS客户端
        dataset_path: 数据集文件路径

    Returns:
        List[LineIndex]: 行索引列表
    """
    # 先获取索引文件路径
    index_cache_path = get_index_cache_path(dataset_path)

    # 检查索引缓存是否存在
    if jfs.exists(index_cache_path):
        logger.info(f"使用缓存的索引: {index_cache_path}")
        try:
            # 读取缓存的索引
            with jfs.open(index_cache_path, 'rb') as f:
                cached_data = pickle.load(f)

            return cached_data['indices']

        except Exception as e:
            logger.warning(f"加载索引缓存失败: {str(e)}，将重建索引")

    # 索引不存在，构建新索引
    logger.info(f"构建新索引: {dataset_path}")
    indices = await build_line_index(executor, jfs, dataset_path)

    # 保存索引缓存
    try:
        cache_data = {
            'indices': indices
        }

        with jfs.open(index_cache_path, 'wb') as f:
            pickle.dump(cache_data, f)

        logger.info(f"索引缓存已保存: {index_cache_path}")
    except Exception as e:
        logger.warning(f"保存索引缓存失败: {str(e)}")

    return indices


def _collect_metadata_fields_from_data(data: Any, fields: set[str], prefix: str = "") -> None:
    """递归收集 JSON 对象字段，嵌套字段使用点号路径。"""
    if isinstance(data, dict):
        for key, value in data.items():
            field_name = f"{prefix}.{key}" if prefix else str(key)
            fields.add(field_name)
            if isinstance(value, (dict, list)):
                _collect_metadata_fields_from_data(value, fields, field_name)
    elif isinstance(data, list):
        for item in data:
            if prefix in {"messages", "conversations", "dialogue"} and isinstance(item, dict):
                role = item.get("role") or item.get("from") or item.get("speaker")
                if isinstance(role, str) and role:
                    fields.add(f"{prefix}.{role}")
            if isinstance(item, (dict, list)):
                _collect_metadata_fields_from_data(item, fields, prefix)


@dataclass
class MetadataFieldsCollectResult:
    fields: List[str]
    total_lines: int
    invalid_lines: int


class MetadataFieldsJsonlParseError(ValueError):
    def __init__(self, line_number: int, message: str):
        self.line_number = line_number
        super().__init__(f"第 {line_number} 行不是合法 JSONL: {message}")


def _collect_metadata_fields_from_jsonl_line(line: str, fields: set[str]) -> None:
    parsed_data = json.loads(line)
    _collect_metadata_fields_from_data(parsed_data, fields)


def _collect_metadata_fields_from_jsonl_line_safe(line: str, fields: set[str]) -> bool:
    try:
        _collect_metadata_fields_from_jsonl_line(line, fields)
    except json.JSONDecodeError:
        return False
    return True


def _collect_metadata_fields_from_jsonl_iterable(
    jsonl_lines: Iterable[str],
    fields: set[str],
    strict: bool = False,
) -> Tuple[int, int]:
    """从已规范化的 JSONL 行中增量收集字段。"""
    total_lines = 0
    invalid_lines = 0
    for line_number, line in enumerate(jsonl_lines, start=1):
        if not line or not line.strip() or line.lstrip().startswith("#"):
            continue
        total_lines += 1
        if strict:
            try:
                _collect_metadata_fields_from_jsonl_line(line, fields)
            except json.JSONDecodeError as exc:
                raise MetadataFieldsJsonlParseError(line_number, exc.msg) from exc
            continue
        if not _collect_metadata_fields_from_jsonl_line_safe(line, fields):
            invalid_lines += 1
    return total_lines, invalid_lines


def _prepare_jsonl_lines_and_collect_metadata(jsonl_lines: Iterable[str], fields: set[str]) -> List[str]:
    """过滤有效 JSONL 行，并在加入合并列表前同步收集字段。"""
    valid_lines = []
    for line in jsonl_lines:
        if not line or not line.strip() or line.lstrip().startswith("#"):
            continue
        valid_lines.append(line)
        _collect_metadata_fields_from_jsonl_line(line, fields)
    return valid_lines


def collect_metadata_fields_from_jsonl_iterable_with_stats(
    jsonl_lines: Iterable[str],
    strict: bool = False,
) -> MetadataFieldsCollectResult:
    """从已规范化的 JSONL 行中收集字段，并返回解析统计信息。"""
    fields: set[str] = set()
    total_lines, invalid_lines = _collect_metadata_fields_from_jsonl_iterable(
        jsonl_lines,
        fields,
        strict=strict,
    )
    return MetadataFieldsCollectResult(
        fields=sorted(fields),
        total_lines=total_lines,
        invalid_lines=invalid_lines,
    )


def collect_metadata_fields_from_jsonl_iterable(jsonl_lines: Iterable[str]) -> List[str]:
    """从已规范化的 JSONL 行中收集完整字段列表。"""
    result = collect_metadata_fields_from_jsonl_iterable_with_stats(jsonl_lines)
    if result.invalid_lines:
        logger.warning(
            "metadata_fields 收集跳过无法解析的 JSONL 行: "
            f"invalid_lines={result.invalid_lines}, total_lines={result.total_lines}"
        )
    return result.fields


def collect_metadata_fields_from_jsonl_lines(jsonl_lines: List[str]) -> List[str]:
    """从已规范化的 JSONL 行列表中收集完整字段列表。"""
    return collect_metadata_fields_from_jsonl_iterable(jsonl_lines)


# ----------------- 文件内容解析入口函数 -----------------
async def analyze_save_dataset_file_multi(
    files: List[UploadFile],
    dataset_type: TrainingTypeCategory,
    dataset_format: DatasetFormat,
    training_method_type: TrainingMethodType,
    namespace: str,
    name: str,
    version: str,
    usage: DatasetUsage,
    storage_service: StorageService,
    executor: ThreadPoolExecutor,
    base_dataset_path: Optional[str] = None,
) -> dict:
    """
    解析多个数据集文件内容并合并保存（多文件）

    Args:
        files: 数据集文件列表
        dataset_type: 数据集类型（图像理解、文本生成等）
        dataset_format: 数据集格式（roled-based, prompt-response）
        training_method_type: 训练方法类型（sft, dpo, business）
        namespace: 命名空间
        name: 数据集名称
        version: 数据集版本
        usage: 数据集用途（验证、测试、训练等）
        storage_service: 存储服务
        executor: 线程池执行器

    Returns:
        dict: 包含 total_samples, total_characters, file_size_bytes, dataset_path,
              origin_dataset_path, metadata_fields
    """
    if not files or len(files) == 0:
        raise HTTPException(status_code=500, detail="文件解析失败：无有效文件")
    
    jfs = await get_juicefs_client(storage_service)

    # 合并所有文件的数据
    all_jsonl_lines = []
    total_samples = 0
    total_characters = 0
    total_file_size_bytes = 0
    metadata_field_set: set[str] = set()
    origin_dataset_paths = []  # 记录所有原始文件路径（用于兼容）

    if base_dataset_path:
        if not jfs.exists(base_dataset_path):
            raise HTTPException(status_code=404, detail=f"继承源文件不存在: {base_dataset_path}")
        with jfs.open(base_dataset_path, 'rb') as base_file:
            base_file_content = base_file.read()
        base_file_size_bytes = len(base_file_content)
        base_text = base_file_content.decode('utf-8')
        base_lines = _prepare_jsonl_lines_and_collect_metadata(
            (line.strip() for line in base_text.splitlines()),
            metadata_field_set
        )
        all_jsonl_lines.extend(base_lines)
        total_samples += len(base_lines)
        total_characters += sum(len(line) for line in base_lines)
        total_file_size_bytes += base_file_size_bytes
    
    # 逐个处理文件
    for index, file in enumerate(files):
        try:
            # 读取文件内容
            file_content = await file.read()
            file_size_bytes = len(file_content)
            total_file_size_bytes += file_size_bytes
            file_type = file.filename.split('.')[-1].lower()

            # 验证文件格式
            validate_dataset_upload_file_type(file_type, dataset_type)

            # 根据数据集类型处理文件
            if dataset_type in (TrainingTypeCategory.IMAGE_UNDERSTANDING, TrainingTypeCategory.IMAGE_GENERATION):
                # 图像类数据集：处理zip文件
                if dataset_type == TrainingTypeCategory.IMAGE_GENERATION:
                    zip_result = await analyze_image_generation_dataset_file_content(file_content, file_type)
                else:
                    zip_result = await analyze_image_understanding_dataset_file_content(
                        file_content,
                        file_type,
                        dataset_format.value if hasattr(dataset_format, "value") else dataset_format,
                        training_method_type.value if hasattr(training_method_type, "value") else training_method_type,
                    )
                jsonl_content = zip_result.jsonl_content
                images = zip_result.images
                parse_total_samples = zip_result.total_samples
                parse_total_characters = zip_result.total_characters

                # 解析jsonl内容为行列表
                if jsonl_content:
                    jsonl_str = jsonl_content.decode('utf-8') if isinstance(jsonl_content, bytes) else jsonl_content
                    lines = jsonl_str.strip().split('\n')
                    valid_lines = _prepare_jsonl_lines_and_collect_metadata(lines, metadata_field_set)
                    all_jsonl_lines.extend(valid_lines)
                    total_samples += parse_total_samples
                    total_characters += parse_total_characters

                image_folder_path = generate_image_folder_path(namespace, name, version, usage.value, dataset_type.value)
                image_dir = image_folder_path
                if not jfs.exists(image_dir):
                    jfs.makedirs(image_dir, exist_ok=True)

                for image_name, image_content in images.items():
                    image_path = os.path.join(image_dir, image_name).replace('\\', '/')
                    if jfs.exists(image_path):
                        with jfs.open(image_path, 'rb') as existing_image:
                            existing_content = existing_image.read()
                        if existing_content != image_content:
                            raise HTTPException(
                                status_code=400,
                                detail=f"图片名 {image_name} 已存在，但是同名图片中新图片和旧图片不一样"
                            )
                    with jfs.open(image_path, 'wb') as f:
                        f.write(image_content)
                logger.info(f"图片文件保存成功: {len(images)} 张图片保存到 {image_dir}")

            elif dataset_type == TrainingTypeCategory.TEXT_GENERATION or dataset_type == TrainingTypeCategory.BUSINESS:
                # 文本生成或业务数据集：解析文件内容
                if dataset_type == TrainingTypeCategory.TEXT_GENERATION:
                    parse_result = await analyze_text_generation_dataset_file_content(
                        file_content,
                        file_type,
                        dataset_format,
                        training_method_type.value if hasattr(training_method_type, "value") else training_method_type,
                    )
                else:
                    parse_result = await analyze_business_dataset_file_content(file_content, file_type)

                # 获取转换后的jsonl内容
                convert_file_content = parse_result.convert_file_content
                if convert_file_content:
                    jsonl_str = convert_file_content.decode('utf-8') if isinstance(convert_file_content, bytes) else convert_file_content
                    lines = jsonl_str.strip().split('\n')
                    valid_lines = _prepare_jsonl_lines_and_collect_metadata(lines, metadata_field_set)
                    all_jsonl_lines.extend(valid_lines)

                # 使用 processed_rows（转化成功样本数）与实际写入行数一致，避免被跳过的行仍被记入导致 total_samples 大于索引条数
                total_samples += parse_result.processed_rows
                total_characters += parse_result.total_characters

            else:
                raise ValueError(f"不支持的数据集类型: {dataset_type}")
        except Exception as e:
            # 文件解析方法抛出的异常统一封装为 HTTPException 500
            logger.error(f"解析数据集格式异常: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"文件 {index + 1} 解析失败:{str(e)}")
    
    # 合并所有数据并保存
    if len(all_jsonl_lines) == 0:
        raise HTTPException(status_code=500, detail="所有文件解析失败：无有效数据")

    metadata_fields = sorted(metadata_field_set)

    try:
        # 生成最终的数据集路径
        dataset_path = generate_dataset_path(
            namespace, name, version, "jsonl", usage, dataset_type
        )
        
        # 保存合并后的jsonl文件
        dataset_dir = os.path.dirname(dataset_path)
        if dataset_dir and not jfs.exists(dataset_dir):
            jfs.makedirs(dataset_dir, exist_ok=True)
        
        with jfs.open(dataset_path, 'wb') as f:
            write_jsonl_lines_in_batches(f, all_jsonl_lines)
        logger.info(f"合并后的jsonl文件保存成功: {dataset_path}, 共 {len(all_jsonl_lines)} 行")
        
        # 构建索引
        new_index_dataset_path = get_index_cache_path(dataset_path)
        logger.info(f"构建新索引: {new_index_dataset_path}")
        
        indices = await build_line_index(executor, jfs, dataset_path)
        
        # 保存索引缓存
        cache_data = {
            'indices': indices
        }
        with jfs.open(new_index_dataset_path, 'wb') as f:
            pickle.dump(cache_data, f)
        logger.info(f"索引缓存已保存: {new_index_dataset_path}")
        
        # 返回处理结果（使用第一个文件的原始路径作为兼容）
        origin_dataset_path = dataset_path  # 合并后的文件路径
        
        return {
            'total_samples': total_samples,
            'total_characters': total_characters,
            'file_size_bytes': total_file_size_bytes,
            'dataset_path': dataset_path,
            'origin_dataset_path': origin_dataset_path,
            'metadata_fields': metadata_fields
        }
    except HTTPException:
        raise
    except Exception as e:
        # 入口方法内（保存/索引等）抛出的异常
        logger.error(f"数据集保存或索引构建失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"{str(e)}")

# ----------------- 辅助函数 -----------------------
async def _read_jsonl_file_content(file_path: str, storage_service: StorageService) -> list[Any]:
    """
    读取jsonl文件中的内容，转化为list，方便后续解析

    Args:
        file_path: 数据集在jfs中的路径
        storage_service: 存储服务
    """
    # 验证数据集文件路径
    if not file_path:
        raise ValueError("数据集文件路径为空")

    # 获取 JuiceFS 客户端
    jfs = await storage_service.JUICEFS_CLIENT()

    # 验证文件是否存在
    if not jfs.exists(file_path):
        raise ValueError(f"数据集文件不存在: {file_path}")

    # 读取 JSONL 文件内容
    all_items = []
    with jfs.open(file_path, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue

            # 跳过注释行
            if line.lstrip().startswith('#'):
                continue

            try:
                parsed_data = json.loads(line)

                # 处理 prompt-response 格式（数组包装）
                if isinstance(parsed_data, list):
                    # prompt-response 格式：每行是一个数组，包含一个或多个对象
                    for item in parsed_data:
                        if isinstance(item, dict):
                            all_items.append(item)
                elif isinstance(parsed_data, dict):
                    # role-based 格式：每行是一个对象
                    all_items.append(parsed_data)
                else:
                    logger.warning(f"跳过无效的数据格式（第{line_num}行）")
                    continue

            except json.JSONDecodeError as e:
                logger.warning(f"跳过无效的JSON行 {line_num} in {file_path}: {str(e)}")
                continue

    if len(all_items) == 0:
        logger.warning(f"数据集中没有有效的数据样本: {file_path}")
        # 不再抛出异常，返回空列表，由调用方处理

    return all_items


# ----------------- 文件导出入口函数 -----------------
async def analyze_export_dataset_file_single(
        db_dataset: TrainingDataset,
        export_file_type: TrainingDatasetExportTypeCategory,
        storage_service: StorageService,
) -> bytes:
    """
    测试/验证/训练数据集多格式导出
        - 支持 文本生成role-based
        - 支持 文本生成prompt-response
        - 支持 文本生成alpaca
        - 支持 xlsx、json、jsonl

    Args:
        db_dataset: 数据集信息
        export_file_type: 需要导出的格式
        storage_service: 存储服务实例
    
    Returns:
        转换后的文件内容（bytes）
    
    转换规则：
    1. JSONL → JSON: 将每行 JSON 对象合并为 JSON 数组
    2. JSONL → XLSX:
       - prompt-response: 列结构为 prompt, response, system（可选）
       - alpaca: 列结构为 instruction, input, chosen, rejected
       - role-based: 从 messages 数组转换为表格列（单轮：system, user, assistant；多轮：system, user1, assistant1, user2, assistant2...）
       - grpo: 列结构为 data_source, prompt, ability, reward_model, extra_info

    注意：
    - prompt-response 格式的 JSONL 每行格式为 [{"prompt": "...", "response": "..."}]（数组包装）
    - alpaca 格式的 JSONL 每行格式为 {"instruction": "...", "input": "...", "chosen": "...", "rejected": "..."}
    - role-based 格式的 JSONL 每行格式为 {"messages": [...]}（标准格式）
    """
    # 先读取数据集文件内容
    if not db_dataset:
        raise ValueError("db_database不能为空")

    all_items = await _read_jsonl_file_content(db_dataset.dataset_path, storage_service)

    # 处理空数据情况：记录日志并返回对应的空文件
    if not all_items:
        logger.info(f"数据集为空，返回空文件: {db_dataset.dataset_path}")
        export_type = export_file_type.value
        
        if export_type == TrainingDatasetExportTypeCategory.JSONL_TYPE.value:
            # JSONL 格式：返回空字符串
            return b''
        
        elif export_type == TrainingDatasetExportTypeCategory.JSON_TYPE.value:
            # JSON 格式：返回空数组
            return json.dumps([], ensure_ascii=False, indent=2).encode('utf-8')
        
        elif export_type == TrainingDatasetExportTypeCategory.XLSX_TYPE.value:
            # XLSX 格式：返回只有表头的 Excel 文件
            dataset_format = db_dataset.dataset_format
            if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
                return _convert_training_prompt_response_to_xlsx([])
            elif dataset_format == DatasetFormat.ALPACA.value:
                return _convert_training_alpaca_to_xlsx([])
            elif dataset_format == DatasetFormat.GRPO.value:
                return _convert_training_grpo_to_xlsx([])
            elif dataset_format == DatasetFormat.ROLE_BASED.value:
                if db_dataset.training_method_type == TrainingMethodType.DPO.value:
                    return _convert_training_dpo_role_based_to_xlsx([])
                return _convert_training_role_based_to_xlsx([])
            else:
                # 未知格式，返回空 Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "数据集"
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                return output.getvalue()
        
        else:
            raise ValueError(f"不支持的导出格式: {export_type}")

    # 根据导出类型进行转换
    export_type = export_file_type.value
    
    if export_type == TrainingDatasetExportTypeCategory.JSONL_TYPE.value:
        # JSONL 格式：直接返回原始 JSONL 内容
        # 注意：prompt-response 格式需要保持数组包装
        dataset_format = db_dataset.dataset_format
        jsonl_lines = []
        
        for item in all_items:
            if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
                # prompt-response 格式：使用数组包装
                jsonl_line = '[' + json.dumps(item, ensure_ascii=False) + ']'
            elif dataset_format == DatasetFormat.ALPACA.value:
                jsonl_line = json.dumps(item, ensure_ascii=False)
            elif dataset_format == DatasetFormat.GRPO.value:
                jsonl_line = json.dumps(item, ensure_ascii=False)
            else:
                # role-based 格式：直接序列化
                jsonl_line = json.dumps(item, ensure_ascii=False)
            jsonl_lines.append(jsonl_line)
        
        jsonl_content = "\n".join(jsonl_lines)
        return jsonl_content.encode('utf-8')
    
    elif export_type == TrainingDatasetExportTypeCategory.JSON_TYPE.value:
        # JSON 格式：将每行 JSON 对象合并为 JSON 数组
        json_content = json.dumps(all_items, ensure_ascii=False, indent=2)
        return json_content.encode('utf-8')
    
    elif export_type == TrainingDatasetExportTypeCategory.XLSX_TYPE.value:
        # XLSX 格式：根据数据集格式进行转换
        dataset_format = db_dataset.dataset_format
        
        if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
            # prompt-response 格式转换为 XLSX
            return _convert_training_prompt_response_to_xlsx(all_items)
        
        elif dataset_format == DatasetFormat.ALPACA.value:
            return _convert_training_alpaca_to_xlsx(all_items)

        elif dataset_format == DatasetFormat.GRPO.value:
            return _convert_training_grpo_to_xlsx(all_items)

        elif dataset_format == DatasetFormat.ROLE_BASED.value:
            # role-based 格式转换为 XLSX
            if db_dataset.training_method_type == TrainingMethodType.DPO.value:
                return _convert_training_dpo_role_based_to_xlsx(all_items)
            return _convert_training_role_based_to_xlsx(all_items)

        elif dataset_format == DatasetFormat.BUSINESS.value:
            # business 格式转化为 XLSX
            return _convert_business_to_xlsx(all_items)
        
        else:
            raise ValueError(f"不支持的数据集格式: {dataset_format}，XLSX 导出仅支持 prompt-response、alpaca、role-based 和 grpo 格式")
    
    else:
        raise ValueError(f"不支持的导出格式: {export_type}")


def _convert_training_prompt_response_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """
    将 prompt-response 格式的训练数据集转换为 XLSX
    
    列结构：
    - prompt（必需）
    - response（必需）
    - system（可选）
    
    Args:
        items: 数据项列表
    
    Returns:
        XLSX 文件的字节内容
    """
    # 收集所有可能的列名
    all_columns = set()
    for item in items:
        all_columns.update(item.keys())
    
    # 定义列的顺序（优先显示常用列）
    priority_columns = ['system', 'prompt', 'response', 'chosen', 'rejected']
    ordered_columns = []
    
    # 先添加优先列（如果数据为空，至少添加默认列）
    if not items:
        # 空数据时，使用默认列
        ordered_columns = priority_columns
    else:
        # 有数据时，按实际列添加
        for col in priority_columns:
            if col in all_columns:
                ordered_columns.append(col)
                all_columns.discard(col)
        
        # 再添加其他列（按字母顺序）
        ordered_columns.extend(sorted(all_columns))
    
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "训练数据集"
    
    # 写入表头
    ws.append(ordered_columns)
    
    # 写入数据行
    for item in items:
        row = [item.get(col, '') for col in ordered_columns]
        # 处理 None 值
        row = ['' if v is None else v for v in row]
        ws.append(row)
    
    # 将工作簿保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def _convert_training_alpaca_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """将 alpaca 格式的训练数据集转换为 XLSX。"""
    all_columns = set()
    for item in items:
        all_columns.update(item.keys())

    priority_columns = ['instruction', 'input', 'chosen', 'rejected']
    ordered_columns = []

    if not items:
        ordered_columns = priority_columns
    else:
        for col in priority_columns:
            if col in all_columns:
                ordered_columns.append(col)
                all_columns.discard(col)
        ordered_columns.extend(sorted(all_columns))

    wb = Workbook()
    ws = wb.active
    ws.title = "训练数据集"
    ws.append(ordered_columns)

    for item in items:
        row = [item.get(col, '') for col in ordered_columns]
        row = ['' if v is None else v for v in row]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _convert_training_role_based_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """
    将 role-based 格式的训练数据集转换为 XLSX
    
    转换规则（反向执行 xlsx 解析逻辑）：
    - 单轮格式：system, user, assistant
    - 多轮格式：system, user1, assistant1, user2, assistant2...
    
    Args:
        items: 数据项列表（每个项包含 messages 数组）
    
    Returns:
        XLSX 文件的字节内容
    """
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "训练数据集"
    
    # 分析所有数据项，确定是单轮还是多轮格式
    max_turns = 0
    
    for item in items:
        if 'messages' not in item:
            continue
        
        messages = item['messages']
        if not isinstance(messages, list):
            continue
        
        # 统计 user-assistant 轮数（排除 system）
        user_assistant_pairs = 0
        for msg in messages:
            if msg.get('role') == 'user':
                user_assistant_pairs += 1
        
        max_turns = max(max_turns, user_assistant_pairs)
    
    # 构建表头
    headers = []
    
    # 添加 system 列
    headers.append('system')
    
    # 添加 user 和 assistant 列
    if max_turns == 1:
        # 单轮格式
        headers.append('user')
        headers.append('assistant')
    else:
        # 多轮格式
        for turn_num in range(1, max_turns + 1):
            headers.append(f'user{turn_num}')
            headers.append(f'assistant{turn_num}')
    
    # 写入表头
    ws.append(headers)
    
    # 写入数据行
    for item in items:
        row_data = {}
        
        # 提取 messages 内容
        if 'messages' in item and isinstance(item['messages'], list):
            messages = item['messages']
            
            # 提取 system 消息
            system_content = ''
            user_assistant_pairs = []
            
            for msg in messages:
                role = msg.get('role', '')
                content = msg.get('content', '')
                
                if role == 'system':
                    system_content = content
                elif role == 'user':
                    user_assistant_pairs.append({'user': content, 'assistant': ''})
                elif role == 'assistant' and user_assistant_pairs:
                    user_assistant_pairs[-1]['assistant'] = content
            
            # 填充 system 列
            row_data['system'] = system_content
            
            # 填充 user 和 assistant 列
            if max_turns == 1:
                # 单轮格式
                if len(user_assistant_pairs) > 0:
                    row_data['user'] = user_assistant_pairs[0].get('user', '')
                    row_data['assistant'] = user_assistant_pairs[0].get('assistant', '')
                else:
                    row_data['user'] = ''
                    row_data['assistant'] = ''
            else:
                # 多轮格式
                for turn_num in range(1, max_turns + 1):
                    if turn_num <= len(user_assistant_pairs):
                        pair = user_assistant_pairs[turn_num - 1]
                        row_data[f'user{turn_num}'] = pair.get('user', '')
                        row_data[f'assistant{turn_num}'] = pair.get('assistant', '')
                    else:
                        row_data[f'user{turn_num}'] = ''
                        row_data[f'assistant{turn_num}'] = ''
        
        # 构建行数据（按照表头顺序）
        row = [row_data.get(header, '') for header in headers]
        ws.append(row)
    
    # 将工作簿保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def _convert_training_dpo_role_based_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """将 DPO role-based 格式的训练数据集转换为 XLSX。"""
    headers = ["messages", "chosen", "rejected"]
    has_images = any("images" in item for item in items)
    if has_images:
        headers.append("images")

    wb = Workbook()
    ws = wb.active
    ws.title = "训练数据集"
    ws.append(headers)

    for item in items:
        row = [
            _convert_value_to_excel(item.get("messages", [])),
            _convert_value_to_excel(item.get("chosen", {})),
            _convert_value_to_excel(item.get("rejected", {})),
        ]
        if has_images:
            row.append(_convert_value_to_excel(item.get("images", [])))
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _convert_training_grpo_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """将 GRPO 格式的训练数据集转换为 XLSX。"""
    all_columns = set()
    for item in items:
        all_columns.update(item.keys())

    priority_columns = ["data_source", "prompt", "ability", "reward_model", "extra_info"]
    ordered_columns = []

    if not items:
        ordered_columns = priority_columns
    else:
        for col in priority_columns:
            if col in all_columns:
                ordered_columns.append(col)
                all_columns.discard(col)
        ordered_columns.extend(sorted(all_columns))

    wb = Workbook()
    ws = wb.active
    ws.title = "训练数据集"
    ws.append(ordered_columns)

    for item in items:
        row = [_convert_value_to_excel(item.get(col, "")) for col in ordered_columns]
        ws.append(row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _convert_value_to_excel(value: Any) -> str:
    """
    将值转换为 Excel 可接受的格式
    
    Args:
        value: 要转换的值
    
    Returns:
        转换后的字符串值
    """
    # None -> 空字符串
    if value is None:
        return ''
    
    # 基本类型（str, int, float, bool）-> 转换为字符串
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    
    # 复杂类型（list, dict）-> 转换为 JSON 字符串
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    
    # 其他类型 -> 转换为字符串
    return str(value)


def _convert_business_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """
    将业务数据集转换为 XLSX 格式
    
    转换规则：
    - 每一行 jsonl 对象就为一个 json 对象
    - 默认第一个对象的属性为表头
    - 后面的 json 对象都以当前表头为准（缺少的字段填充空值）
    
    Args:
        items: 数据项列表
    
    Returns:
        XLSX 文件的字节内容
    """
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "业务数据集"
    
    # 如果数据为空，返回只有表头的空 Excel
    if not items:
        # 空数据时，返回只有表头的空 Excel（表头为空）
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    # 使用第一个对象的键作为表头
    headers = list(items[0].keys())
    
    # 写入表头
    ws.append(headers)
    
    # 写入数据行
    for item in items:
        # 按照表头顺序获取值，如果某个字段不存在则填充空字符串
        row = [item.get(col, '') for col in headers]
        # 使用辅助函数处理每个值，转换为 Excel 可接受的格式
        row = [_convert_value_to_excel(v) for v in row]
        ws.append(row)
    
    # 将工作簿保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
