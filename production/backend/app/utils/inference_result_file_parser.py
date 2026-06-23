"""
推理结果集文件解析工具
"""

import os
import json
import io
import re
from typing import List, Dict, Any, Optional, Tuple
from io import BytesIO
from fastapi import UploadFile, HTTPException
import pandas as pd

from app.core.logging import logger
from app.models import InferenceResultDataset
from app.schemas.inference_result import InferenceResultDatasetUploadType, InferenceResultDatasetExportType
from app.schemas.training_task import TrainingTypeCategory
from app.schemas import DatasetFormat
from app.services.storage.interface import StorageService
from app.utils.dataset_file_parser import analyze_image_understanding_dataset_file_content
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

from app.utils.validators import validate_dataset_upload_file_type


# =========== 文件解析器 ============
class TextGenerationInferenceResultFileParser:
    """文本生成推理结果集文件解析器"""

    def __init__(self):
        """初始化解析器"""
        pass

    # =========== 校验辅助函数 ============
    @staticmethod
    def _detect_dataset_format(file_content: bytes, file_type: str) -> Optional[str]:
        """
        根据文件内容自动检测数据集格式

        检测逻辑：
        - 如果数据项包含 `instruction/input/chosen/rejected` 字段 → alpaca
        - 如果数据项包含 `prompt/completion/reward` 或 `prompt/response/reward` 字段 → completion-reward
        - 如果数据项包含 `messages` 字段且是数组 → role-based
        - 如果数据项包含 `prompt` 和 `response` 字段但不包含 `messages` → prompt-response
        - 如果两者都包含，优先判断为 role-based（因为 role-based 更严格）

        Args:
            file_content: 文件字节内容
            file_type: 文件类型（jsonl, json, csv, xlsx）

        Returns:
            检测到的格式（'role-based' 或 'prompt-response'），如果无法检测则返回 None
        """
        try:
            # 根据文件类型尝试解析前几行/项来检测格式
            sample_items = []

            if file_type == 'jsonl':
                # JSONL: 读取前5行
                lines = file_content.decode('utf-8').strip().split('\n')
                for line in lines[:5]:
                    line = line.strip()
                    if line and not line.lstrip().startswith('#'):
                        try:
                            item = json.loads(line)
                            if isinstance(item, dict):
                                sample_items.append(item)
                        except json.JSONDecodeError:
                            continue

            elif file_type == 'json':
                # JSON: 解析整个文件，取前5项
                content_str = file_content.decode('utf-8').strip()
                parsed_data = json.loads(content_str)
                if isinstance(parsed_data, list):
                    sample_items = parsed_data[:5]
                elif isinstance(parsed_data, dict):
                    sample_items = [parsed_data]

            elif file_type in ['csv', 'xlsx']:
                # CSV/XLSX: 检查列名来判断格式
                try:
                    if file_type == 'csv':
                        encodings_to_try = ['gbk', 'gb18030', 'utf-8', 'latin-1', 'cp1252']
                        df = None
                        for encoding in encodings_to_try:
                            try:
                                df = pd.read_csv(BytesIO(file_content), encoding=encoding, nrows=1)
                                break
                            except Exception:
                                continue
                    else:  # xlsx
                        df = pd.read_excel(BytesIO(file_content), nrows=1)

                    if df is not None and len(df.columns) > 0:
                        # 检查列名来判断格式
                        columns = [col.lower() for col in df.columns]

                        has_alpaca_cols = all(col in columns for col in ['instruction', 'input', 'chosen', 'rejected'])
                        has_completion_reward_cols = (
                            'prompt' in columns
                            and ('completion' in columns or 'response' in columns)
                            and 'reward' in columns
                        )

                        # 检查是否有 role-based 格式的特征列（system、user、assistant 或 user1、assistant1）
                        has_system_col = 'system' in columns
                        has_user_col = 'user' in columns or any(col.startswith('user') for col in columns)
                        has_assistant_col = 'assistant' in columns or any(col.startswith('assistant') for col in columns)

                        # 检查是否有 prompt-response 格式的特征列（prompt、response）
                        has_prompt_col = 'prompt' in columns
                        has_response_col = 'response' in columns or '标准回答' in columns

                        # 判断格式
                        if has_alpaca_cols:
                            return DatasetFormat.ALPACA.value
                        elif has_completion_reward_cols:
                            return DatasetFormat.COMPLETION_REWARD.value
                        elif has_system_col and (has_user_col or has_assistant_col):
                            # 如果有 system 列且同时有 user/assistant 列，判断为 role-based
                            return DatasetFormat.ROLE_BASED.value
                        elif has_prompt_col and has_response_col:
                            # 如果有 prompt 和 response 列，判断为 prompt-response
                            return DatasetFormat.PROMPT_RESPONSE.value
                        else:
                            # 无法确定格式
                            return None
                    else:
                        return None
                except Exception:
                    # 如果解析失败，无法检测格式
                    return None

            # 如果没有样本，无法检测
            if not sample_items:
                return None

            # 检测格式
            has_alpaca = False
            has_completion_reward = False
            has_messages = False
            has_prompt_response = False

            for item in sample_items:
                if not isinstance(item, dict):
                    continue

                if all(field in item for field in ['instruction', 'input', 'chosen', 'rejected']):
                    has_alpaca = True

                if (
                    'prompt' in item
                    and ('completion' in item or 'response' in item)
                    and 'reward' in item
                ):
                    has_completion_reward = True

                # 检查是否有 messages 字段且是数组
                if 'messages' in item and isinstance(item.get('messages'), list):
                    has_messages = True

                # 检查是否有 prompt 和 response 字段
                if 'prompt' in item and 'response' in item:
                    has_prompt_response = True

            # 判断格式
            if has_alpaca:
                return DatasetFormat.ALPACA.value
            elif has_completion_reward:
                return DatasetFormat.COMPLETION_REWARD.value
            elif has_messages:
                # 如果包含 messages，优先判断为 role-based
                return DatasetFormat.ROLE_BASED.value
            elif has_prompt_response:
                # 如果只包含 prompt 和 response，判断为 prompt-response
                return DatasetFormat.PROMPT_RESPONSE.value
            else:
                # 无法确定格式
                return None

        except Exception as e:
            logger.warning(f"格式检测失败: {str(e)}")
            return None

    @staticmethod
    def _validate_prompt_response_item(item: Dict[str, Any], line_num: Optional[int] = None,
                                      item_idx: Optional[int] = None) -> Tuple[bool, str]:
        """
        校验 prompt-response 格式的数据项

        校验规则：
        1. prompt 字段必须存在且不能为空
        2. response 字段必须存在且不能为空

        Args:
            item: 数据项字典
            line_num: 行号（用于错误提示）
            item_idx: 数据项索引（用于错误提示）

        Returns:
            flag: True 如果格式正确，False 如果格式错误
            process_error_msg: 格式错误提示信息
        """
        process_error_msg = ""

        # 1. 校验 prompt 字段
        if 'prompt' not in item:
            if line_num:
                process_error_msg = f"第{line_num}行：缺少prompt字段"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：缺少prompt字段"
            else:
                process_error_msg = "缺少prompt字段"

            logger.warning(process_error_msg)
            return False, process_error_msg

        prompt = item.get('prompt', '')
        if not prompt or not str(prompt).strip():
            if line_num:
                process_error_msg = f"第{line_num}行：prompt不能为空"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：prompt不能为空"
            else:
                process_error_msg = "prompt不能为空"

            logger.warning(process_error_msg)
            return False, process_error_msg

        # 2. 校验 response 字段
        if 'response' not in item:
            if line_num:
                process_error_msg = f"第{line_num}行：缺少response字段"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：缺少response字段"
            else:
                process_error_msg = "缺少response字段"

            logger.warning(process_error_msg)
            return False, process_error_msg

        response = item.get('response', '')
        if not response or not str(response).strip():
            if line_num:
                process_error_msg = f"第{line_num}行：response不能为空"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：response不能为空"
            else:
                process_error_msg = "response不能为空"

            logger.warning(process_error_msg)
            return False, process_error_msg

        return True, process_error_msg

    @staticmethod
    def _validate_alpaca_item(item: Dict[str, Any], line_num: Optional[int] = None,
                              item_idx: Optional[int] = None) -> Tuple[bool, str]:
        process_error_msg = ""
        for field_name in ['instruction', 'input', 'chosen', 'rejected']:
            if field_name not in item:
                process_error_msg = f"第{line_num or item_idx}个样本：缺少{field_name}字段"
                logger.warning(process_error_msg)
                return False, process_error_msg
            value = item.get(field_name)
            if not isinstance(value, str) or not value.strip():
                process_error_msg = f"第{line_num or item_idx}个样本：{field_name}不能为空"
                logger.warning(process_error_msg)
                return False, process_error_msg
        return True, process_error_msg

    @staticmethod
    def _validate_completion_reward_item(item: Dict[str, Any], line_num: Optional[int] = None,
                                         item_idx: Optional[int] = None) -> Tuple[bool, str]:
        process_error_msg = ""
        display_index = line_num or item_idx
        prompt = item.get('prompt')
        completion = item.get('completion', item.get('response'))
        if not isinstance(prompt, str) or not prompt.strip():
            process_error_msg = f"第{display_index}个样本：prompt不能为空"
            logger.warning(process_error_msg)
            return False, process_error_msg
        if not isinstance(completion, str) or not completion.strip():
            process_error_msg = f"第{display_index}个样本：completion/response不能为空"
            logger.warning(process_error_msg)
            return False, process_error_msg
        if 'reward' not in item:
            process_error_msg = f"第{display_index}个样本：缺少reward字段"
            logger.warning(process_error_msg)
            return False, process_error_msg
        return True, process_error_msg

    @staticmethod
    def _normalize_completion_reward_item(item: Dict[str, Any]) -> Dict[str, Any]:
        normalized = dict(item)
        if 'completion' not in normalized and 'response' in normalized:
            normalized['completion'] = normalized['response']
        if 'response' not in normalized and 'completion' in normalized:
            normalized['response'] = normalized['completion']
        return normalized

    @staticmethod
    def _validate_prompt_format(prompt: str, line_num: Optional[int] = None, item_idx: Optional[int] = None) -> Tuple[bool, str]:
        """
        校验 prompt 格式。
        - 单轮（无 <User>/<Assistant> 标签）：仅要求 prompt 非空。
        - 多轮（含标签）：每个 <User> 和 <Assistant> 标签后的内容不能为空。

        Args:
            prompt: prompt 字符串
            line_num: 行号（用于错误提示）
            item_idx: 数据项索引（用于错误提示）

        Returns:
            flag: True 如果格式正确，False 如果格式错误
            process_error_msg: 格式错误提示信息
        """
        process_error_msg = ""
        if not prompt or not prompt.strip():
            if line_num:
                process_error_msg = f"第{line_num}行：prompt不能为空"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：prompt不能为空"
            else:
                process_error_msg = "prompt不能为空"

            logger.warning(process_error_msg)
            return False, process_error_msg

        # 使用正则表达式提取标签后的内容
        # 匹配 <User> 或 <Assistant> 标签及其后的内容
        # 格式：<User> 内容\n<Assistant> 内容\n<User> 内容...
        # 使用非贪婪匹配，匹配到下一个标签或字符串结尾
        # 注意：标签大小写敏感，需要匹配 <User> 和 <Assistant>
        pattern = r'<(User|Assistant)>\s*(.*?)(?=\n<(?:User|Assistant)>|$)'
        matches = re.findall(pattern, prompt, re.DOTALL)

        if not matches:
            # 单轮：无 <User>/<Assistant> 标签时，仅要求 prompt 非空（已在上面校验），视为合法
            if '<User>' not in prompt and '<Assistant>' not in prompt:
                return True, process_error_msg

        # 多轮：检查每个标签后的内容是否为空
        for tag, content in matches:
            # 去除首尾空白后检查是否为空
            content_stripped = content.strip()
            if not content_stripped:
                if line_num:
                    process_error_msg = f"第{line_num}行：prompt中<{tag}>标签后的内容不能为空"
                elif item_idx:
                    process_error_msg = f"第{item_idx}个对象：prompt中<{tag}>标签后的内容不能为空"
                else:
                    process_error_msg = f"prompt中<{tag}>标签后的内容不能为空"

                logger.warning(process_error_msg)
                return False, process_error_msg

        return True, process_error_msg

    @staticmethod
    def _validate_role_based_item(item: Dict[str, Any], line_num: Optional[int] = None,
                                  item_idx: Optional[int] = None) -> Tuple[bool, str]:
        """
        校验 role-based 格式的数据项

        校验规则：
        1. messages 列表中，当 role 不为 system 时，content 必须存在且不能为空
        2. prompt 存在且不能为空；单轮时仅要求非空，多轮时还要求每个 <User>/<Assistant> 标签后内容非空
        3. response 存在且不能为空

        Args:
            item: 数据项字典
            line_num: 行号（用于错误提示）
            item_idx: 数据项索引（用于错误提示）

        Returns:
            flag: True 如果格式正确，False 如果格式错误
            process_error_msg: 格式错误提示信息
        """
        process_error_msg = ""
        # 1. 校验 messages 中的 content（role 不为 system 时）
        if 'messages' in item and isinstance(item['messages'], list):
            for msg_idx, message in enumerate(item['messages']):
                if isinstance(message, dict):
                    role = message.get('role', '')
                    if role != 'system':
                        content = message.get('content', '')
                        if not content or not str(content).strip():
                            if line_num:
                                process_error_msg = f"第{line_num}行第{msg_idx + 1}个message（role={role}）的content不能为空"
                            elif item_idx:
                                process_error_msg = f"第{item_idx}个对象第{msg_idx + 1}个message（role={role}）的content不能为空"
                            else:
                                process_error_msg = f"第{msg_idx + 1}个message（role={role}）的content不能为空"

                            logger.warning(process_error_msg)
                            return False, process_error_msg

        # 2. 校验 prompt
        if 'prompt' not in item:
            if line_num:
                process_error_msg = f"第{line_num}行：缺少prompt字段"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：缺少prompt字段"
            else:
                process_error_msg = "缺少prompt字段"

            logger.warning(process_error_msg)
            return False, process_error_msg

        prompt = item.get('prompt', '')
        flag, process_error_msg = TextGenerationInferenceResultFileParser._validate_prompt_format(prompt, line_num, item_idx)
        if not flag:
            return flag, process_error_msg

        # 3. 校验 response
        if 'response' not in item:
            if line_num:
                process_error_msg = f"第{line_num}行：缺少response字段"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：缺少response字段"
            else:
                process_error_msg = "缺少response字段"

            logger.warning(process_error_msg)
            return False, process_error_msg

        response = item.get('response', '')
        if not response or not str(response).strip():
            if line_num:
                process_error_msg = f"第{line_num}行：response不能为空"
            elif item_idx:
                process_error_msg = f"第{item_idx}个对象：response不能为空"
            else:
                process_error_msg = "response不能为空"

            logger.warning(process_error_msg)
            return False, process_error_msg

        return True, process_error_msg

    @staticmethod
    def _parse_prompt_response_jsonl(content: bytes) -> List[Dict[str, Any]]:
        """
        解析JSONL文件（逐行解析）

        添加字段校验，至少检查 prompt 和 response 字段是否存在
        如果缺少必需字段，抛出明确的错误信息
        """
        items = []
        for line_num, line in enumerate(content.decode('utf-8').strip().split('\n'), start=1):
            line = line.strip()
            if not line:
                continue

            # 跳过注释行
            if line.lstrip().startswith('#'):
                continue

            try:
                item = json.loads(line)
                if isinstance(item, dict):
                    # 校验必需字段
                    flag, error_msg = TextGenerationInferenceResultFileParser._validate_prompt_response_item(
                        item, line_num=line_num
                    )
                    if not flag:
                        raise ValueError(error_msg)
                    items.append(item)
                else:
                    raise ValueError(f"第{line_num}行：数据项必须是JSON对象格式")
            except json.JSONDecodeError as e:
                error_msg = f"第{line_num}行：JSON解析失败 - {str(e)}"
                logger.warning(error_msg)
                raise ValueError(error_msg)
            except ValueError as e:
                # 重新抛出校验错误
                raise ValueError(str(e))

        if len(items) == 0:
            raise ValueError("JSONL文件中没有找到有效的数据样本")

        return items

    @staticmethod
    def _parse_prompt_response_json(content: bytes) -> List[Dict[str, Any]]:
        """
        解析JSON文件（整体解析，支持跨行对象）

        添加字段校验，至少检查 prompt 和 response 字段是否存在
        如果缺少必需字段，抛出明确的错误信息
        """
        items = []
        content_str = content.decode('utf-8').strip()

        try:
            parsed_data = json.loads(content_str)

            if isinstance(parsed_data, list):
                for item_idx, item in enumerate(parsed_data, start=1):
                    if isinstance(item, dict):
                        # 校验必需字段
                        flag, error_msg = TextGenerationInferenceResultFileParser._validate_prompt_response_item(
                            item, item_idx=item_idx
                        )
                        if not flag:
                            raise ValueError(error_msg)
                        items.append(item)
                    else:
                        raise ValueError(f"第{item_idx}个对象：数据项必须是JSON对象格式")

                if len(items) == 0:
                    raise ValueError("JSON文件中没有找到有效的数据样本")
                return items
            elif isinstance(parsed_data, dict):
                # 校验必需字段
                flag, error_msg = TextGenerationInferenceResultFileParser._validate_prompt_response_item(
                    parsed_data, item_idx=1
                )
                if not flag:
                    raise ValueError(error_msg)
                items.append(parsed_data)
                return items
            else:
                raise ValueError(f"JSON格式不是数组或对象，类型: {type(parsed_data).__name__}")
        except json.JSONDecodeError as e:
            raise ValueError(f"解析JSON文件失败: {str(e)}")

    @staticmethod
    def _parse_prompt_response_csv(content: bytes) -> List[Dict[str, Any]]:
        """
        解析CSV文件

        添加字段校验，至少检查 prompt 和 response 字段是否存在
        如果缺少必需字段，抛出明确的错误信息
        """
        try:
            encodings_to_try = ['gbk', 'gb18030', 'utf-8', 'latin-1', 'cp1252']

            for encoding in encodings_to_try:
                try:
                    df = pd.read_csv(BytesIO(content), encoding=encoding)
                    break
                except Exception as e:
                    logger.warning(f"尝试以{encoding}编码解析csv失败！- {str(e)}")
                    continue
            else:
                raise ValueError(f"不支持的文件编码，目前支持: {', '.join(encodings_to_try)}")

            items = []
            for row_num, (_, row) in enumerate(df.iterrows(), start=2):  # 从第2行开始（第1行是表头）
                item = {
                    'system': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('system', '')
                    ),
                    'prompt': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('prompt', '')
                    ),
                    'response': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('标准回答', row.get('response', ''))
                    ),
                    'model_response': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('模型回答', row.get('model_response', ''))
                    ),
                }
                # 校验必需字段
                flag, error_msg = TextGenerationInferenceResultFileParser._validate_prompt_response_item(
                    item, line_num=row_num
                )
                if not flag:
                    raise ValueError(error_msg)
                items.append(item)

            if len(items) == 0:
                raise ValueError("CSV文件中没有找到有效的数据样本")

            return items
        except ValueError:
            # 重新抛出校验错误
            raise
        except Exception as e:
            raise ValueError(f"解析CSV文件失败: {str(e)}")

    @staticmethod
    def _parse_prompt_response_excel(content: bytes) -> List[Dict[str, Any]]:
        """
        解析Excel文件

        添加字段校验，至少检查 prompt 和 response 字段是否存在
        如果缺少必需字段，抛出明确的错误信息
        """
        try:
            df = pd.read_excel(BytesIO(content))
            items = []
            for row_num, (_, row) in enumerate(df.iterrows(), start=2):  # 从第2行开始（第1行是表头）
                item = {
                    'system': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('system', '')
                    ),
                    'prompt': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('prompt', '')
                    ),
                    'response': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('标准回答', row.get('response', ''))
                    ),
                    'model_response': TextGenerationInferenceResultFileParser._prompt_response_field_value(
                        row.get('模型回答', row.get('model_response', ''))
                    ),
                }
                # 校验必需字段
                flag, error_msg = TextGenerationInferenceResultFileParser._validate_prompt_response_item(
                    item, line_num=row_num
                )
                if not flag:
                    raise ValueError(error_msg)
                items.append(item)

            if len(items) == 0:
                raise ValueError("Excel文件中没有找到有效的数据样本")

            return items
        except ValueError:
            # 重新抛出校验错误
            raise
        except Exception as e:
            raise ValueError(f"解析Excel文件失败: {str(e)}")

    @staticmethod
    def _parse_structured_jsonl(
        content: bytes,
        validator,
        normalizer=lambda item: item,
    ) -> List[Dict[str, Any]]:
        items = []
        for line_num, line in enumerate(content.decode('utf-8').strip().split('\n'), start=1):
            line = line.strip()
            if not line or line.lstrip().startswith('#'):
                continue
            try:
                item = json.loads(line)
            except json.JSONDecodeError as e:
                raise ValueError(f"第{line_num}行：JSON解析失败 - {str(e)}")
            if not isinstance(item, dict):
                raise ValueError(f"第{line_num}行：数据项必须是JSON对象格式")
            flag, error_msg = validator(item, line_num=line_num)
            if not flag:
                raise ValueError(error_msg)
            items.append(normalizer(item))
        if len(items) == 0:
            raise ValueError("JSONL文件中没有找到有效的数据样本")
        return items

    @staticmethod
    def _parse_structured_json(
        content: bytes,
        validator,
        normalizer=lambda item: item,
    ) -> List[Dict[str, Any]]:
        content_str = content.decode('utf-8').strip()
        try:
            parsed_data = json.loads(content_str)
        except json.JSONDecodeError as e:
            raise ValueError(f"解析JSON文件失败: {str(e)}")

        data_items = parsed_data if isinstance(parsed_data, list) else [parsed_data]
        if not isinstance(data_items, list):
            raise ValueError(f"JSON格式不是数组或对象，类型: {type(parsed_data).__name__}")

        items = []
        for item_idx, item in enumerate(data_items, start=1):
            if not isinstance(item, dict):
                raise ValueError(f"第{item_idx}个对象：数据项必须是JSON对象格式")
            flag, error_msg = validator(item, item_idx=item_idx)
            if not flag:
                raise ValueError(error_msg)
            items.append(normalizer(item))
        if len(items) == 0:
            raise ValueError("JSON文件中没有找到有效的数据样本")
        return items

    @staticmethod
    def _parse_structured_excel(
        content: bytes,
        required_columns: List[str],
        validator,
        normalizer=lambda item: item,
    ) -> List[Dict[str, Any]]:
        try:
            df = pd.read_excel(BytesIO(content))
            columns = {str(col).strip().lower(): col for col in df.columns}
            missing_columns = [col for col in required_columns if col not in columns]
            if missing_columns:
                raise ValueError(f"Excel文件缺少必需列: {missing_columns}")

            items = []
            for row_num, (_, row) in enumerate(df.iterrows(), start=2):
                item = {
                    column: TextGenerationInferenceResultFileParser._prompt_response_field_value(row.get(columns[column], ''))
                    for column in required_columns
                }
                flag, error_msg = validator(item, line_num=row_num)
                if not flag:
                    raise ValueError(error_msg)
                items.append(normalizer(item))
            if len(items) == 0:
                raise ValueError("Excel文件中没有找到有效的数据样本")
            return items
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"解析Excel文件失败: {str(e)}")

    @staticmethod
    def _parse_alpaca_jsonl(content: bytes) -> List[Dict[str, Any]]:
        return TextGenerationInferenceResultFileParser._parse_structured_jsonl(
            content,
            TextGenerationInferenceResultFileParser._validate_alpaca_item,
        )

    @staticmethod
    def _parse_alpaca_json(content: bytes) -> List[Dict[str, Any]]:
        return TextGenerationInferenceResultFileParser._parse_structured_json(
            content,
            TextGenerationInferenceResultFileParser._validate_alpaca_item,
        )

    @staticmethod
    def _parse_alpaca_excel(content: bytes) -> List[Dict[str, Any]]:
        return TextGenerationInferenceResultFileParser._parse_structured_excel(
            content,
            ['instruction', 'input', 'chosen', 'rejected'],
            TextGenerationInferenceResultFileParser._validate_alpaca_item,
        )

    @staticmethod
    def _parse_completion_reward_jsonl(content: bytes) -> List[Dict[str, Any]]:
        return TextGenerationInferenceResultFileParser._parse_structured_jsonl(
            content,
            TextGenerationInferenceResultFileParser._validate_completion_reward_item,
            TextGenerationInferenceResultFileParser._normalize_completion_reward_item,
        )

    @staticmethod
    def _parse_completion_reward_json(content: bytes) -> List[Dict[str, Any]]:
        return TextGenerationInferenceResultFileParser._parse_structured_json(
            content,
            TextGenerationInferenceResultFileParser._validate_completion_reward_item,
            TextGenerationInferenceResultFileParser._normalize_completion_reward_item,
        )

    @staticmethod
    def _parse_completion_reward_excel(content: bytes) -> List[Dict[str, Any]]:
        try:
            df = pd.read_excel(BytesIO(content))
            columns = {str(col).strip().lower(): col for col in df.columns}
            completion_column = 'completion' if 'completion' in columns else 'response'
            required_columns = ['prompt', completion_column, 'reward']
            missing_columns = [col for col in required_columns if col not in columns]
            if missing_columns:
                raise ValueError(f"Excel文件缺少必需列: {missing_columns}")

            items = []
            for row_num, (_, row) in enumerate(df.iterrows(), start=2):
                item = {
                    'prompt': TextGenerationInferenceResultFileParser._prompt_response_field_value(row.get(columns['prompt'], '')),
                    completion_column: TextGenerationInferenceResultFileParser._prompt_response_field_value(row.get(columns[completion_column], '')),
                    'reward': row.get(columns['reward'], ''),
                }
                flag, error_msg = TextGenerationInferenceResultFileParser._validate_completion_reward_item(item, line_num=row_num)
                if not flag:
                    raise ValueError(error_msg)
                items.append(TextGenerationInferenceResultFileParser._normalize_completion_reward_item(item))
            if len(items) == 0:
                raise ValueError("Excel文件中没有找到有效的数据样本")
            return items
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"解析Excel文件失败: {str(e)}")

    @staticmethod
    def _parse_role_based_jsonl(content: bytes) -> List[Dict[str, Any]]:
        """
        解析role-based格式的JSONL文件（推理结果集版本）
        参考训练数据集的解析规则，但保留所有字段（包括推理结果集特有字段）

        Args:
            content: JSONL文件的字节内容

        Returns:
            解析后的数据项列表
        """
        items = []
        content_str = content.decode('utf-8')
        lines = content_str.splitlines()

        for line_num, line in enumerate(lines, start=1):
            line = line.strip()
            if not line:
                continue

            # 跳过注释行（支持 # 前面有空格的情况）
            if line.lstrip().startswith('#'):
                continue

            try:
                # 解析JSON
                parsed_data = json.loads(line)

                if not isinstance(parsed_data, dict):
                    raise ValueError(f"第{line_num}行格式错误：数据应该是JSON对象格式")

                # 验证 messages 字段
                if 'messages' not in parsed_data:
                    raise ValueError(f"第{line_num}行缺少messages字段")

                messages = parsed_data['messages']
                if not isinstance(messages, list) or len(messages) == 0:
                    raise ValueError(f"第{line_num}行messages必须是非空数组")

                # 验证每个 message
                expected_role = None
                for msg_idx, message in enumerate(messages):
                    if not isinstance(message, dict):
                        raise ValueError(f"第{line_num}行第{msg_idx + 1}个message必须是对象格式")

                    # 验证 role 字段
                    if 'role' not in message:
                        raise ValueError(f"第{line_num}行第{msg_idx + 1}个message缺少role字段")

                    role = message['role']
                    if role not in ['user', 'assistant', 'system']:
                        raise ValueError(f"第{line_num}行第{msg_idx + 1}个message的role必须是user、assistant或system")

                    # 验证 content 字段
                    if 'content' not in message:
                        raise ValueError(f"第{line_num}行第{msg_idx + 1}个message缺少content字段")

                    content = message['content']
                    if not isinstance(content, str) or not content.strip():
                        raise ValueError(f"第{line_num}行第{msg_idx + 1}个message的content不能为空，跳过")

                    # 验证 role 交替出现（system 不影响交替规则）
                    if role != 'system':
                        # 如果是第一个非 system 的消息，必须是 user
                        if expected_role is None:
                            if role != 'user':
                                raise ValueError(
                                    f"第{line_num}行第一个非system的message必须是user，"
                                    f"但第{msg_idx + 1}个message的role为{role}，跳过"
                                )
                            expected_role = 'assistant'
                        else:
                            # 后续消息必须与期望的角色匹配
                            if role != expected_role:
                                raise ValueError(
                                    f"第{line_num}行messages中role必须交替出现（user和assistant），"
                                    f"但第{msg_idx + 1}个message的role为{role}，期望为{expected_role}，跳过"
                                )
                            # 切换期望的角色
                            expected_role = 'assistant' if role == 'user' else 'user'

                # 添加空值校验：校验 messages、prompt、response
                flag, process_error_msg = TextGenerationInferenceResultFileParser._validate_role_based_item(parsed_data, line_num=line_num)
                if not flag:
                    # 空值校验失败，抛出异常，用于错误记录
                    raise ValueError(process_error_msg)

                # 验证通过，保留所有字段（包括 messages、response、system、prompt、model_response 等）
                items.append(parsed_data)

            except Exception as e:
                # 捕获其他异常并抛出
                logger.warning(f"文件格式异常: {str(e)}")
                raise ValueError(str(e))

        if len(items) == 0:
            raise ValueError("JSONL文件中没有找到有效的数据样本")

        return items

    @staticmethod
    def _parse_role_based_json(content: bytes) -> List[Dict[str, Any]]:
        """
        解析role-based格式的JSON文件（推理结果集版本）
        参考训练数据集的解析规则，但保留所有字段（包括推理结果集特有字段）

        Args:
            content: JSON文件的字节内容

        Returns:
            解析后的数据项列表
        """
        items = []
        content_str = content.decode('utf-8').strip()

        try:
            # 解析JSON内容
            data = json.loads(content_str)

            # 将数据转换为列表格式统一处理
            if isinstance(data, list):
                # 数组格式：每个元素是一个对话样本
                data_list = data
            elif isinstance(data, dict):
                # 单个对象格式：转换为列表
                data_list = [data]
            else:
                raise ValueError("JSON文件格式错误：应该是对象或数组格式")

            # 处理每个数据项
            for item_idx, item in enumerate(data_list, start=1):
                try:
                    if not isinstance(item, dict):
                        raise ValueError(f"第{item_idx}个对象不是标准的json对象格式，跳过")

                    # 验证 messages 字段
                    if 'messages' not in item:
                        raise ValueError(f"第{item_idx}个对象缺少messages字段，跳过")

                    messages = item['messages']
                    if not isinstance(messages, list) or len(messages) == 0:
                        raise ValueError(f"第{item_idx}个对象的messages必须是非空数组，跳过")

                    # 验证每个 message
                    expected_role = None
                    for msg_idx, message in enumerate(messages):
                        if not isinstance(message, dict):
                            raise ValueError(f"第{item_idx}个对象第{msg_idx + 1}个message必须是对象格式，跳过")

                        # 验证 role 字段
                        if 'role' not in message:
                            raise ValueError(f"第{item_idx}个对象第{msg_idx + 1}个message缺少role字段，跳过")

                        role = message['role']
                        if role not in ['user', 'assistant', 'system']:
                            raise ValueError(
                                f"第{item_idx}个对象第{msg_idx + 1}个message的role必须是user、assistant或system，跳过"
                            )

                        # 验证 content 字段
                        if 'content' not in message:
                            raise ValueError(f"第{item_idx}个对象第{msg_idx + 1}个message缺少content字段，跳过")

                        content = message['content']
                        if not isinstance(content, str) or not content.strip():
                            raise ValueError(f"第{item_idx}个对象第{msg_idx + 1}个message的content不能为空，跳过")

                        # 验证 role 交替出现（system 不影响交替规则）
                        if role != 'system':
                            # 如果是第一个非 system 的消息，必须是 user
                            if expected_role is None:
                                if role != 'user':
                                    raise ValueError(
                                        f"第{item_idx}个对象第一个非system的message必须是user，"
                                        f"但第{msg_idx + 1}个message的role为{role}，跳过"
                                    )
                                expected_role = 'assistant'
                            else:
                                # 后续消息必须与期望的角色匹配
                                if role != expected_role:
                                    raise ValueError(
                                        f"第{item_idx}个对象messages中role必须交替出现（user和assistant），"
                                        f"但第{msg_idx + 1}个message的role为{role}，期望为{expected_role}，跳过"
                                    )
                                # 切换期望的角色
                                expected_role = 'assistant' if role == 'user' else 'user'

                    # 添加空值校验：校验 messages、prompt、response
                    flag, process_error_msg = TextGenerationInferenceResultFileParser._validate_role_based_item(item, item_idx=item_idx)
                    if not flag:
                        # 空值校验失败，抛出异常，用于错误记录
                        raise ValueError(process_error_msg)

                    # 验证通过，保留所有字段（包括 messages、response、system、prompt、model_response 等）
                    items.append(item)

                except Exception as e:
                    # 捕获其他异常
                    raise ValueError(str(e))

            if len(items) == 0:
                raise ValueError("JSON文件中没有找到有效的数据样本")

            return items

        except json.JSONDecodeError as e:
            raise ValueError(f"解析JSON文件失败: {str(e)}")
        except UnicodeDecodeError:
            raise ValueError("文件编码错误：请确保文件使用UTF-8编码")

    @staticmethod
    def _normalize_cell_newlines(s: str) -> str:
        """将单元格中的字面量 \\n、\\r\\n 转为真实换行。"""
        if not s:
            return s
        return s.replace("\\r\\n", "\n").replace("\\n", "\n")

    @staticmethod
    def _prompt_response_field_value(v: Any) -> str:
        """CSV/Excel 单元格值转字符串，并把字面量 \\n、\\r\\n 转为真实换行。"""
        if v is None:
            return ""
        try:
            if pd.isna(v):
                return ""
        except (TypeError, ValueError):
            pass
        s = str(v).strip()
        return TextGenerationInferenceResultFileParser._normalize_cell_newlines(s)

    @staticmethod
    def _parse_role_based_xlsx(content: bytes) -> List[Dict[str, Any]]:
        """
        解析role-based格式的XLSX文件（推理结果集版本）
        XLSX格式说明：
        - 单轮数据集：表头为 system、user、assistant，每一行对应一个对话样本
        - 多轮数据集：表头为 system、user1、assistant1、user2、assistant2...，每一行对应一个多轮对话样本
        - 保留所有列的数据（包括推理结果集特有字段：response、system、prompt、model_response等）

        Args:
            content: XLSX文件的字节内容

        Returns:
            解析后的数据项列表
        """
        try:
            # 使用只读模式加载工作簿
            file_stream = io.BytesIO(content)
            workbook = load_workbook(
                filename=file_stream,
                read_only=True,
                data_only=True
            )
            worksheet = workbook.active

            if worksheet is None:
                raise ValueError("工作簿中没有找到活动工作表")

            if worksheet.max_row < 2:
                raise ValueError("Excel文件至少需要包含标题行和一行数据")

            # 1. 读取表头
            headers = []
            header_row = next(worksheet.rows)
            for cell in header_row:
                header_value = str(cell.value).strip() if cell.value else ""
                headers.append(header_value)

            # 2. 判断是单轮还是多轮格式，并找到对应的列索引
            system_col = None
            user_cols = []  # 存储所有user列的索引
            assistant_cols = []  # 存储所有assistant列的索引

            # 查找system列
            for i, header in enumerate(headers):
                header_lower = header.lower()
                if header_lower == 'system':
                    system_col = i
                    break

            # 判断格式：检查是否有user1或user（单轮）
            is_multi_turn = False
            for i, header in enumerate(headers):
                header_lower = header.lower()
                if header_lower.startswith('user') and header_lower != 'user':
                    # 发现user1, user2等，说明是多轮格式
                    is_multi_turn = True
                    break

            # 根据格式提取列索引
            if is_multi_turn:
                # 多轮格式：system、user1、assistant1、user2、assistant2...
                for i, header in enumerate(headers):
                    header_lower = header.lower()
                    if header_lower.startswith('user') and header_lower != 'user':
                        # 提取数字，如user1 -> 1, user2 -> 2
                        try:
                            num = int(header_lower.replace('user', ''))
                            user_cols.append((num, i))
                        except ValueError:
                            continue
                    elif header_lower.startswith('assistant') and header_lower != 'assistant':
                        try:
                            num = int(header_lower.replace('assistant', ''))
                            assistant_cols.append((num, i))
                        except ValueError:
                            continue

                # 按数字排序
                user_cols.sort(key=lambda x: x[0])
                assistant_cols.sort(key=lambda x: x[0])
            else:
                # 单轮格式：system、user、assistant
                for i, header in enumerate(headers):
                    header_lower = header.lower()
                    if header_lower == 'user':
                        user_cols.append((1, i))
                    elif header_lower == 'assistant':
                        assistant_cols.append((1, i))

            # 验证必需的列
            if system_col is None:
                raise ValueError("Excel文件必须包含system列")
            if len(user_cols) == 0:
                raise ValueError("Excel文件必须包含user列（单轮）或user1、user2...列（多轮）")
            if len(assistant_cols) == 0:
                raise ValueError("Excel文件必须包含assistant列（单轮）或assistant1、assistant2...列（多轮）")
            if len(user_cols) != len(assistant_cols):
                raise ValueError("Excel文件中user列和assistant列的数量必须相等")

            # 3. 处理数据行
            items = []

            # 跳过第一行（标题行）
            row_iter = iter(worksheet.rows)
            next(row_iter)

            # 处理每一行数据
            for row_num, row in enumerate(row_iter, start=2):
                try:
                    # 构建messages数组
                    messages = []

                    # 添加system消息（如果存在）
                    if system_col < len(row) and row[system_col].value:
                        system_content = TextGenerationInferenceResultFileParser._normalize_cell_newlines(
                            str(row[system_col].value).strip()
                        )
                        if system_content:
                            messages.append({
                                "role": "system",
                                "content": system_content
                            })

                    # 添加user和assistant消息对
                    for turn_num, (user_num, user_col_idx) in enumerate(user_cols, start=1):
                        # 查找对应的assistant列
                        assistant_col_idx = None
                        for ast_num, ast_idx in assistant_cols:
                            if ast_num == user_num:
                                assistant_col_idx = ast_idx
                                break

                        if assistant_col_idx is None:
                            raise ValueError(f"第{row_num}行：找不到与user{user_num}对应的assistant列")

                        # 获取user内容
                        if user_col_idx >= len(row):
                            continue
                        user_cell = row[user_col_idx]
                        user_content = TextGenerationInferenceResultFileParser._normalize_cell_newlines(
                            str(user_cell.value).strip() if user_cell.value else ""
                        )

                        # 获取assistant内容
                        if assistant_col_idx >= len(row):
                            continue
                        assistant_cell = row[assistant_col_idx]
                        assistant_content = TextGenerationInferenceResultFileParser._normalize_cell_newlines(
                            str(assistant_cell.value).strip() if assistant_cell.value else ""
                        )

                        # 如果user或assistant为空，跳过这一轮
                        if not user_content or not assistant_content:
                            if turn_num == 1:
                                # 第一轮必须有内容
                                raise ValueError(f"第{row_num}行：第一轮对话的user和assistant不能为空")
                            else:
                                # 后续轮次可以为空，表示对话结束
                                break

                        # 添加user消息
                        messages.append({
                            "role": "user",
                            "content": user_content
                        })

                        # 添加assistant消息
                        messages.append({
                            "role": "assistant",
                            "content": assistant_content
                        })

                    # 验证messages数组
                    if len(messages) == 0:
                        raise ValueError(f"第{row_num}行：没有找到有效的对话内容")

                    # 验证role交替规则
                    expected_role = None
                    for msg_idx, message in enumerate(messages):
                        role = message['role']
                        if role != 'system':
                            if expected_role is None:
                                if role != 'user':
                                    raise ValueError(
                                        f"第{row_num}行第一个非system的message必须是user，"
                                        f"但第{msg_idx + 1}个message的role为{role}"
                                    )
                                expected_role = 'assistant'
                            else:
                                if role != expected_role:
                                    raise ValueError(
                                        f"第{row_num}行messages中role必须交替出现（user和assistant），"
                                        f"但第{msg_idx + 1}个message的role为{role}，期望为{expected_role}"
                                    )
                                expected_role = 'assistant' if role == 'user' else 'user'

                    # 构建数据对象，先添加messages
                    data = {
                        "messages": messages
                    }

                    # 保留其他列的数据（用于推理结果集：response、model_response、prompt、system等）
                    # 排除已处理的messages相关列
                    processed_cols = set()
                    if system_col is not None:
                        processed_cols.add(system_col)
                    for _, col_idx in user_cols:
                        processed_cols.add(col_idx)
                    for _, col_idx in assistant_cols:
                        processed_cols.add(col_idx)

                    # 遍历所有表头，保留非messages相关的列
                    for col_idx, header in enumerate(headers):
                        if col_idx not in processed_cols:
                            # 跳过空的表头
                            if not header or not header.strip():
                                continue

                            # 获取该列的值
                            if col_idx < len(row):
                                cell_value = row[col_idx].value
                                if cell_value is not None:
                                    # 转换为字符串并去除首尾空格，字面量 \n 转为真实换行
                                    cell_str = TextGenerationInferenceResultFileParser._normalize_cell_newlines(
                                        str(cell_value).strip()
                                    )
                                    if cell_str:
                                        # 使用原始表头作为键（保留大小写）
                                        data[header] = cell_str

                    # 添加空值校验：校验 messages、prompt、response
                    flag, process_error_msg = TextGenerationInferenceResultFileParser._validate_role_based_item(data, line_num=row_num)
                    if not flag:
                        # 空值校验失败，抛出异常，用于错误记录
                        raise ValueError(process_error_msg)

                    # 验证通过，添加到结果列表
                    items.append(data)

                except Exception as e:
                    # 捕获其他异常
                    raise ValueError(str(e))

            if len(items) == 0:
                raise ValueError("Excel文件中没有找到有效的数据样本")

            return items

        except InvalidFileException:
            raise ValueError("无效的Excel文件格式")
        except Exception as e:
            logger.error(f"解析Excel文件失败: {str(e)}")
            raise ValueError(f"解析Excel文件失败: {str(e)}")


    async def process_file(
        self,
        file_content: bytes,
        file_type: str,
        dataset_format: str
    ) -> List[Dict[str, Any]] | None:
        """
        处理文本生成推理结果集文件
            - prompt-response格式：支持xlsx、csv、json、jsonl
            - alpaca格式：支持xlsx、json、jsonl
            - completion-reward格式：支持xlsx、json、jsonl
            - role-based格式：支持xlsx、json、jsonl

        Args:
            file_content: 文件字节内容
            file_type: 文件类型（jsonl, json, csv, xlsx）
            dataset_format: 数据集格式（role-based, prompt-response）

        Returns:
            解析后的数据项列表
        """
        # 验证数据集格式
        if dataset_format not in {
            DatasetFormat.ROLE_BASED.value,
            DatasetFormat.PROMPT_RESPONSE.value,
            DatasetFormat.ALPACA.value,
            DatasetFormat.COMPLETION_REWARD.value,
        }:
            raise ValueError(f"不支持的推理结果集格式: {dataset_format}")

        # 在解析前先检测文件格式
        detected_format = self._detect_dataset_format(file_content, file_type)

        # 如果检测到的格式与用户选择的格式不一致，直接抛出错误
        if detected_format and detected_format != dataset_format:
            raise ValueError(
                f"当前数据集格式异常：检测到的格式为 {detected_format}，但用户选择的格式为 {dataset_format}。"
            )

        # 使用用户选择的格式解析
        try:
            return self._parse_with_format(file_content, file_type, dataset_format)
        except Exception as e:
            # 如果解析失败，直接抛出解析错误
            raise ValueError(f"解析文件失败: {str(e)}")

    def _parse_with_format(
        self,
        file_content: bytes,
        file_type: str,
        dataset_format: str
    ) -> List[Dict[str, Any]]:
        """
        使用指定格式解析文件

        Args:
            file_content: 文件字节内容
            file_type: 文件类型（jsonl, json, csv, xlsx）
            dataset_format: 数据集格式（role-based, prompt-response）

        Returns:
            解析后的数据项列表
        """
        # 根据文件类型选择解析方法
        if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
            # prompt-response
            if file_type == 'jsonl':
                return self._parse_prompt_response_jsonl(file_content)
            elif file_type == 'json':
                return self._parse_prompt_response_json(file_content)
            elif file_type == 'csv':
                return self._parse_prompt_response_csv(file_content)
            elif file_type in ['xlsx', 'xls']:
                return self._parse_prompt_response_excel(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_type}")

        elif dataset_format == DatasetFormat.ROLE_BASED.value:
            # role-based
            if file_type == 'jsonl':
                return self._parse_role_based_jsonl(file_content)
            elif file_type == 'json':
                return self._parse_role_based_json(file_content)
            elif file_type in ['xlsx', 'xls']:
                return self._parse_role_based_xlsx(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_type}")

        elif dataset_format == DatasetFormat.ALPACA.value:
            if file_type == 'jsonl':
                return self._parse_alpaca_jsonl(file_content)
            elif file_type == 'json':
                return self._parse_alpaca_json(file_content)
            elif file_type in ['xlsx', 'xls']:
                return self._parse_alpaca_excel(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_type}")

        elif dataset_format == DatasetFormat.COMPLETION_REWARD.value:
            if file_type == 'jsonl':
                return self._parse_completion_reward_jsonl(file_content)
            elif file_type == 'json':
                return self._parse_completion_reward_json(file_content)
            elif file_type in ['xlsx', 'xls']:
                return self._parse_completion_reward_excel(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_type}")

        else:
            raise ValueError(f"不支持的数据集格式: {dataset_format}")


class ImageUnderstandingInferenceResultFileParser:
    """图像理解推理结果集文件解析器"""

    def __init__(self):
        """初始化解析器"""
        pass

    @staticmethod
    async def _parse_role_based_zip(content: bytes) -> Tuple[Dict[str, bytes], List[Dict[str, Any]]]:
        # 使用图像理解数据集解析器处理zip文件
        parse_result = await analyze_image_understanding_dataset_file_content(content, 'zip')

        jsonl_content = parse_result.jsonl_content
        images = parse_result.images

        # 解析 jsonl_content 为对象列表
        items = []
        if jsonl_content:

            # 将json内容转化为字符串
            jsonl_str = _parse_json_content_to_str(jsonl_content)

            # 解析json，转化为字典list
            items = _parse_json_str_to_item(jsonl_str)

        else:
            logger.warning("图像理解数据集jsonl内容为空")

        return images, items

    async def process_file(
        self,
        file_content: bytes,
        file_type: str,
        dataset_format: str
    ) -> Tuple[Dict[str, bytes], List[Dict[str, Any]]]:
        """
        处理图像理解推理结果集文件

        Args:
            file_content: 文件字节内容
            file_type: 文件类型（zip）
            dataset_format: 数据集格式（必须是role-based）

        Returns:
            (图片字典, 数据项列表)
        """
        # 验证格式
        if dataset_format != DatasetFormat.ROLE_BASED.value:
            raise ValueError(f"不支持的推理结果集格式: {dataset_format}")

        # 根据文件类型选择解析方法
        if dataset_format == DatasetFormat.ROLE_BASED.value:
            # role-based
            if file_type == InferenceResultDatasetUploadType.ZIP_TYPE.value:
                images, items = await self._parse_role_based_zip(file_content)
            else:
                raise ValueError(f"不支持的文件格式: {file_type}")

        else:
            raise ValueError(f"不支持的数据集格式: {dataset_format}")

        return images, items


class BusinessInferenceResultFileParser:
    """业务推理结果集文件解析器"""

    def __init__(self):
        """初始化解析器"""
        pass

    @staticmethod
    def _parse_jsonl(content: bytes) -> List[Dict[str, Any]]:
        """解析业务数据集JSONL文件（宽松规则）"""
        items = []
        content_str = content.decode('utf-8').strip()

        for line_num, line in enumerate(content_str.split('\n'), start=1):
            line = line.strip()
            if line:
                try:
                    item = json.loads(line)
                    if isinstance(item, dict):
                        items.append(item)
                except json.JSONDecodeError as e:
                    logger.warning(f"第 {line_num} 行跳过无效的JSON: {line[:100]}, 错误: {str(e)}")

        return items

    @staticmethod
    def _parse_json(content: bytes) -> List[Dict[str, Any]]:
        """解析业务数据集JSON文件（宽松规则）"""
        items = []
        content_str = content.decode('utf-8').strip()

        try:
            parsed_data = json.loads(content_str)

            if isinstance(parsed_data, list):
                for item in parsed_data:
                    if isinstance(item, dict):
                        items.append(item)
                return items
            elif isinstance(parsed_data, dict):
                items.append(parsed_data)
                return items
            else:
                raise ValueError(f"JSON格式不是数组或对象，类型: {type(parsed_data).__name__}，业务数据集需要字典对象")
        except json.JSONDecodeError as e:
            raise ValueError(f"解析JSON文件失败: {str(e)}")

    @staticmethod
    def _parse_csv(content: bytes) -> List[Dict[str, Any]]:
        """解析业务数据集CSV文件（宽松规则）"""
        try:
            encodings_to_try = ['gbk', 'gb18030', 'utf-8', 'latin-1', 'cp1252']
            df = None

            for encoding in encodings_to_try:
                try:
                    df = pd.read_csv(BytesIO(content), encoding=encoding)
                    break
                except Exception as e:
                    logger.warning(f"尝试以{encoding}编码解析csv失败！- {str(e)}")
                    continue
            else:
                raise ValueError(f"不支持的文件编码，目前支持: {', '.join(encodings_to_try)}")

            items = []
            for _, row in df.iterrows():
                item = row.to_dict()
                item = {k: (None if pd.isna(v) else v) for k, v in item.items()}
                items.append(item)
            return items
        except Exception as e:
            raise ValueError(f"解析业务数据集CSV文件失败: {str(e)}")

    @staticmethod
    def _parse_excel(content: bytes) -> List[Dict[str, Any]]:
        """解析业务数据集Excel文件（宽松规则）"""
        try:
            df = pd.read_excel(BytesIO(content))
            items = []
            for _, row in df.iterrows():
                item = row.to_dict()
                item = {k: (None if pd.isna(v) else v) for k, v in item.items()}
                items.append(item)
            return items
        except Exception as e:
            raise ValueError(f"解析业务数据集Excel文件失败: {str(e)}")

    async def process_file(
        self,
        file_content: bytes,
        file_type: str,
        dataset_format: str
    ) -> List[Dict[str, Any]]:
        """
        处理业务推理结果集文件

        Args:
            file_content: 文件字节内容
            file_type: 文件类型（jsonl, json, csv, xlsx）
            dataset_format: 数据集格式（必须是business）

        Returns:
            解析后的数据项列表
        """
        # 验证数据集格式
        if dataset_format != DatasetFormat.BUSINESS.value:
            raise ValueError(f"不支持的推理结果集格式: {dataset_format}")

        # 根据文件类型选择解析方法
        if file_type == 'jsonl':
            return self._parse_jsonl(file_content)
        elif file_type == 'json':
            return self._parse_json(file_content)
        elif file_type == 'csv':
            return self._parse_csv(file_content)
        elif file_type in ['xlsx', 'xls']:
            return self._parse_excel(file_content)
        else:
            raise ValueError(f"不支持的文件格式: {file_type}")

# ========== 辅助函数 ==========
def _parse_json_content_to_str(content: bytes | str) -> str:
    """
    将json内容统一转化为字符串，无论传入的json内容为bytes或str

    Args:
        content: json内容（bytes或str）

    Returns:
        转化后的json字符串
    """
    return content.decode('utf-8') if isinstance(content, bytes) else content

def _parse_json_str_to_item(content: str) -> List[Dict[str, Any]]:
    """
    将json字符串内容统一转化列表

    Args:
        content: json字符串
    Returns:
        转化后的列表
    """
    items = []
    for line_num, line in enumerate(content.strip().split('\n'), start=1):
        line = line.strip()
        if line:
            try:
                item = json.loads(line)
                if isinstance(item, dict):
                    items.append(item)
                else:
                    logger.warning(f"第 {line_num} 行跳过非字典类型的JSON: {type(item).__name__}")
            except json.JSONDecodeError as e:
                logger.warning(f"第 {line_num} 行跳过无效的JSON: {line[:100]}, 错误: {str(e)}")

    return items

async def _read_jsonl_file_content(file_path: str, storage_service: StorageService) -> list[Any]:
    """
    读取jsonl文件中的内容，转化为list，方便后续解析

    Args:
        file_path: 数据集在jfs中的路径
        storage_service: 存储服务
    """
    # 验证数据集文件路径
    if not file_path:
        raise ValueError("数据集文件路径为空")

    # 获取 JuiceFS 客户端
    jfs = await storage_service.JUICEFS_CLIENT()

    # 验证文件是否存在
    if not jfs.exists(file_path):
        raise ValueError(f"数据集文件不存在: {file_path}")

    # 读取 JSONL 文件内容
    all_items = []
    with jfs.open(file_path, 'r', encoding='utf-8') as f:
        for line_num, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue

            # 跳过注释行
            if line.lstrip().startswith('#'):
                continue

            try:
                item_data = json.loads(line)
                if isinstance(item_data, dict):
                    all_items.append(item_data)
            except json.JSONDecodeError as e:
                logger.warning(f"跳过无效的JSON行 {line_num} in {file_path}: {e}")
                continue

    if len(all_items) == 0:
        logger.warning(f"数据集中没有有效的数据样本: {file_path}")
        # 不再抛出异常，返回空列表，由调用方处理

    return all_items

# ========== 入口函数 ==========

async def analyze_save_inference_result_files(
    upload_files: List[UploadFile],
    file_path: str,
    dataset_format: Optional[str],
    usage: Optional[str],
    dataset_type: Optional[str],
    storage_service: StorageService,
    task: Optional[Any] = None
) -> int:
    """
    处理上传的推理结果集文件，解析并保存到 JuiceFS

    Args:
        upload_files: 上传的文件列表
        file_path: 文件保存路径
        dataset_format: 数据格式
        usage: 数据集用途
        dataset_type: 数据集类型
        storage_service: 存储服务
        task: TaskBase 实例（用于日志记录，可选）

    Returns:
        int: 总数据量
    """
    total_items = 0

    # 获取JuiceFS客户端
    jfs = await storage_service.JUICEFS_CLIENT()

    # 添加调试日志
    if task:
        task._log_info(f"开始处理导入数据集文件: dataset_type={dataset_type}, dataset_format={dataset_format}")

    # 根据数据格式确定实际保存路径
    if dataset_type == TrainingTypeCategory.IMAGE_UNDERSTANDING.value:
        # 图像理解格式：file_path 是文件路径（xxx/inference_result_{name}_{timestamp}.jsonl）
        actual_file_path = file_path
        file_dir = os.path.dirname(file_path)
        images_dir = os.path.join(file_dir, "images").replace('\\', '/')
        remote_dir = file_dir
    elif dataset_type == TrainingTypeCategory.TEXT_GENERATION.value:
        # 文本生成格式：file_path 是文件路径
        actual_file_path = file_path
        images_dir = None
        remote_dir = os.path.dirname(file_path)
    elif dataset_type == TrainingTypeCategory.BUSINESS.value:
        # 业务推理结果集格式: file_path 是文件路径
        actual_file_path = file_path
        images_dir = None
        remote_dir = os.path.dirname(file_path)
    else:
        raise ValueError(f"当前推理结果集类型不支持:{dataset_type}")

    # 确保目录存在
    if remote_dir and not jfs.exists(remote_dir):
        jfs.makedirs(remote_dir, exist_ok=True)

    # 处理每个文件，合并所有数据
    all_items_data = []

    # 记录文件索引
    for index, file in enumerate(upload_files):
        try:
            # 获取文件类型
            file_extension = os.path.splitext(file.filename)[1].lower()
            file_type = file_extension.lstrip('.')  # 去掉点号，如 '.jsonl' -> 'jsonl'

            # 验证文件格式
            validate_dataset_upload_file_type(file_type, dataset_type)

            # 图像理解数据集处理逻辑
            if dataset_type == TrainingTypeCategory.IMAGE_UNDERSTANDING.value:
                parser = ImageUnderstandingInferenceResultFileParser()
                zip_content = await file.read()
                images, items = await parser.process_file(zip_content, file_type, dataset_format)

                if task:
                    task._log_info(
                        f"图像理解数据集zip文件处理成功: 样本数={len(items)}, "
                        f"图片数={len(images)}"
                    )

                # 确保images目录存在
                if not jfs.exists(images_dir):
                    jfs.makedirs(images_dir, exist_ok=True)

                # 保存图片文件
                for image_name, image_content in images.items():
                    image_path = os.path.join(images_dir, image_name).replace('\\', '/')
                    with jfs.open(image_path, 'wb') as f:
                        f.write(image_content)
                if task:
                    task._log_info(f"保存图片文件到 JuiceFS: {len(images)} 张图片保存到 {images_dir}")

                all_items_data.extend(items)
                total_items += len(items)

            # 文本生成/业务数据集处理逻辑
            elif dataset_type == TrainingTypeCategory.TEXT_GENERATION.value or dataset_type == TrainingTypeCategory.BUSINESS.value:

                # 数据集文件类型验证
                if dataset_type == TrainingTypeCategory.TEXT_GENERATION.value and file_extension not in ['.jsonl', '.json', '.xlsx']:
                    raise ValueError("文本生成类型结果集仅支持jsonl、json、xlsx格式的数据集文件导入")

                if dataset_type == TrainingTypeCategory.BUSINESS.value and file_extension not in ['.jsonl', '.json', '.xlsx', '.csv']:
                    raise ValueError("业务类型结果集仅支持jsonl、json、xlsx、csv格式的数据集文件导入")

                # 读取文件内容
                content = await file.read()

                # 根据数据集类型选择解析器
                if dataset_type == TrainingTypeCategory.TEXT_GENERATION.value:
                    # 文本生成类型
                    parser = TextGenerationInferenceResultFileParser()
                    # 处理文件
                    items = await parser.process_file(content, file_type, dataset_format)
                elif dataset_type == TrainingTypeCategory.BUSINESS.value:
                    # 业务类型
                    parser = BusinessInferenceResultFileParser()
                    # 处理文件
                    items = await parser.process_file(content, file_type, dataset_format)
                else:
                    raise HTTPException(status_code=400, detail=f"不支持的结果集类型{dataset_type}")

                # 添加到总列表
                all_items_data.extend(items)
                total_items += len(items)
            else:
                raise ValueError(f"当前推理结果集类型不支持:{dataset_type}")
        except Exception as e:
            # 添加文件索引前缀
            raise ValueError(f"第 {index + 1} 文件解析异常：{str(e)}")

    # 保存合并后的文件到JuiceFS（JSONL格式，每行一个JSON对象）
    if all_items_data:
        jsonl_lines = [json.dumps(item, ensure_ascii=False) for item in all_items_data]
        jsonl_content = "\n".join(jsonl_lines)
        with jfs.open(actual_file_path, 'w', encoding='utf-8') as f:
            f.write(jsonl_content)
        if task:
            task._log_info(f"保存推理结果文件到 JuiceFS: {actual_file_path}, 共 {total_items} 条数据")
    else:
        raise ValueError("上传的文件中没有有效数据")

    return total_items


# =========== 数据集导出入口函数 =============
async def analyze_export_inference_result_file_single(
        db_dataset: InferenceResultDataset,
        export_file_type: InferenceResultDatasetExportType,
        storage_service: StorageService,
) -> bytes:
    """
    推理结果集多格式导出
        - 支持 文本生成role-based
        - 支持 文本生成prompt-response
        - 支持 xlsx、json、jsonl

    Args:
        db_dataset: 结果集信息
        export_file_type: 需要导出的格式
        storage_service: 存储服务实例

    Returns:
        转换后的文件内容（bytes）

    转换规则：
    1. JSONL → JSON: 将每行 JSON 对象合并为 JSON 数组
    2. JSONL → XLSX:
       - prompt-response: 列结构为 prompt, response, system（可选），以及其他字段（如 model_response）
       - role-based: 从 messages 数组转换为表格列（单轮：system, user, assistant；多轮：system, user1, assistant1, user2, assistant2...）
    """
    # 先读取数据集文件内容
    if not db_dataset:
        raise ValueError("db_database不能为空")

    all_items = await _read_jsonl_file_content(db_dataset.file_path, storage_service)

    # 处理空数据情况：记录日志并返回对应的空文件
    if not all_items:
        logger.info(f"推理结果集为空，返回空文件: {db_dataset.file_path}")
        export_type = export_file_type.value

        if export_type == InferenceResultDatasetExportType.JSONL_TYPE.value:
            # JSONL 格式：返回空字符串
            return b''

        elif export_type == InferenceResultDatasetExportType.JSON_TYPE.value:
            # JSON 格式：返回空数组
            return json.dumps([], ensure_ascii=False, indent=2).encode('utf-8')

        elif export_type == InferenceResultDatasetExportType.XLSX_TYPE.value:
            # XLSX 格式：返回只有表头的 Excel 文件
            dataset_format = db_dataset.dataset_format
            if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
                return _convert_prompt_response_to_xlsx([])
            elif dataset_format == DatasetFormat.ROLE_BASED.value:
                return _convert_role_based_to_xlsx([])
            else:
                # 未知格式，返回空 Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "推理结果集"
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                return output.getvalue()

        else:
            raise ValueError(f"不支持的导出格式: {export_type}")

    # 根据导出类型进行转换
    export_type = export_file_type.value

    if export_type == InferenceResultDatasetExportType.JSONL_TYPE.value:
        # JSONL 格式：直接返回原始 JSONL 内容
        jsonl_lines = [json.dumps(item, ensure_ascii=False) for item in all_items]
        jsonl_content = "\n".join(jsonl_lines)
        return jsonl_content.encode('utf-8')

    elif export_type == InferenceResultDatasetExportType.JSON_TYPE.value:
        # JSON 格式：将每行 JSON 对象合并为 JSON 数组
        json_content = json.dumps(all_items, ensure_ascii=False, indent=2)
        return json_content.encode('utf-8')

    elif export_type == InferenceResultDatasetExportType.XLSX_TYPE.value:
        # XLSX 格式：根据数据集格式进行转换
        dataset_format = db_dataset.dataset_format

        if dataset_format == DatasetFormat.PROMPT_RESPONSE.value:
            # prompt-response 格式转换为 XLSX
            return _convert_prompt_response_to_xlsx(all_items)

        elif dataset_format == DatasetFormat.ROLE_BASED.value:
            # role-based 格式转换为 XLSX
            return _convert_role_based_to_xlsx(all_items)

        elif dataset_format == DatasetFormat.BUSINESS.value:
            # business 格式转化为 XLSX
            return _convert_business_to_xlsx(all_items)

        else:
            raise ValueError(f"不支持的数据集格式: {dataset_format}，XLSX 导出仅支持 prompt-response 和 role-based 格式")

    else:
        raise ValueError(f"不支持的导出格式: {export_type}")


def _convert_prompt_response_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """
    将 prompt-response 格式的数据转换为 XLSX

    列结构：
    - prompt（必需）
    - response（必需）
    - system（可选）
    - model_response（可选，推理结果集特有）
    - 其他字段（如 error, error_message 等）

    Args:
        items: 数据项列表

    Returns:
        XLSX 文件的字节内容
    """
    # 收集所有可能的列名
    all_columns = set()
    for item in items:
        all_columns.update(item.keys())

    # 定义列的顺序（优先显示常用列）
    priority_columns = ['prompt', 'response', 'system', 'model_response', 'error', 'error_message']
    ordered_columns = []

    # 先添加优先列（如果数据为空，至少添加默认列）
    if not items:
        # 空数据时，使用默认列
        ordered_columns = priority_columns
    else:
        # 有数据时，按实际列添加
        for col in priority_columns:
            if col in all_columns:
                ordered_columns.append(col)
                all_columns.discard(col)

        # 再添加其他列（按字母顺序）
        ordered_columns.extend(sorted(all_columns))

    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "推理结果集"

    # 写入表头
    ws.append(ordered_columns)

    # 写入数据行
    for item in items:
        row = [item.get(col, '') for col in ordered_columns]
        # 处理 None 值
        row = ['' if v is None else v for v in row]
        ws.append(row)

    # 将工作簿保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def _convert_role_based_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """
    将 role-based 格式的数据转换为 XLSX

    转换规则（反向执行 xlsx 解析逻辑）：
    - 单轮格式：system, user, assistant
    - 多轮格式：system, user1, assistant1, user2, assistant2...
    - 保留其他字段（如 response, system, prompt, model_response 等）

    Args:
        items: 数据项列表（每个项包含 messages 数组和其他字段）

    Returns:
        XLSX 文件的字节内容
    """
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "推理结果集"

    # 分析所有数据项，确定是单轮还是多轮格式
    max_turns = 0
    all_other_fields = set()

    for item in items:
        if 'messages' not in item:
            continue

        messages = item['messages']
        if not isinstance(messages, list):
            continue

        # 统计 user-assistant 轮数（排除 system）
        user_assistant_pairs = 0
        for msg in messages:
            if msg.get('role') == 'user':
                user_assistant_pairs += 1

        max_turns = max(max_turns, user_assistant_pairs)

        # 收集其他字段（非 messages 字段）
        for key in item.keys():
            if key != 'messages':
                all_other_fields.add(key)

    # 构建表头
    headers = []

    # 添加 system 列
    headers.append('system')

    # 添加 user 和 assistant 列
    if max_turns == 1:
        # 单轮格式
        headers.append('user')
        headers.append('assistant')
    else:
        # 多轮格式
        for turn_num in range(1, max_turns + 1):
            headers.append(f'user{turn_num}')
            headers.append(f'assistant{turn_num}')

    # 添加其他字段列（按字母顺序）
    headers.extend(sorted(all_other_fields))

    # 写入表头
    ws.append(headers)

    # 写入数据行
    for item in items:
        row_data = {}

        # 提取 messages 内容
        if 'messages' in item and isinstance(item['messages'], list):
            messages = item['messages']

            # 提取 system 消息
            system_content = ''
            user_assistant_pairs = []

            for msg in messages:
                role = msg.get('role', '')
                content = msg.get('content', '')

                if role == 'system':
                    system_content = content
                elif role == 'user':
                    user_assistant_pairs.append({'user': content, 'assistant': ''})
                elif role == 'assistant' and user_assistant_pairs:
                    user_assistant_pairs[-1]['assistant'] = content

            # 填充 system 列
            row_data['system'] = system_content

            # 填充 user 和 assistant 列
            if max_turns == 1:
                # 单轮格式
                if len(user_assistant_pairs) > 0:
                    row_data['user'] = user_assistant_pairs[0].get('user', '')
                    row_data['assistant'] = user_assistant_pairs[0].get('assistant', '')
                else:
                    row_data['user'] = ''
                    row_data['assistant'] = ''
            else:
                # 多轮格式
                for turn_num in range(1, max_turns + 1):
                    if turn_num <= len(user_assistant_pairs):
                        pair = user_assistant_pairs[turn_num - 1]
                        row_data[f'user{turn_num}'] = pair.get('user', '')
                        row_data[f'assistant{turn_num}'] = pair.get('assistant', '')
                    else:
                        row_data[f'user{turn_num}'] = ''
                        row_data[f'assistant{turn_num}'] = ''

        # 填充其他字段
        for field in all_other_fields:
            value = item.get(field, '')
            # 处理 None 值
            if value is None:
                value = ''
            row_data[field] = value

        # 构建行数据（按照表头顺序）
        row = [row_data.get(header, '') for header in headers]
        ws.append(row)

    # 将工作簿保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def _convert_value_to_excel(value: Any) -> str:
    """
    将值转换为 Excel 可接受的格式

    Args:
        value: 要转换的值

    Returns:
        转换后的字符串值
    """
    # None -> 空字符串
    if value is None:
        return ''

    # 基本类型（str, int, float, bool）-> 转换为字符串
    if isinstance(value, (str, int, float, bool)):
        return str(value)

    # 复杂类型（list, dict）-> 转换为 JSON 字符串
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)

    # 其他类型 -> 转换为字符串
    return str(value)


def _convert_business_to_xlsx(items: List[Dict[str, Any]]) -> bytes:
    """
    将业务推理结果集转换为 XLSX 格式

    转换规则：
    - 每一行 jsonl 对象就为一个 json 对象
    - 默认第一个对象的属性为表头
    - 后面的 json 对象都以当前表头为准（缺少的字段填充空值）

    Args:
        items: 数据项列表

    Returns:
        XLSX 文件的字节内容
    """
    # 创建 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "业务推理结果集"

    # 如果数据为空，返回只有表头的空 Excel
    if not items:
        # 空数据时，返回只有表头的空 Excel（表头为空）
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 使用第一个对象的键作为表头
    headers = list(items[0].keys())

    # 写入表头
    ws.append(headers)

    # 写入数据行
    for item in items:
        # 按照表头顺序获取值，如果某个字段不存在则填充空字符串
        row = [item.get(col, '') for col in headers]
        # 使用辅助函数处理每个值，转换为 Excel 可接受的格式
        row = [_convert_value_to_excel(v) for v in row]
        ws.append(row)

    # 将工作簿保存到 BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
