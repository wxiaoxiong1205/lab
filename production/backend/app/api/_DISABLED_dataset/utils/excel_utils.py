import json
import io
from typing import Dict, List, Any, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.exceptions import InvalidFileException
from fastapi import HTTPException, UploadFile

def parse_meta_info(meta_info_str: str) -> Dict:
    """Parse meta info string to dictionary
    
    Args:
        meta_info_str: Meta info string in JSON format
        
    Returns:
        Dict: Parsed meta info
    """
    if not meta_info_str:
        return {}
        
    try:
        return json.loads(meta_info_str)
    except json.JSONDecodeError:
        # If not valid JSON, return as-is in a dict
        return {"raw": meta_info_str}

async def parse_excel_file(file: UploadFile) -> List[Dict[str, Any]]:
    """Parse Excel file and extract dataset information using pandas for improved performance
    
    Args:
        file: Uploaded Excel file
        
    Returns:
        List[Dict[str, Any]]: List of dataset records
        
    Raises:
        HTTPException: If file format is invalid
    """
    try:
        # Read the file using pandas for better performance
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        # Process dataframe
        all_rows = []
        # Replace NaN values with empty strings
        df = df.fillna("")
        # Convert DataFrame to list of dictionaries
        for _, row in df.iterrows():
            row_data = row.to_dict()
            # Skip empty rows
            if not any(row_data.values()):
                continue
            # 确保文本字段为字符串类型
            row_data["question"] = str(row_data.get("question", ""))
            row_data["ground_truth"] = str(row_data.get("ground_truth", ""))
            # 处理数组类型字段
            for field in ["context", "retrieval_context", "tools", "expected_tools"]:
                if field in row_data and row_data[field]:
                    field_value = row_data[field]
                    # 尝试解析 JSON 字符串
                    if isinstance(field_value, str):
                        try:
                            # 检查是否为 JSON 格式
                            if (field_value.startswith('[') and field_value.endswith(']')):
                                row_data[field] = json.loads(field_value)
                        except json.JSONDecodeError:
                            # 如果 JSON 解析失败，将其作为单个元素数组处理
                            row_data[field] = [str(field_value)]
                else:
                    row_data[field] = []
                    
            # 验证question字段不能为空
            if not row_data.get("question"):
                continue
                     
            all_rows.append(row_data)
            
        return all_rows
        
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid Excel file format: {str(e)}"
        )

def create_excel_template() -> io.BytesIO:
    """Create an Excel template for dataset import
    
    Returns:
        io.BytesIO: Excel file as bytes
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Datasets"
    
    # Define headers
    headers = ["question（必填）", "ground_truth", "retrieval_context", "context", "tools", "expected_tools"]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Apply headers and styles
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Set column widths
    column_widths = {
        "A": 30,  # question
        "B": 30,  # ground_truth
        "C": 40,  # retrieval_context
        "D": 40,  # context
        "E": 30,  # tools
        "F": 30,  # expected_tools
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
        
    # Add example row
    example_data = [
        "What is the capital of France?",
        "Paris",
        json.dumps(["France is a country in Western Europe.", "Paris is a major city in France."]),
        json.dumps(["France is a country in Western Europe.", "It is located in the western part of continental Europe."]),
        json.dumps([{"type":"function","function":{"name":"MetricDataQuery","description":"查询指标和维度等信息","parameters":{"type":"object","properties":{"question":{"description":"原文","type":"string"},"metric":{"description":"与当前问题最相关的指标名称。指标应为字符串列表，必须包括与问题最相关的三个指标名称,其中有必须一个指标在输入原文信息中","type":"array","items":{"type":"string"}},"dimension":{"description":"维度列表","type":"array","items":{"type":"string"}},"date_text":{"description":"抽取用户问题中关于日、月、周、季度等描述时间的实体原文","type":"string"}},"required":["question","metric","dimension","date_text"],"example1":{"question":"今天的销售量是多少？","metric":["销售量","销售额","均折扣"],"dimension":[],"date_text":"今天"},"example2":{"question":"周销售量","metric":["销售量"],"dimension":[],"date_text":"周"},"example3":{"question":"本月女鞋销售量","metric":["销售量"],"dimension":["女","鞋"],"date_text":"本月"}}}},{"type":"function","function":{"name":"ChitChatQuery","description":"其它无关查询","parameters":{"type":"object","properties":{"date_text":{"description":"抽取用户问题中关于日、月、周、季度等描述时间的实体原文","type":"string"},"dimension":{"description":"'当前问题'对应候选维度值，没有则返回空","type":"array","items":{"type":"string"}}},"required":["date_text","dimension"],"example1":{"date_text":"今年","dimension":["张三"]}}}},{"type":"function","function":{"name":"MetricInfoQuery","description":"\"支持围绕具体查询详情, 衍生指标, 基础指标, 维度相关性,指标相关性等信息","parameters":{"type":"object","properties":{"question":{"description":"原文","type":"string"},"metric":{"description":"抽取原文中指标名，字符与原文严格保持一致","type":"array","items":{"type":"string"}},"metric_or_dimension_type":{"description":"候选指标类型中一种","type":"string"},"is_hole_info":{"description":"判定用户意图是否查询全局信息，返回是、否其中一种","type":"string"}},"required":["question","metric","metric_or_dimension_type","is_hole_info"],"example1":{"question":"销售额的详情","metric":["销售量"],"metric_or_dimension_type":"详情","is_hole_info":"否"},"example2":{"question":"系统支持哪些维度","metric":[],"metric_or_dimension_type":"维度","is_hole_info":"是"},"example3":{"question":"系统支持哪些指标","metric":[],"metric_or_dimension_type":"指标","is_hole_info":"是"},"example4":{"question":"销售量的可用维度","metric":["销售量"],"metric_or_dimension_type":"维度","is_hole_info":"否"}}}},{"type":"function","function":{"name":"EntityInfoQuery","description":"支持通过指标目录、指标来源、指标描述等信息查询所属指标名","parameters":{"type":"object","properties":{"question":{"description":"原文","type":"string"},"entity":{"description":"原文中实体信息，与候选实体值一致","type":"string"},"description_text":{"description":"实体类型为指标描述时，抽取当前问题中指标描述完整定义信息。请确保不提取'的指标'","type":"string"},"entity_type":{"description":"当前问题涉及的entity_type","type":"string"}},"required":["question","entity","description_text","entity_type"],"example1":{"question":"财务相关的指标有哪些","entity":"财务","extract_description_text":"","entity_type":"分类目录"},"example2":{"question":"指标描述是今天卖爆了么是哪个指标","entity":"","description_text":"今天卖爆了么","entity_type":"指标描述"}}}},{"type":"function","function":{"name":"AttributionAnalysis","description":"支持对指标、维度波动原因分析","parameters":{"type":"object","properties":{"question":{"description":"原文","type":"string"},"start_time":{"description":"查询开始日期，默认本年","type":"string"},"end_time":{"description":"查询结束日期，默认本年","type":"string"},"metric":{"description":"指标名","type":"string"},"dimension":{"description":"维度名","type":"array","items":{"type":"string"}}},"required":["question","start_time","end_time","metric","dimension"],"example1":{"question":"从3月1日到5月2日，查看店铺对销售量的变化影响","start_time":"2024-03-01","end_time":"2024-05-02","metric":"销售量","dimension":["店铺"]}}}}]),
        json.dumps([{"name":"MetricDataQuery","description":None,"reasoning":None,"output":None,"input_parameters":{"question":"昨天店员销售额排名","metric":["销售额","销售量","均折扣"],"dimension":["店员"],"date_text":"昨天"}}])
    ]
    
    for col_idx, value in enumerate(example_data, 1):
        cell = sheet.cell(row=2, column=col_idx, value=value)
        cell.border = thin_border
        
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output

def create_export_workbook(datasets_with_tags: List[Dict]) -> io.BytesIO:
    """Create an Excel workbook for dataset export
    
    Args:
        datasets_with_tags: List of datasets with their tags
        
    Returns:
        io.BytesIO: Excel file as bytes
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exported Datasets"
    
    # Define headers - only include required fields
    headers = ["question", "ground_truth", "retrieval_context", "context", "tools", "expected_tools"]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Apply headers and styles
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Set column widths - only include required fields
    column_widths = {
        "A": 30,  # question
        "B": 30,  # ground_truth
        "C": 40,  # retrieval_context
        "D": 40,  # context
        "E": 30,  # tools
        "F": 30,  # expected_tools
    }
    
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
        
    # Add data rows
    for row_idx, dataset_with_tags in enumerate(datasets_with_tags, 2):
        dataset = dataset_with_tags["dataset"]
        
        # 处理数组类型字段
        context_value = json.dumps(dataset.context) if dataset.context else ""
        retrieval_context_value = json.dumps(dataset.retrieval_context) if dataset.retrieval_context else ""
        tools_value = json.dumps(dataset.tools) if dataset.tools else ""
        expected_tools_value = json.dumps(dataset.expected_tools) if dataset.expected_tools else ""
        
        # Row data - only include required fields
        row_data = [
            dataset.question,
            dataset.ground_truth,
            retrieval_context_value,
            context_value,
            tools_value,
            expected_tools_value
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output 